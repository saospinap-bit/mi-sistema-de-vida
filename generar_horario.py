#!/usr/bin/env python3
"""
Generador del Sistema de Vida — Horario semanal.
Crea:  Horario.xlsx  ·  Horario.pdf  ·  Horario.ics
Zona horaria: America/Bogota (UTC-5)
Ejecuta:  python generar_horario.py
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ----------------------------------------------------------------------------
# 1) DATOS DEL HORARIO  (categoria define el color)
# ----------------------------------------------------------------------------
DIAS = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

# Categorias y su color (hex sin #)
CAT_COLOR = {
    "ingles":      "B3E5FC",  # azul claro
    "clase":       "FFCDD2",  # rojo claro (obligatorio)
    "estructural": "F44336",  # rojo fuerte (NO FALTAR)
    "proyecto":    "FFE0B2",  # naranja (proyecto)
    "novia":       "F8BBD0",  # rosa
    "gym":         "C8E6C9",  # verde
    "box":         "A5D6A7",  # verde
    "psicologia":  "E1BEE7",  # morado
    "lectura":     "FFF9C4",  # amarillo
    "viaje":       "ECEFF1",  # gris
    "descanso":    "F5F5F5",  # gris muy claro
    "libre":       "FFFFFF",
}

# Cada evento: (dia, inicio "HH:MM", fin "HH:MM", titulo, categoria)
EVENTOS = [
    # --- LUNES (dia de casa) ---
    ("Lunes", "06:00", "07:30", "Ingles (virtual)", "ingles"),
    ("Lunes", "08:00", "09:30", "Gimnasio (fuerza)", "gym"),
    ("Lunes", "10:00", "12:30", "PROYECTO estructural", "proyecto"),
    ("Lunes", "13:00", "13:30", "Lectura 30 min", "lectura"),
    ("Lunes", "14:00", "16:30", "PROYECTO / estudio", "proyecto"),
    ("Lunes", "16:30", "17:30", "Psicologia", "psicologia"),
    ("Lunes", "18:00", "20:00", "Clase opcional (evaluar viaje)", "clase"),
    # --- MARTES ---
    ("Martes", "06:00", "07:30", "Ingles (virtual)", "ingles"),
    ("Martes", "08:00", "09:00", "Viaje a la U", "viaje"),
    ("Martes", "09:00", "11:00", "Clase OBLIGATORIA", "clase"),
    ("Martes", "11:00", "12:30", "Estudio / pasantia", "proyecto"),
    ("Martes", "12:30", "13:30", "Almuerzo con novia", "novia"),
    ("Martes", "13:30", "16:00", "PROYECTO estructural", "proyecto"),
    ("Martes", "17:30", "18:00", "Recoger a novia", "novia"),
    ("Martes", "19:30", "21:00", "Box", "box"),
    # --- MIERCOLES (estructural) ---
    ("Miercoles", "06:00", "07:30", "Ingles (virtual)", "ingles"),
    ("Miercoles", "08:00", "09:00", "Viaje a la U", "viaje"),
    ("Miercoles", "09:00", "11:00", "Clase OBLIGATORIA", "clase"),
    ("Miercoles", "11:00", "12:30", "Repaso / tareas", "proyecto"),
    ("Miercoles", "12:30", "13:30", "Almuerzo con novia", "novia"),
    ("Miercoles", "14:00", "16:00", "CLASE ESTRUCTURAL - NO FALTAR", "estructural"),
    ("Miercoles", "16:00", "17:30", "Estudio / dudas profe", "proyecto"),
    ("Miercoles", "17:30", "18:00", "Recoger a novia", "novia"),
    # --- JUEVES ---
    ("Jueves", "06:00", "07:30", "Ingles (virtual)", "ingles"),
    ("Jueves", "08:00", "09:00", "Viaje a la U", "viaje"),
    ("Jueves", "09:00", "11:00", "Clase OBLIGATORIA", "clase"),
    ("Jueves", "11:00", "12:30", "Estudio / pasantia", "proyecto"),
    ("Jueves", "12:30", "13:30", "Almuerzo con novia", "novia"),
    ("Jueves", "13:30", "16:00", "PROYECTO estructural", "proyecto"),
    ("Jueves", "17:30", "18:00", "Recoger a novia", "novia"),
    ("Jueves", "19:30", "21:00", "Box", "box"),
    # --- VIERNES (estructural) ---
    ("Viernes", "06:00", "07:30", "Ingles (virtual)", "ingles"),
    ("Viernes", "08:00", "09:00", "Viaje a la U", "viaje"),
    ("Viernes", "09:00", "11:00", "Clase OBLIGATORIA", "clase"),
    ("Viernes", "12:30", "13:30", "Almuerzo con novia", "novia"),
    ("Viernes", "14:00", "16:00", "CLASE ESTRUCTURAL - NO FALTAR", "estructural"),
    ("Viernes", "17:30", "18:00", "Recoger a novia", "novia"),
    ("Viernes", "19:00", "21:00", "Tiempo de calidad con novia", "novia"),
    # --- SABADO ---
    ("Sabado", "08:00", "09:30", "Gimnasio (fuerza)", "gym"),
    ("Sabado", "10:00", "12:30", "PROYECTO estructural (bloque largo)", "proyecto"),
    ("Sabado", "14:00", "18:00", "Tiempo con novia / social", "novia"),
    ("Sabado", "20:00", "20:30", "Lectura 30 min", "lectura"),
    # --- DOMINGO ---
    ("Domingo", "10:00", "11:00", "Descanso / recargar", "descanso"),
    ("Domingo", "17:00", "17:15", "Revision semanal (15 min)", "psicologia"),
    ("Domingo", "17:15", "17:45", "Planear prioridades de la semana", "proyecto"),
]

# Franjas horarias para la rejilla (06:00 a 21:00 en bloques de 30 min)
def franjas():
    out = []
    t = datetime.strptime("06:00", "%H:%M")
    fin = datetime.strptime("21:00", "%H:%M")
    while t < fin:
        out.append(t.strftime("%H:%M"))
        t += timedelta(minutes=30)
    return out

FRANJAS = franjas()

def hm(s):
    return datetime.strptime(s, "%H:%M")

# Mapa (dia, franja) -> evento que arranca ahi, y set de franjas ocupadas
def construir_rejilla():
    inicio = {}   # (dia, franja_inicio) -> (titulo, cat, n_filas)
    ocupado = {}  # (dia, franja) -> cat (para pintar celdas combinadas)
    for dia, ini, fin, titulo, cat in EVENTOS:
        t0, t1 = hm(ini), hm(fin)
        n = int((t1 - t0).total_seconds() // 1800)  # bloques de 30 min
        inicio[(dia, ini)] = (titulo, cat, max(n, 1))
        tt = t0
        while tt < t1:
            ocupado[(dia, tt.strftime("%H:%M"))] = cat
            tt += timedelta(minutes=30)
    return inicio, ocupado


# ----------------------------------------------------------------------------
# 2) EXCEL
# ----------------------------------------------------------------------------
def generar_excel(path="Horario.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Titulo
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "MI HORARIO SEMANAL  -  9o Sem. Ing. Civil"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="37474F")
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # Encabezados
    headers = ["Hora"] + DIAS
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="546E7A")
        cell.alignment = center
        cell.border = border

    inicio, ocupado = construir_rejilla()

    # Filas de franjas
    for i, fr in enumerate(FRANJAS):
        r = i + 3
        hcell = ws.cell(row=r, column=1, value=fr)
        hcell.font = Font(bold=True, size=9)
        hcell.alignment = center
        hcell.fill = PatternFill("solid", fgColor="ECEFF1")
        hcell.border = border
        ws.row_dimensions[r].height = 22
        for j, dia in enumerate(DIAS, start=2):
            cell = ws.cell(row=r, column=j)
            cell.border = border
            cell.alignment = center
            if (dia, fr) in inicio:
                titulo, cat, n = inicio[(dia, fr)]
                cell.value = titulo
                cell.fill = PatternFill("solid", fgColor=CAT_COLOR.get(cat, "FFFFFF"))
                cell.font = Font(size=9, bold=cat in ("estructural", "clase"))
                if n > 1:
                    ws.merge_cells(start_row=r, start_column=j, end_row=r + n - 1, end_column=j)
            elif (dia, fr) in ocupado:
                cell.fill = PatternFill("solid", fgColor=CAT_COLOR.get(ocupado[(dia, fr)], "FFFFFF"))

    # Anchos
    ws.column_dimensions["A"].width = 8
    for j in range(2, 9):
        ws.column_dimensions[get_column_letter(j)].width = 20

    # Hoja de leyenda / tracker de habitos
    ws2 = wb.create_sheet("Tracker Habitos")
    ws2["A1"] = "TRACKER SEMANAL DE HABITOS"
    ws2["A1"].font = Font(bold=True, size=13)
    habitos = ["Dormir 7-8h", "Ingles", "Proyecto estructural", "Gym / Box",
               "Lectura 30 min", "Tiempo c/ novia", "Agua + comer bien"]
    cols = ["Habito"] + ["L", "M", "X", "J", "V", "S", "D"]
    for j, h in enumerate(cols, start=1):
        cc = ws2.cell(row=3, column=j, value=h)
        cc.font = Font(bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor="546E7A")
        cc.alignment = center
        cc.border = border
    for i, hb in enumerate(habitos, start=4):
        ws2.cell(row=i, column=1, value=hb).border = border
        for j in range(2, 9):
            cc = ws2.cell(row=i, column=j)
            cc.border = border
            cc.alignment = center
    ws2.column_dimensions["A"].width = 24
    for j in range(2, 9):
        ws2.column_dimensions[get_column_letter(j)].width = 5

    wb.save(path)
    print(f"OK -> {path}")


# ----------------------------------------------------------------------------
# 3) PDF
# ----------------------------------------------------------------------------
def generar_pdf(path="Horario.pdf"):
    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle("t", parent=styles["Title"], fontSize=16, spaceAfter=6)
    cell_style = ParagraphStyle("c", parent=styles["Normal"], fontSize=6.5, leading=7.5, alignment=1)

    elems = [Paragraph("Mi Horario Semanal - 9o Sem. Ing. Civil", titulo),
             Spacer(1, 0.2*cm)]

    inicio, ocupado = construir_rejilla()
    rl_colors = {k: colors.HexColor("#" + v) for k, v in CAT_COLOR.items()}

    # Construir matriz de la tabla
    header = ["Hora"] + DIAS
    data = [header]
    for fr in FRANJAS:
        row = [fr]
        for dia in DIAS:
            if (dia, fr) in inicio:
                titulo_e, cat, n = inicio[(dia, fr)]
                row.append(Paragraph(titulo_e, cell_style))
            else:
                row.append("")
        data.append(row)

    n_cols = len(header)
    col_w = [1.5*cm] + [(27.5*cm) / 7] * 7
    t = Table(data, colWidths=col_w, repeatRows=1)

    ts = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#546E7A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (0, -1), 7),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#ECEFF1")),
    ]

    # Pintar y combinar celdas por evento
    for ri, fr in enumerate(FRANJAS, start=1):
        for ci, dia in enumerate(DIAS, start=1):
            if (dia, fr) in inicio:
                _, cat, n = inicio[(dia, fr)]
                color = rl_colors.get(cat, colors.white)
                ts.append(("BACKGROUND", (ci, ri), (ci, ri + n - 1), color))
                if n > 1:
                    ts.append(("SPAN", (ci, ri), (ci, ri + n - 1)))
            elif (dia, fr) in ocupado:
                ts.append(("BACKGROUND", (ci, ri), (ci, ri), rl_colors.get(ocupado[(dia, fr)], colors.white)))

    t.setStyle(TableStyle(ts))
    elems.append(t)

    # Leyenda
    elems.append(Spacer(1, 0.3*cm))
    leyenda = [
        ("Ingles", "ingles"), ("Clase oblig.", "clase"), ("Estructural NO FALTAR", "estructural"),
        ("Proyecto/estudio", "proyecto"), ("Novia", "novia"), ("Gym", "gym"),
        ("Box", "box"), ("Psicologia/Repaso", "psicologia"), ("Lectura", "lectura"), ("Viaje", "viaje"),
    ]
    leg_cells = [Paragraph(f'<font size=7>{txt}</font>', cell_style) for txt, _ in leyenda]
    leg_table = Table([leg_cells], colWidths=[2.75*cm]*len(leyenda))
    leg_style = [("GRID", (0, 0), (-1, -1), 0.3, colors.white),
                 ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                 ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]
    for i, (_, cat) in enumerate(leyenda):
        leg_style.append(("BACKGROUND", (i, 0), (i, 0), rl_colors.get(cat, colors.white)))
    leg_table.setStyle(TableStyle(leg_style))
    elems.append(leg_table)

    doc.build(elems)
    print(f"OK -> {path}")


# ----------------------------------------------------------------------------
# 4) ICS (Google Calendar)  - eventos recurrentes semanales
# ----------------------------------------------------------------------------
ICS_DAY = {"Lunes": "MO", "Martes": "TU", "Miercoles": "WE", "Jueves": "TH",
           "Viernes": "FR", "Sabado": "SA", "Domingo": "SU"}
# Fecha base por dia (semana del lunes 2026-06-15)
BASE = {"Lunes": "20260615", "Martes": "20260616", "Miercoles": "20260617",
        "Jueves": "20260618", "Viernes": "20260619", "Sabado": "20260620", "Domingo": "20260621"}

def generar_ics(path="Horario.ics"):
    lines = [
        "BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Plan Vida//Horario//ES",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH", "X-WR-CALNAME:Mi Horario Semanal",
        "X-WR-TIMEZONE:America/Bogota",
        "BEGIN:VTIMEZONE", "TZID:America/Bogota",
        "BEGIN:STANDARD", "DTSTART:19930101T000000", "TZOFFSETFROM:-0500",
        "TZOFFSETTO:-0500", "TZNAME:-05", "END:STANDARD", "END:VTIMEZONE",
    ]
    stamp = "20260616T120000Z"
    for idx, (dia, ini, fin, titulo, cat) in enumerate(EVENTOS):
        base = BASE[dia]
        dtstart = f"{base}T{ini.replace(':','')}00"
        dtend = f"{base}T{fin.replace(':','')}00"
        uid = f"plan-vida-{idx}@horario"
        lines += [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{stamp}",
            f"DTSTART;TZID=America/Bogota:{dtstart}",
            f"DTEND;TZID=America/Bogota:{dtend}",
            f"RRULE:FREQ=WEEKLY;BYDAY={ICS_DAY[dia]}",
            f"SUMMARY:{titulo}",
            f"CATEGORIES:{cat}",
            "END:VEVENT",
        ]
    lines.append("END:VCALENDAR")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\r\n".join(lines))
    print(f"OK -> {path}")


if __name__ == "__main__":
    generar_excel()
    generar_pdf()
    generar_ics()
    print("\nTodo generado: Horario.xlsx, Horario.pdf, Horario.ics")
