# -*- coding: utf-8 -*-
"""
Calculo de ACERO por peso usando la tabla de masas de varilla.
Metodo: sum(longitud_total_por_calibre * masa_kg_por_m) + 10% desperdicio/traslapos.
Datos de barras tomados de los DESPIECES del plano ESTRUCTURAL - MODULO 1 - FC3.
"""
import openpyxl

# Tabla de masa lineal (kg/m) por calibre (N.)
MASA = {2:0.250, 3:0.560, 4:0.994, 5:1.552, 6:2.235, 7:3.042, 8:3.973}

def kg(barras):
    """barras = lista de (n, calibre, longitud_m). Devuelve (kg_total, metros_por_calibre)."""
    total = 0.0
    porcal = {}
    for n, cal, L in barras:
        m = n * L
        porcal[cal] = porcal.get(cal, 0.0) + m
        total += m * MASA[cal]
    return total, porcal

def mostrar(nombre, long_barras, est_barras):
    kL, mL = kg(long_barras)
    kE, mE = kg(est_barras)
    sub = kL + kE
    con10 = sub * 1.10
    print(f"\n===== {nombre} =====")
    print(f"  Longitudinal: {kL:8.1f} kg   (metros/calibre: " +
          ", ".join(f'#{c}:{m:.1f}m' for c,m in sorted(mL.items())) + ")")
    print(f"  Estribos:     {kE:8.1f} kg   (metros/calibre: " +
          ", ".join(f'#{c}:{m:.1f}m' for c,m in sorted(mE.items())) + ")")
    print(f"  Subtotal:     {sub:8.1f} kg")
    print(f"  +10%:         {con10:8.1f} kg")
    return round(kL,1), round(kE,1), round(sub,1), round(con10,1)

# ============================================================
# 4.7.1 COLUMNAS  (8 uds: 4 Tipo1 + 4 Tipo2)
# T1 0.40x0.40: 6#6 L=5.85 ; 42 estribos #3 L=1.48 ; 42 grapas #3 L=0.54
# T2 0.50x0.50: 4#6+4#7 L=5.95 ; 42 estribos #3 L=1.88 ; 42 grapas #3 L=0.64
# ============================================================
col_long = [
    (4*6, 6, 5.85),           # T1 longitudinal #6
    (4*4, 6, 5.95),           # T2 longitudinal #6
    (4*4, 7, 5.95),           # T2 longitudinal #7
]
col_est = [
    (4*42, 3, 1.48),          # T1 estribos
    (4*42, 3, 0.54),          # T1 grapas
    (4*42, 3, 1.88),          # T2 estribos
    (4*42, 3, 0.64),          # T2 grapas
]
col = mostrar("4.7.1 COLUMNAS (4 T1 + 4 T2)", col_long, col_est)

# ============================================================
# 2.3.1 VIGAS DE CIMENTACION  B=0.40 H=0.45, estribos #3 L=1.58
# Transversales (5): VC-1, VC-2, VC-3(=VC-2), VC-3', VC-4(=VC-3')
# Longitudinales (2): VC-A, VC-B
# ============================================================
vc_long = [
    # ---- VC-1 ----
    (3,6,3.00),(3,6,6.90),(3,5,5.00),(3,5,4.60),
    # ---- VC-2 ----
    (3,6,3.00),(3,6,6.80),(3,5,5.00),(3,5,4.50),
    # ---- VC-3 (=VC-2) ----
    (3,6,3.00),(3,6,6.80),(3,5,5.00),(3,5,4.50),
    # ---- VC-3' ----
    (3,6,3.00),(3,6,6.70),(3,5,5.00),(3,5,4.40),
    # ---- VC-4 (=VC-3') ----
    (3,6,3.00),(3,6,6.70),(3,5,5.00),(3,5,4.40),
    # ---- VC-A y VC-B (x2) ----  top:3#5 5.50 + 3#6 7.60 + 3#6 7.40 ; bot:3#5 4.00 + 2#5 7.80 + 3#5 6.30 + 1#5 4.00
    (2*3,5,5.50),(2*3,6,7.60),(2*3,6,7.40),
    (2*3,5,4.00),(2*2,5,7.80),(2*3,5,6.30),(2*1,5,4.00),
]
vc_est = [
    (5*36, 3, 1.58),          # 5 vigas transversales, 36 estribos c/u
    (2*89, 3, 1.58),          # 2 vigas longitudinales, 89 estribos c/u (29+32+28)
]
vcim = mostrar("2.3.1 VIGAS DE CIMENTACION (5 transv + 2 long)", vc_long, vc_est)

# ============================================================
# 4.8.1 VIGAS DE CUBIERTA  B=0.40 H=0.50, estribos #3 L=1.68
# Transversales (4): VG-1..VG-4
# Longitudinales (2): VG-6, VG-7 (se toma despiece VG-6 para ambas)
# ============================================================
vg_long = [
    # VG-1..4 (x4): top 2#5 5.00 + 2#5 4.40 ; bot 2#6 3.00 + 2#6 6.70
    (4*2,5,5.00),(4*2,5,4.40),(4*2,6,3.00),(4*2,6,6.70),
    # VG-6 y VG-7 (x2): top 3#5 4.00+3#5 7.80+3#5 6.30+3#5 6.00 + 1#6 4.00 ; bot 3#5 5.50 + 3#6 7.60+3#6 7.40+3#6 3.50
    (2*3,5,4.00),(2*3,5,7.80),(2*3,5,6.30),(2*3,5,6.00),(2*1,6,4.00),
    (2*3,5,5.50),(2*3,6,7.60),(2*3,6,7.40),(2*3,6,3.50),
]
vg_est = [
    (4*73, 3, 1.68),          # 4 vigas transversales, 73 estribos c/u
    (2*182, 3, 1.68),         # 2 vigas longitudinales, 182 estribos c/u
]
vcub = mostrar("4.8.1 VIGAS DE CUBIERTA (4 transv + 2 long)", vg_long, vg_est)

# ============================================================
# VIGUETAS (referencia - placa aligerada)  B=0.20 H=0.40, estribos #3 L=1.08
# 6 viguetas: top 3#4(4.00+7.80+7.10+5.20) ; bot 3#4(5.50+7.60+7.40+3.50) ; 186 estribos c/u
# ============================================================
vgt_long = [
    (6*3,4,4.00),(6*3,4,7.80),(6*3,4,7.10),(6*3,4,5.20),
    (6*3,4,5.50),(6*3,4,7.60),(6*3,4,7.40),(6*3,4,3.50),
]
vgt_est = [(6*186, 3, 1.08)]
vgt = mostrar("VIGUETAS (6 uds - referencia, placa aligerada)", vgt_long, vgt_est)

# ============================================================
# 2.3.2 ZAPATAS  #5 @0.20 en dos sentidos
# Z-1 (4 uds) 2.30x2.30 ; Z-2 (4 uds) 2.15x2.15 ; recubrimiento => Lbarra = lado-0.15
# n barras por sentido = lado/0.20 + 1
# ============================================================
def zapata(n_zap, lado):
    Lb = lado - 0.15
    n_por_sentido = int(lado/0.20) + 1
    n_total = n_zap * 2 * n_por_sentido      # dos sentidos
    return [(n_total, 5, Lb)]
zap_barras = zapata(4, 2.30) + zapata(4, 2.15)
kZ, mZ = kg(zap_barras)
print("\n===== 2.3.2 ZAPATAS (4 Z-1 + 4 Z-2, #5@0.20 dos sentidos) =====")
print(f"  Acero #5: {kZ:8.1f} kg  (#5:{mZ[5]:.1f} m)")
print(f"  +10%:     {kZ*1.10:8.1f} kg")
zap = (round(kZ,1), round(kZ*1.10,1))

# ============================================================
# 5.2.1 MAMPOSTERIA CONFINADA
# Columnetas 0.20x0.15: 2#4 vert + estribo #3 c/0.20
# Cinta V2: 2#3 horiz + gancho #2 (1/4") c/0.30
# Conteo estimado del plano de mamposteria (por confirmar en DWG).
# ============================================================
# Columnetas: ~24 uds, altura libre 2.70 m (long barra 2.70+traslapo) ; estribos c/0.20 -> ~14 c/u
N_COLNETA = 24; H_COLNETA = 2.70; N_EST_COL = 14; L_EST_COL = 0.60  # per(0.20+0.15)*2+ganchos
# Cintas: longitud total de muros ~ 70 m (perimetral + divisorias)
L_CINTA = 70.0; N_GANCHO = int(L_CINTA/0.30)
mamp_long = [
    (N_COLNETA*2, 4, H_COLNETA),      # 2#4 por columneta
    (2, 3, L_CINTA),                  # 2#3 corridos en cinta
]
mamp_est = [
    (N_COLNETA*N_EST_COL, 3, L_EST_COL),  # estribos columnetas
    (N_GANCHO, 2, 0.35),                  # ganchos cinta #2
]
mamp = mostrar("5.2.1 MAMPOSTERIA (estimado: 24 columnetas + 70 m cinta)", mamp_long, mamp_est)

# ============================================================
# RESUMEN
# ============================================================
print("\n\n############ RESUMEN ACERO (kg, con +10%) ############")
print(f"  2.3.1 Vigas cimentacion : {vcim[3]:8.1f}")
print(f"  2.3.2 Zapatas           : {zap[1]:8.1f}")
print(f"  4.7.1 Columnas          : {col[3]:8.1f}")
print(f"  4.8.1 Vigas cubierta    : {vcub[3]:8.1f}")
print(f"  5.2.1 Mamposteria       : {mamp[3]:8.1f}")
print(f"  (ref) Viguetas          : {vgt[3]:8.1f}")
tot = vcim[3]+zap[1]+col[3]+vcub[3]+mamp[3]
print(f"  TOTAL (5 items memoria) : {tot:8.1f}")

# ============================================================
# ESCRIBIR EN EL EXCEL
# ============================================================
FN = 'MEMORIAS DE CALCULO.xlsx'
wb = openpyxl.load_workbook(FN)

def set_item(sheet, k17, r17, k18, r18):
    ws = wb[sheet]
    ws['K17'] = k17; ws['R17'] = r17
    ws['K18'] = k18; ws['R18'] = r18
    ws['R19'] = round(r17 + r18, 1)
    ws['R20'] = round((r17 + r18) * 1.10, 1)

set_item('4.7.1',
    'Long. #6+#7 (T1:6#6 L5.85 / T2:4#6+4#7 L5.95) x8 col', col[0],
    'Estribos+grapas #3 (T1:42x1.48+42x0.54 / T2:42x1.88+42x0.64) x8', col[1])
wb['4.7.1']['B40'] = ('Calculado por peso con tabla de masas de varilla: L(m)xkg/m + 10%. '
    'kg/m usados: #3=0.560, #6=2.235, #7=3.042. Barras del despiece de columnas (Son 4 c/tipo).')

set_item('2.3.1',
    'Long. #6 (236.7 m) + #5 (276.2 m) - VC-1..4,3\',A,B', vcim[0],
    'Estribos #3 L=1.58 (5x36 transv + 2x89 long)', vcim[1])
wb['2.3.1']['B39'] = 'OBSERVACIONES'
wb['2.3.1']['B40'] = ('Por peso con tabla de masas: #5=1.552, #6=2.235, #3=0.560 kg/m; +10%. '
    'Barras del despiece: VC-1/2/3/3\'/4 (transv) y VC-A/VC-B (long). VC-3 se asume = VC-2.')

set_item('4.8.1',
    'Long. #5 (252.8 m) + #6 (196.6 m) - VG-1..4 + VG-6/VG-7', vcub[0],
    'Estribos #3 L=1.68 (4x73 transv + 2x182 long)', vcub[1])
wb['4.8.1']['B39'] = 'OBSERVACIONES'
wb['4.8.1']['B40'] = ('Por peso con tabla de masas: #5=1.552, #6=2.235, #3=0.560 kg/m; +10%. '
    'NO incluye viguetas (placa aligerada): 6 viguetas 0.20x0.40 con #4 y est.#3 = '
    f'{vgt[3]} kg aprox, contabilizar en la placa de cubierta si aplica.')

set_item('2.3.2 ACERO ZAPATAS',
    'Z-1 (4) 2.30x2.30  #5@0.20 dos sentidos', round(kg(zapata(4,2.30))[0],1),
    'Z-2 (4) 2.15x2.15  #5@0.20 dos sentidos', round(kg(zapata(4,2.15))[0],1))
wb['2.3.2 ACERO ZAPATAS']['B40'] = ('Por peso con tabla de masas: #5=1.552 kg/m; +10%. '
    'Lbarra=lado-0.15 (recubr.); n=lado/0.20+1 por sentido. Confirmar lados exactos en DWG.')

set_item('5.2.1',
    'Columnetas 2#4 (0.20x0.15) + Cinta V2 2#3', mamp[0],
    'Estribo #3 c/0.20 (columneta) + gancho #2 (1/4") c/0.30 (cinta)', mamp[1])
wb['5.2.1']['B39'] = 'OBSERVACIONES'
wb['5.2.1']['B40'] = ('ESTIMACION por peso (tabla masas): columneta 2#4+est#3; cinta V2 2#3+gancho#2. '
    'Cantidades (24 columnetas h=2.70; 70 m de cinta) POR CONFIRMAR en el plano de mamposteria FC3.')

wb.save(FN)
print("\nExcel actualizado OK:", FN)
