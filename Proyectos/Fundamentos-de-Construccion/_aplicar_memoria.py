# -*- coding: utf-8 -*-
"""Aplica los valores de acero recalculados y agrega observaciones cortas
(estilo del usuario) en la memoria con fotos, y reemplaza las cantidades de
acero en el presupuesto."""
import openpyxl

MEM = 'Memoria de calculo Fundamentos.xlsx'
CANT = 'Cantidades de obra fundamentos.xlsx'

# ---------------------------------------------------------------
# 1) MEMORIA: valores de acero (R17, R18) + descripciones (K17, K18)
#    R19 y R20 son formulas: =SUM(R17:R18) y =R19*1.1  -> se recalculan solas
# ---------------------------------------------------------------
wb = openpyxl.load_workbook(MEM)

acero = {
    '2.3.1': dict(
        k17='Long. #6 (236.7 m) + #5 (276.2 m)', r17=957.7,
        k18='Estribos #3 L=1.58 (5x36 transv + 2x89 long)', r18=316.8),
    '4.7.1': dict(
        k17='Long. T1 6#6 L=5.85 + T2 4#6+4#7 L=5.95 (8 col)', r17=816.2,
        k18='Estribos + grapas #3 (T1 y T2)', r18=427.1),
    '4.8.1': dict(
        k17='Long. #5 (252.8 m) + #6 (196.6 m)', r17=831.7,
        k18='Estribos #3 c/0.10 (4x73 + 2x182)', r18=617.2),
    '5.2.1': dict(
        k17='Long. columnetas 2#4 + cinta 2#3', r17=207.2,
        k18='Estribos #3 (columneta) + ganchos #2 (cinta)', r18=133.3),
    '2.3.2 ACERO ZAPATAS': dict(
        k17='Z-1: 4 zap 2.30x2.30, #5@0.20 dos sentidos', r17=320.3,
        k18='Z-2: 4 zap 2.15x2.15, #5@0.20 dos sentidos', r18=273.2),
}
for sh, d in acero.items():
    ws = wb[sh]
    ws['K17'] = d['k17']; ws['R17'] = d['r17']
    ws['K18'] = d['k18']; ws['R18'] = d['r18']

# limpiar dimensiones basura (copiadas) en filas de acero
for sh in acero:
    ws = wb[sh]
    for r in (17, 18):
        for c in ('N', 'O', 'P', 'Q'):
            ws[f'{c}{r}'] = None

# ---------------------------------------------------------------
# 2) OBSERVACIONES (B40) - solo en las hojas que NO tienen, estilo del usuario
# ---------------------------------------------------------------
obs = {
 '4.3.1 Placa-Contrapiso':
   'placa de contrapiso e=0.10 m sobre toda la huella del modulo (9.85 x 21.95 m). area tomada del plano arquitectonico de plantas.',
 '4.3.3 PLACA CUBIERTA':
   'placa aligerada de cubierta; ancho 7.30 m y luces (5.55+6.10+6.87) del plano estructural de cubierta. incluye 6 viguetas de 0.20x0.32.',
 '8.1 PISOS ACABADOS':
   'adoquin en zonas exteriores; areas 1 a 9 medidas directamente sobre el plano de adoquin (planta de acabados).',
 '2.1.4':
   'recebo de nivelacion e=0.20 m bajo la placa, en toda la huella (9.85 x 21.95 m). completa la excavacion masiva de 0.60 (0.20 recebo + 0.30 subbase + 0.10 placa).',
 '2.1.6':
   'geotextil NT 1600 en toda el area de la huella bajo la placa (9.85 x 21.95 m).',
 '2.2.1':
   'vigas de cimentacion 0.40x0.45 (seccion del despiece); 2 longitudinales de 19.80 m y 5 transversales de 7.30 m (ejes del plano de cimentacion).',
 '4.2.1':
   'vigas de cubierta 0.40x0.50; 2 longitudinales de 19.80 m y 5 transversales de 7.30 m, luces del plano estructural de cubierta.',
 '5.1.1':
   'area de muros: longitud total 173.60 m x altura 2.90 m, menos 20% por vanos de puertas y ventanas.',
 '5.1.2':
   'columnetas de confinamiento cada ~3 m en los muros; 58 unidades de 2.90 m de altura (medido en ml).',
 '5.1.4':
   'panete sobre el 65% del area de muros (402.75 m2).',
 '2.3.1':
   'acero por peso: longitud de cada barra del despiece (VC-1 a VC-4, A y B) x masa por metro (#5=1.552, #6=2.235, #3=0.560 kg/m) + 10% de traslapos.',
 '4.7.1':
   'acero por peso: longitudinal #6 y #7 del despiece (tipo1 6#6; tipo2 4#6+4#7) + estribos y grapas #3, por su masa por metro + 10%.',
 '4.8.1':
   'acero por peso: barras #5 y #6 del despiece (VG-1 a VG-4 y VG longitudinales) + estribos #3 c/0.10, por su masa por metro + 10%.',
 '5.2.1':
   'acero por peso: columnetas 2#4 con estribo #3 y viga cinta 2#3 con gancho #2 (detalle de confinamiento), por su masa por metro + 10%.',
}
added = []
for sh, txt in obs.items():
    ws = wb[sh]
    if not ws['B40'].value:            # no pisar las que ya tienen
        ws['B40'] = txt
        added.append(sh)

wb.save(MEM)
print('MEMORIA guardada. Observaciones agregadas en:', len(added), 'hojas ->', added)
print('Acero actualizado en:', list(acero.keys()))

# ---------------------------------------------------------------
# 3) CANTIDADES DE OBRA: reemplazar cantidades (D) de acero por los nuevos totales (+10%)
# ---------------------------------------------------------------
wb2 = openpyxl.load_workbook(CANT)
ws2 = wb2.active
def tot(sh):
    d = acero[sh]
    return round((d['r17'] + d['r18']) * 1.1, 1)
nuevos = {
    'D18': tot('2.3.1'),   # acero vigas cimentacion
    'D38': tot('4.7.1'),   # acero columnas
    'D41': tot('4.8.1'),   # acero vigas cubierta
    'D54': tot('5.2.1'),   # acero mamposteria
}
for celda, val in nuevos.items():
    viejo = ws2[celda].value
    ws2[celda] = val
    print(f'  {celda}: {viejo} -> {val}')
wb2.save(CANT)
print('CANTIDADES guardadas.')
