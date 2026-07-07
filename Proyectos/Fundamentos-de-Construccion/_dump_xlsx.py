import openpyxl, sys

def dump(path, max_rows=300, max_cols=30):
    print("="*100)
    print("ARCHIVO:", path)
    print("="*100)
    wb = openpyxl.load_workbook(path, data_only=False)
    print("HOJAS:", wb.sheetnames)
    for ws in wb.worksheets:
        print("\n" + "#"*80)
        print(f"HOJA: {ws.title}  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
        print("#"*80)
        rows = min(ws.max_row, max_rows)
        cols = min(ws.max_column, max_cols)
        for r in range(1, rows+1):
            line = []
            empty = True
            for c in range(1, cols+1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is not None and str(v).strip() != "":
                    empty = False
                    line.append(f"{cell.coordinate}={v!r}")
            if not empty:
                print(f"  R{r}: " + " | ".join(line))

for p in sys.argv[1:]:
    try:
        dump(p)
    except Exception as e:
        print("ERROR con", p, e)
