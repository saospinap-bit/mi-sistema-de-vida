# -*- coding: utf-8 -*-
import openpyxl
f = "MEMORIAS DE CALCULO.xlsx"
wb = openpyxl.load_workbook(f)
ws = wb["2.1.1"]
ws["K17"] = "Excavacion masiva plataforma (sub-base 0.30 + placa 0.10)"
ws["O17"] = 0.40
ws["B40"] = ("CORREGIDO: 0.60 -> 0.40 m. El detalle de la losa solo muestra sub-base 0.30 + "
             "placa 0.10. El recebo (2.1.4) NO va aqui: es relleno que va en otra parte "
             "(backfill sobre las vigas). Huella 9.85x21.95 aun por confirmar en DWG.")
wb.save(f)
print("2.1.1 masiva =", round(9.85*0.40*21.95,2), "m3")
