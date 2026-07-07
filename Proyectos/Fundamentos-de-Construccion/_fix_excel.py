# -*- coding: utf-8 -*-
"""Corrige la memoria de calculo con los valores verificados en los planos."""
import openpyxl

f = "MEMORIAS DE CALCULO.xlsx"
wb = openpyxl.load_workbook(f)

def obs(ws, texto):
    # escribe una observacion debajo del rotulo OBSERVACIONES (B39)
    ws["B40"] = texto

# ---------------------------------------------------------------
# 4.1.1  COLUMNAS  -> 8 columnas en 2 tipos (verificado en despiece)
# ---------------------------------------------------------------
ws = wb["4.1.1"]
ws["K17"] = "Columnas tipo 1 (0.40x0.40) h=4.70 - ejes 1A,1B,3'B,4A"
ws["N17"] = 0.40; ws["O17"] = 0.40; ws["P17"] = 4.70; ws["Q17"] = 4
ws["R17"] = "=N17*O17*P17*Q17"
ws["K18"] = "Columnas tipo 2 (0.50x0.50) h=4.90 - ejes 2A,2B,3A,3B"
ws["N18"] = 0.50; ws["O18"] = 0.50; ws["P18"] = 4.90; ws["Q18"] = 4
ws["R18"] = "=N18*O18*P18*Q18"
obs(ws, "CORREGIDO: el despiece de columnas muestra 8 columnas en 2 tipos "
        "(4 de 0.40x0.40 y 4 de 0.50x0.50). Antes se habia puesto 10 col de 0.40x0.40.")

# ---------------------------------------------------------------
# 4.3.3  PLACA DE CUBIERTA (M2) -> area de losa; viguetas 0.20x0.40
# ---------------------------------------------------------------
ws = wb["4.3.3 PLACA CUBIERTA"]
ws["N16"] = "Ancho (m)"; ws["O16"] = "Alto (m)"; ws["P16"] = "Largo (m)"
ws["K17"] = "Placa aligerada e=0.08 (area en planta)"
ws["N17"] = 7.30; ws["O17"] = 1; ws["P17"] = "=5.55+6.10+6.87"; ws["Q17"] = 1
ws["R17"] = "=N17*O17*P17*Q17"
ws["K18"] = "Viguetas 0.20x0.40 (incluidas en la placa aligerada - no suma area)"
ws["N18"] = 0.20; ws["O18"] = 0.40; ws["P18"] = "=5.55+6.10+6.87"; ws["Q18"] = 0
ws["R18"] = "=N18*O18*P18*Q18"
obs(ws, "CORREGIDO: item por M2 -> se toma AREA de losa (Alto=1). Viguetas son "
        "0.20x0.40 (el plano dice 20x40, no 20x32) y van incluidas en el m2 de losa "
        "aligerada, por eso su fila no suma (Cantidad=0). El concreto de viguetas se "
        "detalla en la cartilla / APU.")

# ---------------------------------------------------------------
# 4.7.1  ACERO COLUMNAS -> descripcion segun 2 tipos
# ---------------------------------------------------------------
ws = wb["4.7.1"]
ws["K17"] = "Long. Tipo1 (4col x 6#6 L=5.85) + Tipo2 (4col x 4#6+4#7 L=5.95)"
obs(ws, "REVISAR: el refuerzo longitudinal difiere por tipo de columna "
        "(Tipo1: 6#6; Tipo2: 4#6+4#7). El peso exacto (kg) debe salir de la "
        "cartilla de hierros (DLNET). El valor mostrado es una estimacion con +10%.")

# ---------------------------------------------------------------
# Observaciones de trazabilidad (huella no acotada en los PDF)
# ---------------------------------------------------------------
huella = ("VERIFICAR EN DWG: la huella 9.85 x 21.95 no esta acotada en los PDF. "
          "La reticula estructural acotada es 19.80 (ejes 1-4) x 7.30/7.70 (ejes A-B). "
          "Confirmar la dimension en el plano arquitectonico/DWG y pegar el pantallazo.")
for hoja in ["2.1.1", "2.1.2", "2.1.4", "2.1.6", "4.3.1 Placa-Contrapiso"]:
    if hoja in wb.sheetnames:
        obs(wb[hoja], huella)

obs(wb["5.1.1"], "VERIFICAR EN DWG: L=173.60 m es la suma de longitudes de todos los "
                 "muros (medida sobre plano). Pegar pantallazo con la medicion. H=2.90 m "
                 "= piso a viga de cubierta.")
obs(wb["5.1.2"], "VERIFICAR EN DWG: 58 columnetas = una cada ~3 m sobre los muros. "
                 "Contar sobre el plano de mamposteria y pegar pantallazo.")
obs(wb["8.1 PISOS ACABADOS"], "VERIFICAR EN DWG: las 9 areas de adoquin se miden sobre "
                 "el plano de adoquin/arquitectonico (zonas exteriores). Pegar pantallazo "
                 "con las areas resaltadas.")

wb.save(f)
print("Excel corregido y guardado:", f)

# Reporte de totales corregidos (evaluando formulas simples)
wb2 = openpyxl.load_workbook(f, data_only=False)
def val(ws, r):
    return (ws.cell(r,14).value, ws.cell(r,15).value, ws.cell(r,16).value, ws.cell(r,17).value)
print("4.1.1 R17:", val(wb2["4.1.1"],17), " R18:", val(wb2["4.1.1"],18))
print("4.3.3 R17:", val(wb2["4.3.3 PLACA CUBIERTA"],17))
