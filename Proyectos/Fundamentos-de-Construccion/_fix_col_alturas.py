# -*- coding: utf-8 -*-
import openpyxl
f = "MEMORIAS DE CALCULO.xlsx"
wb = openpyxl.load_workbook(f)
ws = wb["4.1.1"]
ws["P17"] = 4.75  # Tipo 1 altura (de -0.05 a +4.70)
ws["P18"] = 4.95  # Tipo 2 altura (de -0.05 a +4.90)
ws["K17"] = "Columnas tipo 1 (0.40x0.40) h=4.75 - ejes 1A,1B,3'B,4A"
ws["K18"] = "Columnas tipo 2 (0.50x0.50) h=4.95 - ejes 2A,2B,3A,3B"
ws["B40"] = ("CORREGIDO: 8 columnas en 2 tipos. Alturas de -0.05 a +4.70 (T1=4.75) y "
             "a +4.90 (T2=4.95), leidas de los niveles del despiece. Vol total = 7.99 m3.")
wb.save(f)
print("Alturas corregidas. Vol =", round(0.4*0.4*4.75*4 + 0.5*0.5*4.95*4,2), "m3")
