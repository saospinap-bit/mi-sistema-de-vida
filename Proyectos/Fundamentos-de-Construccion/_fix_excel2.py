# -*- coding: utf-8 -*-
"""Corrige TODO el capitulo 2 (cimentacion) con la logica de niveles verificada."""
import openpyxl

f = "MEMORIAS DE CALCULO.xlsx"
wb = openpyxl.load_workbook(f)

def obs(ws, texto):
    ws["B40"] = texto

# ---------------------------------------------------------------
# 2.1.1  EXCAVACION MASIVA  -> profundidad 0.60 (recebo0.20+subbase0.30+placa0.10)
# ---------------------------------------------------------------
ws = wb["2.1.1"]
ws["K17"] = "Excavacion masiva plataforma (reemplazo bajo placa)"
ws["O17"] = 0.60
obs(ws, "CORREGIDO: prof. 0.60 m = recebo 0.20 + sub-base 0.30 + placa 0.10 (paquete de "
        "reemplazo bajo la placa, hasta nivel -0.65). Antes 0.55 no cuadraba. "
        "Huella 9.85x21.95 sigue pendiente de confirmar en DWG.")

# ---------------------------------------------------------------
# 2.1.2  SUB-BASE  -> e = 0.30 (el detalle de losa dice SUB-BASE 30 cm)
# ---------------------------------------------------------------
ws = wb["2.1.2"]
ws["K17"] = "Capa sub-base bajo placa e=0.30"
ws["O17"] = 0.30
obs(ws, "CORREGIDO: e=0.30 m. El detalle de losa de contrapiso dice 'SUB-BASE DE 30 cm' "
        "y 'Minimo 0.30m'. Antes estaba en 0.15.")

# ---------------------------------------------------------------
# 2.1.3  EXCAVACION MANUAL  -> zanjas de vigas (prof 0.45) + fosos de zapatas
# ---------------------------------------------------------------
ws = wb["2.1.3"]
ws["K17"] = "Zanjas vigas long. A,B (0.40x0.45, de -0.65 a -1.10)"
ws["N17"] = 0.40; ws["O17"] = 0.45; ws["P17"] = 19.80; ws["Q17"] = 2
ws["R17"] = "=N17*O17*P17*Q17"
ws["K18"] = "Zanjas vigas transv. (0.40x0.45)"
ws["N18"] = 0.40; ws["O18"] = 0.45; ws["P18"] = 7.30; ws["Q18"] = 5
ws["R18"] = "=N18*O18*P18*Q18"
ws["K19"] = "Fosos zapatas Z-1 (2.30x2.30) prof 0.90 (de -0.65 a -1.55)"
ws["N19"] = 2.30; ws["O19"] = 0.90; ws["P19"] = 2.30; ws["Q19"] = 4
ws["R19"] = "=N19*O19*P19*Q19"
ws["K20"] = "Fosos zapatas Z-2 (2.15x2.15) prof 0.90"
ws["N20"] = 2.15; ws["O20"] = 0.90; ws["P20"] = 2.15; ws["Q20"] = 4
ws["R20"] = "=N20*O20*P20*Q20"
obs(ws, "CORREGIDO: excavacion manual por debajo del nivel masivo (-0.65): zanjas de vigas "
        "0.45 m de prof (altura de la viga) + fosos de zapatas 0.90 m hasta el desplante "
        "(-1.55). Estimacion sin descontar solapes viga-zapata (conservador).")

# ---------------------------------------------------------------
# 2.2.2  SOLADO  -> agregar solado bajo zapatas
# ---------------------------------------------------------------
ws = wb["2.2.2"]
ws["K19"] = "Solado bajo zapatas Z-1 (2.30x2.30)"
ws["N19"] = 2.30; ws["O19"] = 1; ws["P19"] = 2.30; ws["Q19"] = 4
ws["R19"] = "=N19*O19*P19*Q19"
ws["K20"] = "Solado bajo zapatas Z-2 (2.15x2.15)"
ws["N20"] = 2.15; ws["O20"] = 1; ws["P20"] = 2.15; ws["Q20"] = 4
ws["R20"] = "=N20*O20*P20*Q20"
obs(ws, "AGREGADO: solado de limpieza e=0.05 tambien bajo las 8 zapatas (el detalle de "
        "zapata lo muestra). Medido por m2.")

# ---------------------------------------------------------------
# NUEVA HOJA: 2.2.3  ZAPATAS CONCRETO 3000 psi
# ---------------------------------------------------------------
zc = wb.copy_worksheet(wb["2.2.1"])
zc.title = "2.2.3 ZAPATAS"
zc["C13"] = "2"; zc["D13"] = "CIMENTACION"; zc["J13"] = "2.2.3"
zc["K13"] = "Zapatas concreto 3000psi"; zc["R13"] = "M3"
zc["K17"] = "Zapatas Z-1 (2.30x2.30x0.40) - ejes 2A,3A,2B,3B"
zc["N17"] = 2.30; zc["O17"] = 0.40; zc["P17"] = 2.30; zc["Q17"] = 4
zc["R17"] = "=N17*O17*P17*Q17"
zc["K18"] = "Zapatas Z-2 (2.15x2.15x0.40) - ejes 1A,4A,1B,4B"
zc["N18"] = 2.15; zc["O18"] = 0.40; zc["P18"] = 2.15; zc["Q18"] = 4
zc["R18"] = "=N18*O18*P18*Q18"
obs(zc, "AGREGADO: 8 zapatas (4 Z-1 + 4 Z-2), H=0.40. Falta la linea correspondiente en el "
        "presupuesto (2.2.3). Concreto 3000 psi.")

# ---------------------------------------------------------------
# NUEVA HOJA: 2.3.2  ACERO ZAPATAS (#5 @0.20 en ambos sentidos) - estimacion
# ---------------------------------------------------------------
za = wb.copy_worksheet(wb["2.3.1"])
za.title = "2.3.2 ACERO ZAPATAS"
za["C13"] = "2"; za["D13"] = "REFUERZOS / ACERO"; za["J13"] = "2.3.2"
za["K13"] = "Acero figurado zapatas"; za["R13"] = "KG"
za["K17"] = "Z-1: 4 zap, #5@0.20 dos sentidos (est.)"; za["R17"] = 327.8
za["K18"] = "Z-2: 4 zap, #5@0.20 dos sentidos (est.)"; za["R18"] = 280.0
za["K19"] = "Subtotal (kg)"; za["R19"] = 607.8
za["K20"] = "Con 10% desperd/traslapo (kg)"; za["R20"] = 668.6
za["R37"] = "=R20"
obs(za, "AGREGADO (ESTIMACION): #5 @0.20 en ambos sentidos, ambas zapatas. El peso exacto "
        "debe salir de la cartilla de hierros (DLNET). Falta linea en el presupuesto.")

wb.save(f)
print("Capitulo 2 corregido. Hojas:", wb.sheetnames)

# reporte de calculo
def V(*args):
    r=0
    for a in args: r+=a
    return round(r,2)
print("2.1.1 =", round(9.85*0.60*21.95,2), "m3")
print("2.1.2 =", round(9.85*0.30*21.95,2), "m3")
print("2.1.3 =", round(0.4*0.45*19.8*2 + 0.4*0.45*7.3*5 + 2.30*0.90*2.30*4 + 2.15*0.90*2.15*4,2), "m3")
print("2.2.2 =", round(0.5*19.8*2 + 0.5*7.3*5 + 2.30*2.30*4 + 2.15*2.15*4,2), "m2")
print("2.2.3 zapatas =", round(2.30*0.40*2.30*4 + 2.15*0.40*2.15*4,2), "m3")
