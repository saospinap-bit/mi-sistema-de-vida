#!/usr/bin/env python3
"""Genera el AutoCAD de las vigas de carga VC1–VC7 como vigas completas.

Fuentes coordinadas:
- VIGAS CARGA.pdf: ejes, continuidad y trazado en planta.
- Diseño de vigas proyecto diseño DEF.xlsx: sección y refuerzo por grupo.
- Despiece-Viga-EJEMPLO.dxf: lenguaje gráfico del detalle.

Toda la geometría de modelspace está en milímetros a tamaño real 1:1. El DXF
contiene siete layouts A1, uno por grupo, con viewports bloqueados a 1:25.
"""
from __future__ import annotations

import math
from pathlib import Path

import ezdxf
from ezdxf.enums import TextEntityAlignment
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
SOURCE_PDF = BASE / "VIGAS CARGA.pdf"
SOURCE_XLSX = BASE / "Diseño de vigas proyecto diseño DEF.xlsx"
REFERENCE_DXF = BASE / "Despiece-Viga-EJEMPLO.dxf"
OUTPUT_DIR = BASE / "Planos-Autocad-Carga-Revision"
OUTPUT_DXF = OUTPUT_DIR / "DESPIECE-ACADEMICO-VIGAS-CARGA-VC1-VC7-REVISION.dxf"

PAPER_W = 841.0
PAPER_H = 594.0
SCALE = 25.0
PANEL_W = 19000.0
PANEL_H = 7000.0
PANEL_GAP = 1200.0
BEAM_H = 550.0
COLUMN_W = 600.0
COVER = 40.0
DB_LONG = 15.9
DATE_TEXT = "18-07-2026"

WHITE = 7
# El ejemplo usa linework monocromático; se conservan capas semánticas,
# pero todas plotean blanco/negro como en el DXF de referencia.
RED = YELLOW = GREEN = CYAN = BLUE = MAGENTA = GRAY = BROWN = WHITE

GROUP_GEOMETRY = {
    "VC1": {
        "grid_lines": ("15", "14"),
        "members": [{"axes": ("B", "C", "D"), "spans": (3750.0, 3750.0)}],
        "plan": ((0.0, 0.0), (3750.0, 0.0), (7500.0, 0.0)),
    },
    "VC2": {
        "grid_lines": ("13", "1"),
        "members": [{"axes": ("A", "B", "C", "D", "E"),
                     "spans": (4000.0, 3750.0, 3750.0, 4000.0)}],
        "plan": ((0.0, 0.0), (4000.0, 0.0), (7750.0, 0.0),
                 (11500.0, 0.0), (15500.0, 0.0)),
    },
    "VC3": {
        "grid_lines": ("12", "2"),
        "members": [{"axes": ("A", "B", "C", "D", "E"),
                     "spans": (4000.0, 3814.7739, 3814.7739, 4000.0)}],
        "plan": ((0.0, 0.0), (4000.0, 0.0), (7750.0, -700.0),
                 (11500.0, 0.0), (15500.0, 0.0)),
        "mirrored_plan_lines": ("12", "2"),
        "secondary_grids": (("F", 1630.0), ("F", 4850.0),
                            ("G", 10650.0), ("G", 13870.0)),
    },
    "VC4": {
        "grid_lines": ("11", "3"),
        "members": [
            {"axes": ("A", "B"), "spans": (3850.0,), "label": "TRAMO A–B"},
            {"axes": ("D", "E"), "spans": (3850.0,), "label": "TRAMO D–E"},
        ],
        "plan_segments": (
            ((0.0, 0.0), (3850.0, 0.0)),
            ((11650.0, 0.0), (15500.0, 0.0)),
        ),
    },
    "VC5": {
        "grid_lines": ("10", "4"),
        "members": [{"axes": ("A", "B", "C", "D", "E"),
                     "spans": (3850.0, 3900.0, 3900.0, 3850.0)}],
        "plan": ((0.0, 0.0), (3850.0, 0.0), (7750.0, 0.0),
                 (11650.0, 0.0), (15500.0, 0.0)),
    },
    "VC6": {
        "grid_lines": ("9", "5"),
        "members": [{"axes": ("A", "B", "C", "D", "E"),
                     "spans": (3850.0, 3900.0, 3900.0, 3850.0)}],
        "plan": ((0.0, 0.0), (3850.0, 0.0), (7750.0, 0.0),
                 (11650.0, 0.0), (15500.0, 0.0)),
    },
    "VC7": {
        "grid_lines": ("8", "6"),
        "members": [{"axes": ("A", "B", "C", "D", "E"),
                     "spans": (2650.0, 5100.0, 5100.0, 2650.0)}],
        "plan": ((0.0, 0.0), (2650.0, 0.0), (7750.0, 0.0),
                 (12850.0, 0.0), (15500.0, 0.0)),
    },
}


def load_design_rows() -> dict[str, dict]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    sheet = workbook["Vigas de Carga (7)"]
    headers = [cell.value for cell in sheet[3]]
    rows = {
        values[0]: dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=4, values_only=True)
    }
    expected = set(GROUP_GEOMETRY)
    if set(rows) != expected:
        raise ValueError(f"Grupos de carga inesperados: {sorted(rows)}")
    return rows


def validate_sources() -> None:
    for path in (SOURCE_PDF, SOURCE_XLSX, REFERENCE_DXF):
        if not path.exists():
            raise FileNotFoundError(path)
    reference = ezdxf.readfile(REFERENCE_DXF)
    required = {"LINE", "TEXT", "CIRCLE", "LWPOLYLINE", "DIMENSION"}
    actual = {entity.dxftype() for entity in reference.modelspace()}
    if not required.issubset(actual):
        raise ValueError("El DXF de ejemplo no contiene el detalle esperado")


def create_document() -> ezdxf.document.Drawing:
    doc = ezdxf.new("R2018", setup=True)
    doc.units = 4
    doc.header["$INSUNITS"] = 4
    doc.header["$LUNITS"] = 2
    doc.header["$LUPREC"] = 2
    doc.header["$MEASUREMENT"] = 1
    doc.header["$PSLTSCALE"] = 1
    layers = {
        "VIGA": WHITE,
        "COLUMNAS": CYAN,
        "EJES": GRAY,
        "REF_SUP_CONT": BLUE,
        "REF_SUP_BASTON": YELLOW,
        "REF_INF": RED,
        "REF_TORSION": BROWN,
        "ESTRIBOS": GREEN,
        "CORTES": MAGENTA,
        "COTAS": GREEN,
        "TEXTO": WHITE,
        "TRAZADO_PLANTA": BLUE,
        "MARCO": WHITE,
        "VIEWPORT": GRAY,
    }
    for name, color in layers.items():
        if name not in doc.layers:
            doc.layers.add(name, color=color)
    return doc


def add_text(space, text, point, height=90.0, layer="TEXTO", color=None,
             align=TextEntityAlignment.LEFT, rotation=0.0):
    attrs = {"height": height, "layer": layer, "rotation": rotation}
    if color is not None:
        attrs["color"] = color
    entity = space.add_text(str(text), dxfattribs=attrs)
    entity.set_placement(point, align=align)
    return entity


def line(space, start, end, layer="VIGA", color=None, lineweight=25,
         linetype=None):
    attrs = {"layer": layer, "lineweight": lineweight}
    if color is not None:
        attrs["color"] = color
    if linetype:
        attrs["linetype"] = linetype
    return space.add_line(start, end, dxfattribs=attrs)


def polyline(space, points, layer="VIGA", color=None, lineweight=25,
             close=False):
    attrs = {"layer": layer, "lineweight": lineweight}
    if color is not None:
        attrs["color"] = color
    return space.add_lwpolyline(points, close=close, dxfattribs=attrs)


def rectangle(space, x, y, width, height, layer="VIGA", color=None,
              lineweight=25):
    return polyline(
        space,
        [(x, y), (x + width, y), (x + width, y + height), (x, y + height)],
        layer,
        color,
        lineweight,
        True,
    )


def fmt_mm(value: float) -> str:
    rounded = round(value)
    if math.isclose(value, rounded, abs_tol=0.05):
        return str(int(rounded))
    return f"{value:.1f}".replace(".", ",")


def positions_between(start: float, end: float,
                      max_spacing: float) -> list[float]:
    if end <= start + 1e-6:
        return [start]
    intervals = max(1, math.ceil((end - start) / max_spacing))
    return [start + i * (end - start) / intervals for i in range(intervals + 1)]


def span_stirrups(left_axis: float, right_axis: float, s_end: float,
                  s_center: float) -> list[float]:
    left_face = left_axis + COLUMN_W / 2
    right_face = right_axis - COLUMN_W / 2
    clear = right_face - left_face
    zone = min(2 * BEAM_H, clear / 2)
    first = left_face + min(50.0, s_end)
    last = right_face - min(50.0, s_end)
    left_boundary = min(left_face + zone, last)
    right_boundary = max(right_face - zone, first)
    points: set[float] = set()
    for start, end, spacing in (
        (first, left_boundary, s_end),
        (left_boundary, right_boundary, s_center),
        (right_boundary, last, s_end),
    ):
        points.update(round(x, 6) for x in positions_between(start, end, spacing))
    return sorted(points)


def draw_dimension(space, x0: float, x1: float, y: float, label: str,
                   extension=120.0):
    """Crea una cota CAD real, como la entidad DIMENSION del ejemplo."""
    dimension = space.add_linear_dim(
        base=(x0, y),
        p1=(x0, y + extension),
        p2=(x1, y + extension),
        text=label,
        angle=0,
        dimstyle="EZDXF",
        override={
            "dimtxt": 85.0,
            "dimasz": 45.0,
            "dimexe": 70.0,
            "dimexo": 25.0,
            "dimtad": 1,
            "dimclrd": WHITE,
            "dimclre": WHITE,
            "dimclrt": WHITE,
        },
        dxfattribs={"layer": "COTAS"},
    )
    dimension.render()


def draw_axis(space, x: float, beam_y: float, label: str):
    line(space, (x, beam_y - 950), (x, beam_y + 1550),
         "EJES", GRAY, 18, "CENTER")
    center = (x, beam_y + 1250)
    space.add_circle(center, 170.0,
                     dxfattribs={"layer": "EJES", "color": GRAY,
                                 "lineweight": 25})
    add_text(space, label, center, 130, "EJES", GRAY,
             TextEntityAlignment.MIDDLE_CENTER)


def draw_cut_marker(space, x: float, beam_y: float, label: str):
    line(space, (x, beam_y - 110), (x, beam_y + BEAM_H + 110),
         "CORTES", MAGENTA, 22)
    for y in (beam_y - 110, beam_y + BEAM_H + 110):
        space.add_circle((x, y), 75.0,
                         dxfattribs={"layer": "CORTES", "color": MAGENTA,
                                     "lineweight": 22})
        add_text(space, label, (x, y), 70, "CORTES", MAGENTA,
                 TextEntityAlignment.MIDDLE_CENTER)


def draw_hooked_end(space, x: float, y: float, outward: int, upward: int,
                    layer: str, color: int):
    horizontal = 305.0
    tail = 195.0
    line(space, (x, y), (x + outward * horizontal, y),
         layer, color, 40)
    line(space, (x + outward * horizontal, y),
         (x + outward * horizontal, y + upward * tail),
         layer, color, 40)


def draw_external_baston(space, axis: float, y: float, toward: int,
                          lc: float, ldh: float, tail: float,
                          dint: float) -> float:
    """Dibuja un bastón extremo cuya trayectoria física mide Lc."""
    radius = dint / 2 + DB_LONG / 2
    arc_length = math.pi * radius / 2
    embed = ldh - tail - arc_length
    straight = lc - ldh
    if embed < -1e-6 or straight < -1e-6:
        raise ValueError("Geometría de bastón incompatible con Lc/ldh/gancho")
    embed = max(embed, 0.0)
    attrs = {"layer": "REF_SUP_BASTON", "color": YELLOW,
             "lineweight": 38}
    if toward > 0:
        tangent = axis - embed
        cutoff = axis + straight
        line(space, (tangent, y), (cutoff, y),
             "REF_SUP_BASTON", YELLOW, 38)
        center = (tangent, y - radius)
        space.add_arc(center, radius, 90.0, 180.0, dxfattribs=attrs)
        outer = (tangent - radius, y - radius)
    else:
        tangent = axis + embed
        cutoff = axis - straight
        line(space, (cutoff, y), (tangent, y),
             "REF_SUP_BASTON", YELLOW, 38)
        center = (tangent, y - radius)
        space.add_arc(center, radius, 0.0, 90.0, dxfattribs=attrs)
        outer = (tangent + radius, y - radius)
    line(space, outer, (outer[0], outer[1] - tail),
         "REF_SUP_BASTON", YELLOW, 38)
    return lc


def draw_internal_baston(space, axis: float, y: float, toward: int,
                          lc: float, ldh: float) -> float:
    """Dibuja un bastón interior cruzando el apoyo con desarrollo ldh."""
    span_part = lc - ldh
    if span_part < -1e-6:
        raise ValueError("Lc de bastón interior menor que ldh")
    if toward > 0:
        start, end = axis - ldh, axis + span_part
    else:
        start, end = axis - span_part, axis + ldh
    line(space, (start, y), (end, y),
         "REF_SUP_BASTON", YELLOW, 38)
    return abs(end - start)


def draw_longitudinal_reinforcement(space, axes: list[float], beam_y: float,
                                    row: dict) -> list[float]:
    x0, x1 = axes[0], axes[-1]
    top = beam_y + BEAM_H - 70
    bottom = beam_y + 70
    mid = beam_y + BEAM_H / 2
    line(space, (x0, top), (x1, top), "REF_SUP_CONT", BLUE, 40)
    draw_hooked_end(space, x0, top, -1, -1, "REF_SUP_CONT", BLUE)
    draw_hooked_end(space, x1, top, 1, -1, "REF_SUP_CONT", BLUE)
    line(space, (x0, bottom), (x1, bottom), "REF_INF", RED, 40)
    draw_hooked_end(space, x0, bottom, -1, 1, "REF_INF", RED)
    draw_hooked_end(space, x1, bottom, 1, 1, "REF_INF", RED)

    support_length = float(row["Lc sup. apoyo Ln máx. (m)"]) * 1000
    ldh = float(row["ldh Nº5 adopt. (mm)"])
    tail = float(row["Gancho 90° cola Nº5 (mm)"])
    dint = float(row["Gancho 90° Dint Nº5 (mm)"])
    baston_lengths: list[float] = []
    for index, axis in enumerate(axes):
        y = top - 105 - (index % 2) * 70
        if index == 0:
            baston_lengths.append(draw_external_baston(
                space, axis, y, 1, support_length, ldh, tail, dint
            ))
        elif index == len(axes) - 1:
            baston_lengths.append(draw_external_baston(
                space, axis, y, -1, support_length, ldh, tail, dint
            ))
        else:
            baston_lengths.append(draw_internal_baston(
                space, axis, y, -1, support_length, ldh
            ))
            baston_lengths.append(draw_internal_baston(
                space, axis, y - 55, 1, support_length, ldh
            ))

    n_torsion = int(row["Nº5 torsión dedicadas"] or 0)
    if n_torsion:
        line(space, (x0, mid), (x1, mid),
             "REF_TORSION", BROWN, 32)
    return baston_lengths


def draw_section(space, center: tuple[float, float], row: dict,
                 title: str, support: bool):
    cx, cy = center
    width = float(row["b adopt. (mm)"])
    height = float(row["h adopt. (mm)"])
    stirrup_db = float(row["db estribo (mm)"])
    n_top = int(row["Nº5 sup"]) if support else 2
    n_bottom = int(row["Nº5 inf"])
    n_torsion = int(row["Nº5 torsión dedicadas"] or 0)
    left, bottom = cx - width / 2, cy - height / 2
    rectangle(space, left, bottom, width, height, "VIGA", WHITE, 35)
    rectangle(space, left + COVER, bottom + COVER,
              width - 2 * COVER, height - 2 * COVER,
              "ESTRIBOS", GREEN, 25)
    if int(row["Ramas"]) == 4:
        for fraction in (1 / 3, 2 / 3):
            x = left + COVER + fraction * (width - 2 * COVER)
            line(space, (x, bottom + COVER), (x, bottom + height - COVER),
                 "ESTRIBOS", GREEN, 20)
    edge = COVER + stirrup_db + DB_LONG / 2

    def bar_row(count: int, y: float, layer: str, color: int):
        if count == 1:
            xs = [cx]
        else:
            xs = [left + edge + i * (width - 2 * edge) / (count - 1)
                  for i in range(count)]
        for x in xs:
            space.add_circle((x, y), DB_LONG / 2,
                             dxfattribs={"layer": layer, "color": color,
                                         "lineweight": 35})

    bar_row(n_top, bottom + height - edge, "REF_SUP_CONT", BLUE)
    bar_row(n_bottom, bottom + edge, "REF_INF", RED)
    if n_torsion:
        if n_torsion % 2:
            raise ValueError(f"Cantidad impar de barras de torsión: {n_torsion}")
        per_side = n_torsion // 2
        usable_bottom = bottom + edge + 60
        usable_top = bottom + height - edge - 60
        ys = ([cy] if per_side == 1 else
              [usable_bottom + i * (usable_top - usable_bottom) /
               (per_side - 1) for i in range(per_side)])
        for y in ys:
            for x in (left + edge, left + width - edge):
                space.add_circle((x, y), DB_LONG / 2,
                                 dxfattribs={"layer": "REF_TORSION",
                                             "color": BROWN,
                                             "lineweight": 30})
    add_text(space, title, (cx, bottom - 120), 90, "CORTES", MAGENTA,
             TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"{int(width)} × {int(height)} mm",
             (cx, bottom - 235), 80, "TEXTO", WHITE,
             TextEntityAlignment.MIDDLE_CENTER)


def interpolate_plan_y(plan, x: float) -> float:
    for (x0, y0), (x1, y1) in zip(plan, plan[1:]):
        if x0 - 1e-6 <= x <= x1 + 1e-6:
            factor = (x - x0) / (x1 - x0)
            return y0 + factor * (y1 - y0)
    raise ValueError(f"Cruce de retícula fuera del trazado: {x}")


def draw_plan_trace(space, geometry: dict, origin_y: float):
    add_text(space, "TRAZADO EN PLANTA SEGÚN PDF",
             (PANEL_W / 2, origin_y + 700), 100, "TRAZADO_PLANTA", BLUE,
             TextEntityAlignment.MIDDLE_CENTER)
    if "plan" in geometry:
        plan = geometry["plan"]
        total = max(point[0] for point in plan)
        offset = (PANEL_W - total) / 2
        axes = geometry["members"][0]["axes"]
        if "mirrored_plan_lines" in geometry:
            variants = (
                (geometry["mirrored_plan_lines"][0], 500.0, 1.0),
                (geometry["mirrored_plan_lines"][1], -1000.0, -1.0),
            )
            for line_label, shift, mirror in variants:
                points = [
                    (offset + x, origin_y + shift + mirror * y)
                    for x, y in plan
                ]
                polyline(space, points, "TRAZADO_PLANTA", BLUE, 35)
                add_text(space, f"EJE {line_label}",
                         (offset - 180, points[0][1]), 75,
                         "TRAZADO_PLANTA", BLUE,
                         TextEntityAlignment.MIDDLE_RIGHT)
                for (x, y), axis in zip(points, axes):
                    space.add_circle((x, y), 60.0,
                                     dxfattribs={"layer": "TRAZADO_PLANTA",
                                                 "color": BLUE,
                                                 "lineweight": 22})
                    label_y = y + (105 if shift > 0 else -105)
                    add_text(space, axis, (x, label_y), 65,
                             "TRAZADO_PLANTA", BLUE,
                             TextEntityAlignment.MIDDLE_CENTER)
                for grid_label, grid_x in geometry["secondary_grids"]:
                    base_y = interpolate_plan_y(plan, grid_x)
                    y = origin_y + shift + mirror * base_y
                    x = offset + grid_x
                    line(space, (x, y - 75), (x, y + 75),
                         "TRAZADO_PLANTA", BLUE, 18)
                    label_y = y + (125 if shift > 0 else -125)
                    add_text(space, grid_label, (x, label_y), 58,
                             "TRAZADO_PLANTA", BLUE,
                             TextEntityAlignment.MIDDLE_CENTER)
        else:
            points = [(offset + x, origin_y + y) for x, y in plan]
            polyline(space, points, "TRAZADO_PLANTA", BLUE, 35)
            for (x, y), axis in zip(points, axes):
                space.add_circle((x, y), 90.0,
                                 dxfattribs={"layer": "TRAZADO_PLANTA",
                                             "color": BLUE,
                                             "lineweight": 25})
                add_text(space, axis, (x, y + 160), 85,
                         "TRAZADO_PLANTA", BLUE,
                         TextEntityAlignment.MIDDLE_CENTER)
    else:
        all_points = [point for segment in geometry["plan_segments"]
                      for point in segment]
        total = max(point[0] for point in all_points)
        offset = (PANEL_W - total) / 2
        for segment, member in zip(geometry["plan_segments"],
                                   geometry["members"]):
            points = [(offset + x, origin_y + y) for x, y in segment]
            polyline(space, points, "TRAZADO_PLANTA", BLUE, 35)
            for (x, y), axis in zip(points, member["axes"]):
                space.add_circle((x, y), 90.0,
                                 dxfattribs={"layer": "TRAZADO_PLANTA",
                                             "color": BLUE,
                                             "lineweight": 25})
                add_text(space, axis, (x, y + 160), 85,
                         "TRAZADO_PLANTA", BLUE,
                         TextEntityAlignment.MIDDLE_CENTER)


def draw_member(space, member: dict, row: dict, beam_y: float,
                x_shift: float | None = None) -> dict:
    spans = list(member["spans"])
    total = sum(spans)
    x0 = (PANEL_W - total) / 2 if x_shift is None else x_shift
    axes = [x0]
    for span in spans:
        axes.append(axes[-1] + span)

    rectangle(space, axes[0] - COLUMN_W / 2, beam_y,
              total + COLUMN_W, BEAM_H, "VIGA", WHITE, 40)
    for x, label in zip(axes, member["axes"]):
        rectangle(space, x - COLUMN_W / 2, beam_y - 450,
                  COLUMN_W, BEAM_H + 900, "COLUMNAS", CYAN, 35)
        draw_axis(space, x, beam_y, label)

    s_end = float(row["s extremo DMO (mm; zona 2h)"])
    s_center = float(row["s centro (mm)"])
    all_stirrups: list[float] = []
    interval_audit: list[tuple[float, float]] = []
    for left, right in zip(axes, axes[1:]):
        positions = span_stirrups(left, right, s_end, s_center)
        all_stirrups.extend(positions)
        boundaries = [left + COLUMN_W / 2, *positions, right - COLUMN_W / 2]
        for start, end in zip(boundaries, boundaries[1:]):
            interval_audit.append((end - start, max(s_end, s_center)))
    for x in sorted(set(round(value, 6) for value in all_stirrups)):
        line(space, (x, beam_y + COVER),
             (x, beam_y + BEAM_H - COVER), "ESTRIBOS", GREEN, 15)

    baston_lengths = draw_longitudinal_reinforcement(space, axes, beam_y, row)
    for left, right, span in zip(axes, axes[1:], spans):
        draw_dimension(space, left, right, beam_y - 420,
                       f"{fmt_mm(span)} mm")
    draw_dimension(space, axes[0], axes[-1], beam_y - 760,
                   f"LONGITUD ENTRE EJES = {fmt_mm(total)} mm")

    first_clear_mid = (axes[0] + axes[1]) / 2
    draw_cut_marker(space, axes[0] + COLUMN_W / 2 + 120, beam_y, "A")
    draw_cut_marker(space, first_clear_mid, beam_y, "B")
    if member.get("label"):
        add_text(space, member["label"],
                 ((axes[0] + axes[-1]) / 2, beam_y + BEAM_H + 420),
                 105, "TEXTO", WHITE, TextEntityAlignment.MIDDLE_CENTER)
    return {
        "axes": axes,
        "axis_labels": member["axes"],
        "spans": spans,
        "total": total,
        "intervals": interval_audit,
        "baston_lengths": baston_lengths,
    }


def draw_group_panel(space, group: str, row: dict,
                     geometry: dict, origin_y: float) -> dict:
    local_start = len(space)
    lines = " Y ".join(geometry["grid_lines"])
    add_text(space, f"DESPIECE COMPLETO VIGA DE CARGA {group}",
             (PANEL_W / 2, origin_y + 6650), 190, "TEXTO", MAGENTA,
             TextEntityAlignment.MIDDLE_CENTER)
    add_text(
        space,
        f"APLICA EN EJES {lines} | Sección {int(row['b adopt. (mm)'])} × "
        f"{int(row['h adopt. (mm)'])} mm | {int(row['N vigas'])} vigas del grupo",
        (PANEL_W / 2, origin_y + 6400), 105, "TEXTO", WHITE,
        TextEntityAlignment.MIDDLE_CENTER,
    )
    check_vt = str(row["Chequeo V-T"] or "NO VALIDADO")
    status_text = (
        "DETALLE ACADÉMICO — NO EMITIR PARA CONSTRUCCIÓN | "
        f"Verificación V-T: {check_vt} | Estado global del Excel: {row['ESTADO']}"
    )
    add_text(space, status_text, (PANEL_W / 2, origin_y + 6200),
             92, "TEXTO", WHITE, TextEntityAlignment.MIDDLE_CENTER)
    draw_plan_trace(space, geometry, origin_y + 5300)

    members: list[dict] = []
    if group == "VC4":
        members.append(draw_member(space, geometry["members"][0], row,
                                   origin_y + 3500, 2500.0))
        members.append(draw_member(space, geometry["members"][1], row,
                                   origin_y + 3500, 12650.0))
    elif group == "VC3":
        members.append(draw_member(space, geometry["members"][0], row,
                                   origin_y + 2700))
    else:
        members.append(draw_member(space, geometry["members"][0], row,
                                   origin_y + 3300))

    section_y = origin_y + 900
    draw_section(space, (PANEL_W / 2 - 1300, section_y), row,
                 "CORTE A–A APOYO", True)
    draw_section(space, (PANEL_W / 2 + 1300, section_y), row,
                 "CORTE B–B CENTRO", False)

    n_sup = int(row["Nº5 sup"])
    n_inf = int(row["Nº5 inf"])
    n_tor = int(row["Nº5 torsión dedicadas"] or 0)
    s_end = int(row["s extremo DMO (mm; zona 2h)"])
    s_center = int(row["s centro (mm)"])
    stirrup = row["Estribo"]
    support_lc = float(row["Lc sup. apoyo Ln máx. (m)"]) * 1000
    add_text(
        space,
        f"SUPERIOR: {n_sup} Nº5 (2 continuas + {max(n_sup - 2, 0)} bastones); "
        f"Lc bastón desde cada apoyo = {fmt_mm(support_lc)} mm",
        (350.0, origin_y + 1450), 90, "REF_SUP_CONT", BLUE,
    )
    add_text(space, f"INFERIOR: {n_inf} Nº5 continuas", (350.0, origin_y + 1300),
             90, "REF_INF", RED)
    add_text(space, f"TORSIÓN: {n_tor} Nº5 dedicadas", (350.0, origin_y + 1150),
             90, "REF_TORSION", BROWN)
    add_text(
        space,
        f"ESTRIBOS: {stirrup}, {int(row['Ramas'])} ramas, primer estribo ≤50 mm; "
        f"@{s_end} mm en 2h desde cada apoyo y @{s_center} mm al centro",
        (350.0, origin_y + 1000), 82, "ESTRIBOS", GREEN,
    )
    add_text(
        space,
        f"ANCLAJES Nº5: ld inf/sup/ldh = {int(row['ld inf. Nº5 adopt. (mm)'])}/"
        f"{int(row['ld sup. Nº5 adopt. (mm)'])}/"
        f"{int(row['ldh Nº5 adopt. (mm)'])} mm; gancho 90° cola "
        f"{int(row['Gancho 90° cola Nº5 (mm)'])} mm",
        (350.0, origin_y + 850), 82, "TEXTO", WHITE,
    )
    add_text(
        space,
        "NOTA: el trazado y los ejes provienen de VIGAS CARGA.pdf; "
        "la elevación desarrolla la viga completa, usa apoyos nominales de 600 mm "
        "y marca cada tramo entre ejes.",
        (350.0, origin_y + 180), 76, "TEXTO", WHITE,
    )
    return {
        "group": group,
        "center": (PANEL_W / 2, origin_y + PANEL_H / 2),
        "members": members,
        "grid_lines": geometry["grid_lines"],
        "support_length": support_lc,
        "expected_torsion_circles": 2 * n_tor,
        "entities": list(space)[local_start:],
    }


def paper_text(layout, text, point, height=3.0,
               align=TextEntityAlignment.LEFT, color=WHITE):
    return add_text(layout, text, point, height, "TEXTO", color, align)


def add_layout(doc, panel: dict):
    name = f"{panel['group']}_CARGA"
    layout = doc.layouts.new(name)
    layout.page_setup(
        size=(PAPER_W, PAPER_H), margins=(0, 0, 0, 0), units="mm",
        rotation=0, scale=(1, 1), name="ISO A1 horizontal 841x594 mm",
        device="DWG To PDF.pc3",
    )
    rectangle(layout, 10, 10, PAPER_W - 20, PAPER_H - 20,
              "MARCO", WHITE, 40)
    rectangle(layout, 10, 10, PAPER_W - 20, 65,
              "MARCO", WHITE, 35)
    for x in (130, 500, 650, 745):
        line(layout, (x, 10), (x, 75), "MARCO", WHITE, 25)
    paper_text(layout, "UNIVERSIDAD NACIONAL DE COLOMBIA",
               (70, 55), 3.4, TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, "DISEÑO ESTRUCTURAL — GRUPO 6",
               (70, 34), 2.8, TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, f"DESPIECE COMPLETO {panel['group']} — VIGAS DE CARGA",
               (315, 50), 5.0, TextEntityAlignment.MIDDLE_CENTER, MAGENTA)
    paper_text(layout,
               f"Ejes de planta: {' / '.join(panel['grid_lines'])} | "
               "Geometría modelspace 1:1 mm",
               (315, 28), 3.0, TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, "REVISIÓN ACADÉMICA", (575, 51), 3.7,
               TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, "NO EMITIR PARA CONSTRUCCIÓN", (575, 29), 2.3,
               TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, "ESC. 1:25 — A1", (697.5, 51), 3.5,
               TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, "841 × 594 mm", (697.5, 29), 2.7,
               TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, DATE_TEXT, (793, 51), 3.0,
               TextEntityAlignment.MIDDLE_CENTER)
    paper_text(layout, panel["group"], (793, 29), 4.5,
               TextEntityAlignment.MIDDLE_CENTER)

    viewport = layout.add_viewport(
        center=(PAPER_W / 2, 340),
        size=(PAPER_W - 20, 320),
        view_center_point=panel["center"],
        view_height=320 * SCALE,
        status=2,
    )
    viewport.dxf.flags = int(viewport.dxf.flags or 0) | 16384
    return layout


def audit(doc, panels: list[dict]) -> None:
    if doc.units != 4:
        raise AssertionError("El DXF no está en milímetros")
    if tuple(panel["group"] for panel in panels) != tuple(GROUP_GEOMETRY):
        raise AssertionError("Los grupos no están en orden VC1–VC7")
    for panel in panels:
        layout = doc.layouts.get(f"{panel['group']}_CARGA")
        viewports = [vp for vp in layout.query("VIEWPORT")
                     if int(vp.dxf.id or 0) > 1]
        if len(viewports) != 1:
            raise AssertionError(f"Layout inválido para {panel['group']}")
        viewport = viewports[0]
        ratio = float(viewport.dxf.view_height) / float(viewport.dxf.height)
        if not math.isclose(ratio, SCALE, abs_tol=1e-9):
            raise AssertionError(f"Escala incorrecta en {panel['group']}: 1:{ratio}")
        if not (int(viewport.dxf.flags or 0) & 16384):
            raise AssertionError(f"Viewport sin bloquear en {panel['group']}")
        for member in panel["members"]:
            if not math.isclose(sum(member["spans"]), member["total"],
                                abs_tol=1e-6):
                raise AssertionError(f"Tramos incoherentes en {panel['group']}")
            for length in member["baston_lengths"]:
                if not math.isclose(length, panel["support_length"],
                                    abs_tol=1e-6):
                    raise AssertionError(
                        f"Lc física de bastón incorrecta en {panel['group']}: "
                        f"{length:.1f} mm"
                    )
            for gap, limit in member["intervals"]:
                if gap > limit + 1e-6:
                    raise AssertionError(
                        f"Estribos fuera de separación en {panel['group']}: "
                        f"{gap:.1f} > {limit:.1f} mm"
                    )
        torsion_circles = sum(
            1 for entity in panel["entities"]
            if entity.dxftype() == "CIRCLE"
            and entity.dxf.layer == "REF_TORSION"
        )
        if torsion_circles != panel["expected_torsion_circles"]:
            raise AssertionError(
                f"Barras de torsión incorrectas en {panel['group']}: "
                f"{torsion_circles} != {panel['expected_torsion_circles']}"
            )
        dimensions = [
            entity for entity in panel["entities"]
            if entity.dxftype() == "DIMENSION"
        ]
        expected_dimensions = sum(len(member["spans"]) + 1
                                  for member in panel["members"])
        if len(dimensions) != expected_dimensions:
            raise AssertionError(
                f"Cotas CAD incompletas en {panel['group']}: "
                f"{len(dimensions)} != {expected_dimensions}"
            )
    vc3 = GROUP_GEOMETRY["VC3"]["plan"]
    if math.isclose(vc3[2][1], 0.0, abs_tol=1e-9):
        raise AssertionError("Se perdió el quiebre de 700 mm de VC3")
    if len(panels[3]["members"]) != 2:
        raise AssertionError("VC4 debe conservar A–B y D–E independientes")


def main() -> None:
    validate_sources()
    rows = load_design_rows()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    doc = create_document()
    modelspace = doc.modelspace()
    panels = []
    for index, group in enumerate(GROUP_GEOMETRY):
        origin_y = -index * (PANEL_H + PANEL_GAP)
        panels.append(draw_group_panel(
            modelspace, group, rows[group], GROUP_GEOMETRY[group], origin_y
        ))
    for panel in panels:
        add_layout(doc, panel)
    if "Layout1" in doc.layout_names():
        doc.layouts.delete("Layout1")
    audit(doc, panels)
    doc.saveas(OUTPUT_DXF)
    reloaded = ezdxf.readfile(OUTPUT_DXF)
    if reloaded.units != 4 or not OUTPUT_DXF.read_bytes().startswith(b"  0"):
        raise AssertionError("El DXF guardado no superó la verificación final")
    print(f"OK: {OUTPUT_DXF}")
    print("VC1–VC7 completas, en orden; modelspace 1:1 mm; layouts A1 1:25")


if __name__ == "__main__":
    main()
