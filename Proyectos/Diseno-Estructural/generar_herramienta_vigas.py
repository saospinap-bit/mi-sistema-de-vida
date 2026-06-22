#!/usr/bin/env python3
"""
HERRAMIENTA-GUIA de diseno de elementos en concreto reforzado - NSR-10 / ACI 318 (SI).
Genera: Herramienta-Diseno-Viguetas-Vigas.xlsx
Hojas: Instrucciones | Materiales | Viguetas | Vigas | Riostras
Filosofia: el usuario SOLO mete f'c, fy, geometria y los MOMENTOS/FUERZAS (celdas amarillas).
La hoja calcula y VERIFICA todo: acero a flexion, No. de barras, estribos, torsion,
deflexiones y longitudes de desarrollo, con semaforo CUMPLE / NO CUMPLE.
Unidades: MPa, mm, kN, kN*m, mm2.
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

NAVY="1A237E"; GRIS="37474F"; CLARO="ECEFF1"
AZUL_IN="BBDEFB"; AMA_IN="FFF59D"; VERDE="C8E6C9"; ROJO="FFCDD2"; AMAR="FFF9C4"
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def titulo(ws,a,t,span,color=NAVY):
    ws.merge_cells(f"{a}:{span}"); c=ws[a]; c.value=t
    c.font=Font(bold=True,size=13,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER

def seccion(ws,row,txt,color=GRIS,span=4):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    c=ws.cell(row=row,column=1,value=txt); c.font=Font(bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=LEFT

def inp(ws,row,label,val,unit="",nota="",dato=AZUL_IN,fmt=None):
    ws.cell(row=row,column=1,value=label).alignment=LEFT
    c=ws.cell(row=row,column=2,value=val); c.alignment=CENTER; c.fill=PatternFill("solid",fgColor=dato)
    if fmt: c.number_format=fmt
    ws.cell(row=row,column=3,value=unit).alignment=CENTER
    ws.cell(row=row,column=4,value=nota).alignment=LEFT
    for j in range(1,5): ws.cell(row=row,column=j).border=BORDER

def out(ws,row,label,formula,unit="",fmt="0.00",bold=False,hl=None):
    ws.cell(row=row,column=1,value=label).alignment=LEFT
    c=ws.cell(row=row,column=2,value=formula); c.alignment=CENTER; c.number_format=fmt
    if bold: c.font=Font(bold=True)
    if hl: c.fill=PatternFill("solid",fgColor=hl)
    ws.cell(row=row,column=3,value=unit).alignment=CENTER
    for j in range(1,5): ws.cell(row=row,column=j).border=BORDER

def nota(ws,row,txt):
    ws.cell(row=row,column=1,value=txt).font=Font(italic=True,size=9)

def anchos(ws):
    ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=16
    ws.column_dimensions["C"].width=10; ws.column_dimensions["D"].width=48

def semaforo(ws,rng):
    # verde = bueno ; rojo = malo
    for tok in ['"CUMPLE"','"DESPRECIABLE"','"NO CALCULAR"','"OK"']:
        ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=[tok],fill=PatternFill("solid",fgColor=VERDE)))
    for tok in ['"NO CUMPLE"','"REVISAR"','"DISENAR"','"CALCULAR"']:
        ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=[tok],fill=PatternFill("solid",fgColor=ROJO)))

# ============================================================ INSTRUCCIONES
def hoja_instrucciones(wb):
    ws=wb.active; ws.title="Instrucciones"; anchos(ws)
    titulo(ws,"A1","COMO USAR ESTA HERRAMIENTA","D1")
    pasos=[
        ("1","Ve a la hoja del elemento: Viguetas, Vigas o Riostras."),
        ("2","Llena las celdas AZULES (materiales y geometria): f'c, fy, b, h, recubrimiento y diametros de barra."),
        ("3","Llena las celdas AMARILLAS con los MOMENTOS y FUERZAS de tu modelo (Mu+, Mu-, Vu, Tu)."),
        ("4","Listo: la hoja calcula As, numero de barras, estribos, torsion, deflexiones y longitudes."),
        ("5","Mira la columna de CHEQUEOS: verde = CUMPLE, rojo = NO CUMPLE / REVISAR."),
        ("6","Si algo sale rojo, aumenta seccion (h o b), f'c, o el acero, hasta que quede verde."),
    ]
    r=3
    for n,t in pasos:
        ws.cell(row=r,column=1,value=n).font=Font(bold=True); ws.cell(row=r,column=1).alignment=CENTER
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
        ws.cell(row=r,column=2,value=t).alignment=LEFT
        for j in range(1,5): ws.cell(row=r,column=j).border=BORDER
        r+=1
    seccion(ws,r+1,"LEYENDA DE COLORES")
    leg=[("Celda AZUL","Dato de entrada: materiales y geometria",AZUL_IN),
         ("Celda AMARILLA","Dato de entrada: MOMENTOS y FUERZAS (de tu modelo SAP/ETABS)",AMA_IN),
         ("Verde","El chequeo CUMPLE",VERDE),
         ("Rojo","El chequeo NO CUMPLE / hay que revisar",ROJO)]
    rr=r+2
    for a,b,col in leg:
        c=ws.cell(row=rr,column=1,value=a); c.fill=PatternFill("solid",fgColor=col); c.border=BORDER; c.alignment=CENTER
        ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4)
        ws.cell(row=rr,column=2,value=b).alignment=LEFT; ws.cell(row=rr,column=2).border=BORDER
        rr+=1
    nota(ws,rr+1,"Base normativa: NSR-10 (Titulos B cargas, A sismo, C concreto) / ACI 318. Unidades SI: MPa, mm, kN.")

# ============================================================ MATERIALES (tabla de barras)
def hoja_materiales(wb):
    ws=wb.create_sheet("Materiales"); anchos(ws)
    titulo(ws,"A1","TABLA DE BARRAS DE REFUERZO (referencia)","D1")
    for j,t in enumerate(["Barra No.","Diametro db (mm)","Area (mm2)","Uso tipico"],1):
        c=ws.cell(row=2,column=j,value=t); c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=GRIS); c.alignment=CENTER; c.border=BORDER
    barras=[("No.2",6.4,32,"Estribos de viguetas"),("No.3",9.5,71,"Estribos / refuerzo menor"),
            ("No.4",12.7,129,"Refuerzo longitudinal"),("No.5",15.9,199,"Refuerzo longitudinal"),
            ("No.6",19.1,284,"Vigas / columnas"),("No.7",22.2,387,"Vigas / columnas"),
            ("No.8",25.4,510,"Columnas")]
    r=3
    for n,d,a,u in barras:
        ws.cell(row=r,column=1,value=n); ws.cell(row=r,column=2,value=d); ws.cell(row=r,column=3,value=a); ws.cell(row=r,column=4,value=u)
        for j in range(1,5): ws.cell(row=r,column=j).border=BORDER; ws.cell(row=r,column=j).alignment=CENTER if j<4 else LEFT
        r+=1
    nota(ws,r+1,"Areas nominales. En las hojas el area se calcula como pi/4*db^2 a partir del diametro que elijas.")

# ----- bloque comun de geometria/flexion/cortante -----
def bloque_comun(ws, T_section=False, vigueta=False):
    """Construye un elemento. Devuelve dict de filas clave. Layout fijo por hoja."""
    seccion(ws,3,"DATOS DE ENTRADA  (azul = materiales/geometria, amarillo = fuerzas)")
    inp(ws,4,"f'c",21,"MPa")
    inp(ws,5,"fy (longitudinal)",420,"MPa")
    inp(ws,6,"fyt (transversal)",420,"MPa")
    inp(ws,7,"b / bw (ancho)",100 if vigueta else 300,"mm")
    inp(ws,8,"h (altura total)",400 if vigueta else 500,"mm")
    inp(ws,9,"recubrimiento libre cc",25 if vigueta else 40,"mm")
    inp(ws,10,"db estribo",6.4 if vigueta else 9.5,"mm","No.2=6.4 / No.3=9.5")
    inp(ws,11,"db barra longitudinal",15.9,"mm","No.4=12.7 / No.5=15.9 / No.6=19.1")
    if T_section:
        inp(ws,12,"hf (espesor loseta)",50,"mm")
        inp(ws,13,"bf (ancho efectivo ala)",850,"mm","min(L/4, bw+16hf, sep nervios)")
    else:
        inp(ws,12,"hf (no aplica)",0,"mm","0 para seccion rectangular")
        inp(ws,13,"bf (= b)","=B7","mm")
    inp(ws,14,"L (luz libre)",5000,"mm")
    inp(ws,15,"Factor apoyo",18.5,"-","16 simple / 18.5 un extr. cont. / 21 ambos / 8 voladizo")
    inp(ws,16,"Mu+  (momento positivo)",25 if vigueta else 120,"kN*m","<-- DATO de tu modelo",dato=AMA_IN)
    inp(ws,17,"Mu-  (momento negativo)",30 if vigueta else 150,"kN*m","<-- DATO de tu modelo",dato=AMA_IN)
    inp(ws,18,"Vu  (cortante)",35 if vigueta else 140,"kN","<-- DATO de tu modelo",dato=AMA_IN)
    inp(ws,19,"Tu  (torsion)",0.3 if vigueta else 12,"kN*m","<-- DATO de tu modelo",dato=AMA_IN)
    inp(ws,20,"phi flexion",0.9,"-")
    inp(ws,21,"phi cortante/torsion",0.75,"-")

    seccion(ws,23,"GEOMETRIA Y MATERIALES CALCULADOS")
    out(ws,24,"d (altura efectiva) = h-cc-db_estr-db_long/2","=B8-B9-B10-B11/2","mm","0.0")
    out(ws,25,"Area 1 barra long = pi/4*db^2","=PI()/4*B11^2","mm2","0")
    out(ws,26,"Av estribo (2 ramas) = 2*pi/4*db_estr^2","=2*PI()/4*B10^2","mm2","0")
    out(ws,27,"beta1","=IF(B4<=28,0.85,MAX(0.65,0.85-0.05*(B4-28)/7))","-","0.000")
    out(ws,28,"As minimo = max(1.4/fy ; 0.25raizf'c/fy)*b*d","=MAX(1.4/B5,0.25*SQRT(B4)/B5)*B7*B24","mm2","0")
    out(ws,29,"rho maximo (ductil et=0.005)","=0.85*B27*B4/B5*(3/8)","-","0.00000")

def flexion(ws, r0, signo, Mcell, ancho_cell, label):
    """Bloque de flexion. r0 fila inicial de la seccion."""
    seccion(ws,r0,f"FLEXION {signo}  ({label})")
    out(ws,r0+1,"Rn = Mu/(phi*ancho*d^2)",f"={Mcell}*10^6/(B20*{ancho_cell}*B24^2)","MPa","0.000")
    out(ws,r0+2,"rho",f"=(0.85*B4/B5)*(1-SQRT(1-2*B{r0+1}/(0.85*B4)))","-","0.00000")
    out(ws,r0+3,"As requerido",f"=B{r0+2}*{ancho_cell}*B24","mm2","0")
    out(ws,r0+4,"As ADOPTADO = max(req, As_min)",f"=MAX(B{r0+3},B28)","mm2","0",bold=True,hl=VERDE)
    out(ws,r0+5,"a (profundidad bloque)",f"=B{r0+4}*B5/(0.85*B4*{ancho_cell})","mm","0.0")
    out(ws,r0+6,"Chequeo ductilidad (rho<=rho_max)",f'=IF(B{r0+2}<=B29,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    out(ws,r0+7,"No. de barras",f"=ROUNDUP(B{r0+4}/B25,0)","barras","0",bold=True,hl=AMAR)
    out(ws,r0+8,"As provisto",f"=B{r0+7}*B25","mm2","0")
    out(ws,r0+9,"Chequeo acero (As_prov>=As_adopt)",f'=IF(B{r0+8}>=B{r0+4},"CUMPLE","NO CUMPLE")',"","General",bold=True)
    return r0+9

# ============================================================ VIGUETAS
def hoja_viguetas(wb):
    ws=wb.create_sheet("Viguetas"); anchos(ws)
    titulo(ws,"A1","DISENO DE VIGUETA (seccion T) - NSR-10","D1")
    bloque_comun(ws, T_section=True, vigueta=True)
    # Flexion + usa ala bf (B13); Flexion - usa bw (B7)
    flexion(ws,31,"POSITIVA (+)","B16","B13","abajo, comprime la loseta -> ancho bf")
    flexion(ws,42,"NEGATIVA (-)","B17","B7","arriba en apoyos -> ancho bw")
    # Cortante (vigueta admite 1.1 Vc)
    seccion(ws,53,"CORTANTE (viguetas: NSR-10 C.8.13 permite 1.1*Vc)")
    out(ws,54,"Vc = 0.17*raizf'c*bw*d","=0.17*SQRT(B4)*B7*B24/1000","kN","0.00")
    out(ws,55,"1.1*Vc","=1.1*B54","kN","0.00")
    out(ws,56,"phi*1.1Vc","=B21*B55","kN","0.00")
    out(ws,57,"Vs requerido = Vu/phi - 1.1Vc","=MAX(B18/B21-B55,0)","kN","0.00")
    out(ws,58,"s requerido = Av*fyt*d/Vs","=IF(B57<=0,9999,B26*B6*B24/(B57*1000))","mm","0")
    out(ws,59,"s maximo = min(d/2,600)","=MIN(B24/2,600)","mm","0")
    out(ws,60,"s ADOPTADO","=IF(B18<=B56,B59,MIN(B58,B59))","mm","0",bold=True,hl=AMAR)
    out(ws,61,"RESULTADO","=IF(B18<=B56,\"Concreto resiste: estribos de montaje @ \"&B59&\" mm\",\"Estribos #2 @ \"&ROUND(B60,0)&\" mm\")","","General")
    out(ws,62,"phi*Vn con s adoptado","=B21*(B55+B26*B6*B24/B60/1000)","kN","0.00")
    out(ws,63,"Chequeo cortante (Vu<=phiVn)",'=IF(B18<=B62,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    # Torsion
    seccion(ws,65,"TORSION (umbral)")
    out(ws,66,"Acp=bw*h","=B7*B8","mm2","0")
    out(ws,67,"pcp=2(bw+h)","=2*(B7+B8)","mm","0")
    out(ws,68,"phi*Tth = phi*0.083raizf'c*Acp^2/pcp","=B21*0.083*SQRT(B4)*(B66^2/B67)/10^6","kN*m","0.000")
    out(ws,69,"Chequeo torsion",'=IF(B19<B68,"DESPRECIABLE","DISENAR")',"","General",bold=True)
    # Deflexiones
    seccion(ws,71,"DEFLEXIONES (NSR-10 C.9.5)")
    out(ws,72,"h minimo = L/factor","=B14/B15","mm","0")
    out(ws,73,"Chequeo deflexion",'=IF(B8>=B72,"NO CALCULAR","CALCULAR")',"","General",bold=True)
    # Longitudes de desarrollo
    seccion(ws,75,"LONGITUDES DE DESARROLLO Y GANCHOS (barras <=No.6)")
    out(ws,76,"ld inferior = (fy/(2.1*raizf'c))*db","=(B5/(2.1*SQRT(B4)))*B11","mm","0")
    out(ws,77,"ld superior (psi_t=1.3)","=1.3*B76","mm","0")
    out(ws,78,"gancho ldh = (0.24*fy/raizf'c)*db","=MAX(0.24*B5/SQRT(B4)*B11,8*B11,150)","mm","0")
    semaforo(ws,"B31:B80")
    nota(ws,82,"As+ usa bf (ala a compresion); As- usa solo bw. Verifica que 'a' <= hf para que valga la formula rectangular.")

# ============================================================ VIGAS
def hoja_vigas(wb):
    ws=wb.create_sheet("Vigas"); anchos(ws)
    titulo(ws,"A1","DISENO DE VIGA rectangular - flexion, cortante y torsion - NSR-10","D1")
    bloque_comun(ws, T_section=False, vigueta=False)
    flexion(ws,31,"POSITIVA (+)","B16","B7","centro de luz")
    flexion(ws,42,"NEGATIVA (-)","B17","B7","apoyos")
    seccion(ws,53,"CORTANTE (vigas: casi siempre requieren estribos)")
    out(ws,54,"Vc = 0.17*raizf'c*b*d","=0.17*SQRT(B4)*B7*B24/1000","kN","0.00")
    out(ws,55,"phi*Vc","=B21*B54","kN","0.00")
    out(ws,56,"phi*Vc/2","=B21*B54/2","kN","0.00")
    out(ws,57,"Vs requerido = Vu/phi - Vc","=MAX(B18/B21-B54,0)","kN","0.00")
    out(ws,58,"Limite 0.33raizf'c*b*d","=0.33*SQRT(B4)*B7*B24/1000","kN","0.00")
    out(ws,59,"s max (segun Vs)","=IF(B57<=B58,MIN(B24/2,600),MIN(B24/4,300))","mm","0")
    out(ws,60,"s por resistencia = Av*fyt*d/Vs","=IF(B57<=0,9999,B26*B6*B24/(B57*1000))","mm","0")
    out(ws,61,"s por minimo (Av,min)","=B26*B6/MAX(0.062*SQRT(B4)*B7,0.35*B7)","mm","0")
    out(ws,62,"s ADOPTADO","=IF(B18<=B56,B59,MIN(B59,B60,B61))","mm","0",bold=True,hl=AMAR)
    out(ws,63,"RESULTADO","=IF(B18<=B56,\"No requiere estribos por resistencia\",\"Estribos @ \"&ROUND(B62,0)&\" mm\")","","General")
    out(ws,64,"phi*Vn con s adoptado","=B21*(B54+B26*B6*B24/B62/1000)","kN","0.00")
    out(ws,65,"Chequeo cortante (Vu<=phiVn)",'=IF(B18<=B64,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    seccion(ws,67,"TORSION (diseno si Tu>=phi*Tth)")
    out(ws,68,"Acp=b*h","=B7*B8","mm2","0")
    out(ws,69,"pcp=2(b+h)","=2*(B7+B8)","mm","0")
    out(ws,70,"phi*Tth (umbral)","=B21*0.083*SQRT(B4)*(B68^2/B69)/10^6","kN*m","0.000")
    out(ws,71,"Aoh=(b-2cc)(h-2cc)","=(B7-2*B9)*(B8-2*B9)","mm2","0")
    out(ws,72,"ph=2((b-2cc)+(h-2cc))","=2*((B7-2*B9)+(B8-2*B9))","mm","0")
    out(ws,73,"Ao=0.85*Aoh","=0.85*B71","mm2","0")
    out(ws,74,"At/s (1 rama) = Tu/(phi*2*Ao*fyt)","=IF(B19<B70,0,B19*10^6/(B21*2*B73*B6))","mm2/mm","0.000")
    out(ws,75,"Al (acero long. torsion)=(At/s)*ph*(fyt/fy)","=B74*B72*(B6/B5)","mm2","0")
    out(ws,76,"Chequeo torsion",'=IF(B19<B70,"DESPRECIABLE","DISENAR")',"","General",bold=True)
    seccion(ws,78,"DEFLEXIONES (C.9.5)")
    out(ws,79,"h minimo = L/factor","=B14/B15","mm","0")
    out(ws,80,"Chequeo deflexion",'=IF(B8>=B79,"NO CALCULAR","CALCULAR")',"","General",bold=True)
    seccion(ws,82,"LONGITUDES DE DESARROLLO Y EMPALMES")
    out(ws,83,"ld inferior=(fy/(1.7raizf'c))*db (>=No.7) o (fy/(2.1raizf'c))*db",
        "=IF(B11>=19.1,(B5/(1.7*SQRT(B4)))*B11,(B5/(2.1*SQRT(B4)))*B11)","mm","0")
    out(ws,84,"ld superior (psi_t=1.3)","=1.3*B83","mm","0")
    out(ws,85,"Empalme clase B = 1.3*ld","=1.3*B83","mm","0")
    out(ws,86,"gancho ldh","=MAX(0.24*B5/SQRT(B4)*B11,8*B11,150)","mm","0")
    semaforo(ws,"B31:B90")
    nota(ws,88,"Estribo total por rama = (cortante: 0.5*Av/s) + (torsion: At/s). Si DMO/DES, verifica cortante por capacidad (Mpr).")

# ============================================================ RIOSTRAS
def hoja_riostras(wb):
    ws=wb.create_sheet("Riostras"); anchos(ws)
    titulo(ws,"A1","DISENO DE RIOSTRA - flexion y cortante simple - NSR-10","D1")
    seccion(ws,3,"DATOS DE ENTRADA")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa"); inp(ws,6,"fyt",420,"MPa")
    inp(ws,7,"b",100,"mm"); inp(ws,8,"h",400,"mm"); inp(ws,9,"cc",25,"mm")
    inp(ws,10,"db estribo",6.4,"mm"); inp(ws,11,"db long",12.7,"mm")
    inp(ws,12,"Mu",18,"kN*m","<-- DATO",dato=AMA_IN); inp(ws,13,"Vu",20,"kN","<-- DATO",dato=AMA_IN)
    inp(ws,14,"phi flexion",0.9,"-"); inp(ws,15,"phi cortante",0.75,"-")
    seccion(ws,17,"GEOMETRIA CALCULADA")
    out(ws,18,"d","=B8-B9-B10-B11/2","mm","0.0")
    out(ws,19,"Area barra","=PI()/4*B11^2","mm2","0")
    out(ws,20,"Av (2 ramas)","=2*PI()/4*B10^2","mm2","0")
    out(ws,21,"As minimo","=MAX(1.4/B5,0.25*SQRT(B4)/B5)*B7*B18","mm2","0")
    seccion(ws,23,"FLEXION")
    out(ws,24,"Rn","=B12*10^6/(B14*B7*B18^2)","MPa","0.000")
    out(ws,25,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B24/(0.85*B4)))","-","0.00000")
    out(ws,26,"As requerido","=B25*B7*B18","mm2","0")
    out(ws,27,"As ADOPTADO","=MAX(B26,B21)","mm2","0",bold=True,hl=VERDE)
    out(ws,28,"No. de barras","=ROUNDUP(B27/B19,0)","barras","0",bold=True,hl=AMAR)
    out(ws,29,"As provisto","=B28*B19","mm2","0")
    out(ws,30,"Chequeo acero",'=IF(B29>=B27,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    seccion(ws,32,"CORTANTE")
    out(ws,33,"Vc","=0.17*SQRT(B4)*B7*B18/1000","kN","0.00")
    out(ws,34,"phi*Vc","=B15*B33","kN","0.00")
    out(ws,35,"Vs requerido","=MAX(B13/B15-B33,0)","kN","0.00")
    out(ws,36,"s calculado","=IF(B35<=0,9999,B20*B6*B18/(B35*1000))","mm","0")
    out(ws,37,"s ADOPTADO","=IF(B13<=B34,MIN(B18/2,600),MIN(B36,B18/2,600))","mm","0",bold=True,hl=AMAR)
    out(ws,38,"RESULTADO",'=IF(B13<=B34,"Concreto resiste","Estribos @ "&ROUND(B37,0)&" mm")',"","General")
    semaforo(ws,"B17:B40")

# ============================================================ LOSA
def hoja_losa(wb):
    ws=wb.create_sheet("Losa"); anchos(ws)
    titulo(ws,"A1","LOSA NERVADA - espesor y acero de loseta - NSR-10","D1")
    seccion(ws,3,"DATOS DE ENTRADA")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa")
    inp(ws,6,"Ln (luz libre mayor)",5000,"mm")
    inp(ws,7,"Factor apoyo",18.5,"-","16 simple / 18.5 un extr / 21 ambos")
    inp(ws,8,"h losa total",400,"mm"); inp(ws,9,"hf loseta",50,"mm")
    inp(ws,10,"cc loseta",20,"mm"); inp(ws,11,"db loseta",6.4,"mm","malla / No.2")
    seccion(ws,13,"1) ESPESOR (C.9.5)")
    out(ws,14,"h minimo = Ln/factor","=B6/B7","mm","0")
    out(ws,15,"Chequeo espesor",'=IF(B8>=B14,"CUMPLE","REVISAR")',"","General",bold=True)
    seccion(ws,17,"2) ACERO DE RETRACCION Y TEMPERATURA (loseta, C.7.12)")
    out(ws,18,"As temp = 0.0018*1000*hf","=0.0018*1000*B9","mm2/m","0")
    out(ws,19,"Area 1 barra","=PI()/4*B11^2","mm2","0")
    out(ws,20,"separacion = 1000*Abarra/As","=1000*B19/B18","mm","0")
    out(ws,21,"s maximo = min(5hf,450)","=MIN(5*B9,450)","mm","0")
    out(ws,22,"s ADOPTADO","=MIN(B20,B21)","mm","0",bold=True,hl=AMAR)
    out(ws,23,"RESULTADO",'="Malla loseta @ "&ROUND(B22,0)&" mm en ambas direcciones"',"","General")
    nota(ws,25,"Los nervios (viguetas) se disenan en la hoja 'Viguetas'. Aqui solo loseta superior.")
    semaforo(ws,"B13:B25")

# ============================================================ ESCALERAS
def hoja_escaleras(wb):
    ws=wb.create_sheet("Escaleras"); anchos(ws)
    titulo(ws,"A1","ESCALERA (losa inclinada de 1 direccion) - NSR-10","D1")
    seccion(ws,3,"DATOS DE ENTRADA  (azul=geometria, amarillo=cargas/momento)")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa")
    inp(ws,6,"Ln (luz horizontal)",3.5,"m")
    inp(ws,7,"t (espesor garganta)",150,"mm")
    inp(ws,8,"CH (contrahuella)",175,"mm"); inp(ws,9,"H (huella)",280,"mm")
    inp(ws,10,"acabados",1.5,"kN/m2",dato=AMA_IN)
    inp(ws,11,"carga viva",3.0,"kN/m2","escaleras residencial",dato=AMA_IN)
    inp(ws,12,"cc",25,"mm"); inp(ws,13,"db",12.7,"mm")
    inp(ws,14,"factor momento (8 o 10)",8,"-")
    inp(ws,15,"Mu manual (opcional)",0,"kN*m","si lo pones, manda sobre el calculado",dato=AMA_IN)
    seccion(ws,17,"1) CARGAS")
    out(ws,18,"angulo theta","=DEGREES(ATAN(B8/B9))","grados","0.0")
    out(ws,19,"peso losa inclinada = 24*t/cos(theta)","=24*B7/1000/COS(ATAN(B8/B9))","kN/m2","0.00")
    out(ws,20,"peso escalones = 24*(CH/2)","=24*(B8/1000)/2","kN/m2","0.00")
    out(ws,21,"wD total = losa+escalones+acabados","=B19+B20+B10","kN/m2","0.00")
    out(ws,22,"wu = 1.2D + 1.6L","=1.2*B21+1.6*B11","kN/m2","0.00")
    out(ws,23,"Mu calculado = wu*Ln^2/factor","=B22*B6^2/B14","kN*m/m","0.00")
    out(ws,24,"Mu DE DISENO","=IF(B15>0,B15,B23)","kN*m/m","0.00",bold=True,hl=AMAR)
    seccion(ws,26,"2) FLEXION (por metro de ancho)")
    out(ws,27,"d = t-cc-db/2","=B7-B12-B13/2","mm","0.0")
    out(ws,28,"Rn","=B24*10^6/(0.9*1000*B27^2)","MPa","0.000")
    out(ws,29,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B28/(0.85*B4)))","-","0.00000")
    out(ws,30,"As req","=B29*1000*B27","mm2/m","0")
    out(ws,31,"As min (temp 0.0018)","=0.0018*1000*B7","mm2/m","0")
    out(ws,32,"As ADOPTADO","=MAX(B30,B31)","mm2/m","0",bold=True,hl=VERDE)
    out(ws,33,"sep = 1000*Abarra/As","=1000*(PI()/4*B13^2)/B32","mm","0")
    out(ws,34,"s max = min(3t,450)","=MIN(3*B7,450)","mm","0")
    out(ws,35,"RESULTADO",'="Barras No."&ROUND(B13/3.175,0)&" @ "&ROUND(MIN(B33,B34),0)&" mm"',"","General",bold=True,hl=AMAR)
    seccion(ws,37,"3) ESPESOR")
    out(ws,38,"t min = Ln/20","=B6*1000/20","mm","0")
    out(ws,39,"Chequeo espesor",'=IF(B7>=B38,"CUMPLE","REVISAR")',"","General",bold=True)
    semaforo(ws,"B17:B40")

# ============================================================ COLUMNAS
def hoja_columnas(wb):
    ws=wb.create_sheet("Columnas"); anchos(ws)
    titulo(ws,"A1","COLUMNA (carga axial + momento) - NSR-10","D1")
    seccion(ws,3,"DATOS DE ENTRADA  (azul=geometria, amarillo=fuerzas)")
    inp(ws,4,"f'c",28,"MPa"); inp(ws,5,"fy",420,"MPa")
    inp(ws,6,"b",400,"mm"); inp(ws,7,"h",400,"mm")
    inp(ws,8,"cc",40,"mm"); inp(ws,9,"db longitudinal",19.1,"mm","No.6"); inp(ws,10,"db estribo",9.5,"mm","No.3")
    inp(ws,11,"Lu (altura libre)",2600,"mm"); inp(ws,12,"k (factor longitud)",1.0,"-")
    inp(ws,13,"rho (cuantia, 0.01-0.06)",0.02,"-")
    inp(ws,14,"Pu (carga axial)",1500,"kN",dato=AMA_IN)
    inp(ws,15,"Mux",80,"kN*m",dato=AMA_IN); inp(ws,16,"Muy",40,"kN*m",dato=AMA_IN)
    inp(ws,17,"phi compresion (estribos)",0.65,"-")
    seccion(ws,19,"1) ACERO LONGITUDINAL")
    out(ws,20,"Ag = b*h","=B6*B7","mm2","0")
    out(ws,21,"Ast = rho*Ag","=B13*B20","mm2","0")
    out(ws,22,"Chequeo cuantia (0.01<=rho<=0.06)",'=IF(AND(B13>=0.01,B13<=0.06),"CUMPLE","NO CUMPLE")',"","General",bold=True)
    out(ws,23,"No. de barras","=ROUNDUP(B21/(PI()/4*B9^2),0)","barras","0",bold=True,hl=AMAR)
    out(ws,24,"As provisto","=B23*(PI()/4*B9^2)","mm2","0")
    seccion(ws,26,"2) CAPACIDAD AXIAL (cota superior, e=0)")
    out(ws,27,"phi*Pn,max = phi*0.80*(0.85f'c(Ag-Ast)+fy*Ast)","=B17*0.80*(0.85*B4*(B20-B24)+B5*B24)/1000","kN","0.0")
    out(ws,28,"Chequeo axial (Pu<=phiPnmax)",'=IF(B14<=B27,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    out(ws,29,"excentricidad ex = Mux/Pu","=B15/B14*1000","mm","0")
    out(ws,30,"excentricidad minima = 15+0.03h","=15+0.03*B7","mm","0")
    seccion(ws,32,"3) ESBELTEZ (NSR-10 C.10.10)")
    out(ws,33,"r = 0.30*h","=0.30*B7","mm","0")
    out(ws,34,"k*Lu/r","=B12*B11/B33","-","0.0")
    out(ws,35,"Chequeo esbeltez",'=IF(B34<=22,"CUMPLE","CALCULAR")',"","General",bold=True)
    seccion(ws,37,"4) ESTRIBOS")
    out(ws,38,"s max = min(16db, 48db_estr, b, h)","=MIN(16*B9,48*B10,B6,B7)","mm","0")
    out(ws,39,"so confinamiento DES = min(6db,150)","=MIN(6*B9,150)","mm","0")
    nota(ws,41,"Para flexo-compresion COMBINADA (P + Mx + My) usa el DIAGRAMA DE INTERACCION P-M. Esta hoja te da")
    nota(ws,42,"el acero, la cota axial pura, la esbeltez y los estribos. Si k*Lu/r>22 hay que magnificar momentos.")
    semaforo(ws,"B19:B42")

# ============================================================ CIMENTACION
def hoja_cimentacion(wb):
    ws=wb.create_sheet("Cimentacion"); anchos(ws)
    titulo(ws,"A1","ZAPATA AISLADA - NSR-10 / ACI 318","D1")
    seccion(ws,3,"DATOS DE ENTRADA  (azul=geometria/suelo, amarillo=cargas)")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa")
    inp(ws,6,"P servicio (D+L)",1100,"kN",dato=AMA_IN)
    inp(ws,7,"Pu (mayorada)",1500,"kN",dato=AMA_IN)
    inp(ws,8,"q admisible del suelo",150,"kN/m2","Grupo 6 Santa Marta: 15 ton/m2 ~ 150 kN/m2")
    inp(ws,9,"columna b",400,"mm"); inp(ws,10,"columna h",400,"mm")
    inp(ws,11,"h zapata",500,"mm"); inp(ws,12,"cc zapata",75,"mm"); inp(ws,13,"db",15.9,"mm")
    seccion(ws,15,"1) DIMENSIONAMIENTO")
    out(ws,16,"Area requerida = P/qadm","=B6/B8","m2","0.00")
    out(ws,17,"Lado B (cuadrada) redondeado","=CEILING(SQRT(B16),0.05)","m","0.00",bold=True,hl=VERDE)
    out(ws,18,"qu (presion ultima) = Pu/B^2","=B7/(B17^2)","kN/m2","0.0")
    out(ws,19,"d (efectiva) = h-cc-db","=B11-B12-B13","mm","0")
    seccion(ws,21,"2) CORTANTE UNIDIRECCIONAL (a 'd' de la cara)")
    out(ws,22,"volado = (B - col)/2","=(B17-B9/1000)/2","m","0.000")
    out(ws,23,"Vu1 = qu*B*(volado - d)","=B18*B17*(B22-B19/1000)","kN","0.0")
    out(ws,24,"phi*Vc1 = 0.75*0.17raizf'c*B*d","=0.75*0.17*SQRT(B4)*(B17*1000)*B19/1000","kN","0.0")
    out(ws,25,"Chequeo cortante 1D",'=IF(B23<=B24,"CUMPLE","REVISAR")',"","General",bold=True)
    seccion(ws,27,"3) PUNZONAMIENTO (a d/2)")
    out(ws,28,"bo = 2((col_b+d)+(col_h+d))","=2*((B9+B19)+(B10+B19))","mm","0")
    out(ws,29,"Vu2 = Pu - qu*(col_b+d)(col_h+d)","=B7-B18*((B9+B19)/1000)*((B10+B19)/1000)","kN","0.0")
    out(ws,30,"phi*Vc2 = 0.75*0.33raizf'c*bo*d","=0.75*0.33*SQRT(B4)*B28*B19/1000","kN","0.0")
    out(ws,31,"Chequeo punzonamiento",'=IF(B29<=B30,"CUMPLE","REVISAR")',"","General",bold=True)
    seccion(ws,33,"4) FLEXION (en la cara de la columna)")
    out(ws,34,"Mu = qu*B*volado^2/2","=B18*B17*B22^2/2","kN*m","0.0")
    out(ws,35,"Rn","=B34*10^6/(0.9*(B17*1000)*B19^2)","MPa","0.000")
    out(ws,36,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B35/(0.85*B4)))","-","0.00000")
    out(ws,37,"As req (ancho B)","=B36*(B17*1000)*B19","mm2","0")
    out(ws,38,"As min = 0.0018*B*h","=0.0018*(B17*1000)*B11","mm2","0")
    out(ws,39,"As ADOPTADO","=MAX(B37,B38)","mm2","0",bold=True,hl=VERDE)
    out(ws,40,"No. de barras","=ROUNDUP(B39/(PI()/4*B13^2),0)","barras","0",bold=True,hl=AMAR)
    semaforo(ws,"B15:B42")

# ============================================================ MUROS
def hoja_muros(wb):
    ws=wb.create_sheet("Muros"); anchos(ws)
    titulo(ws,"A1","MURO ESTRUCTURAL - cortante en el plano - NSR-10","D1")
    seccion(ws,3,"DATOS DE ENTRADA")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa")
    inp(ws,6,"lw (longitud del muro)",4000,"mm"); inp(ws,7,"tw (espesor)",200,"mm")
    inp(ws,8,"Vu (cortante en el plano)",800,"kN",dato=AMA_IN)
    inp(ws,9,"rho_t (cuantia horizontal)",0.0025,"-","min 0.0025")
    inp(ws,10,"alpha_c (0.17 a 0.25)",0.17,"-","0.25 si hw/lw<=1.5 ; 0.17 si >=2")
    inp(ws,11,"phi",0.75,"-")
    seccion(ws,13,"CORTANTE EN EL PLANO")
    out(ws,14,"Acv = lw*tw","=B6*B7","mm2","0")
    out(ws,15,"phi*Vn = phi*(alpha*raizf'c + rho_t*fy)*Acv","=B11*(B10*SQRT(B4)+B9*B5)*B14/1000","kN","0.0")
    out(ws,16,"Vn max = 0.83*raizf'c*Acv","=0.83*SQRT(B4)*B14/1000","kN","0.0")
    out(ws,17,"Chequeo cuantia (rho_t>=0.0025)",'=IF(B9>=0.0025,"CUMPLE","NO CUMPLE")',"","General",bold=True)
    out(ws,18,"Chequeo cortante (Vu<=phiVn y Vu<=phiVnmax)",'=IF(AND(B8<=B15,B8<=B11*B16),"CUMPLE","NO CUMPLE")',"","General",bold=True)
    semaforo(ws,"B13:B20")

def main():
    wb=Workbook()
    hoja_instrucciones(wb); hoja_materiales(wb)
    hoja_losa(wb); hoja_viguetas(wb); hoja_vigas(wb); hoja_riostras(wb)
    hoja_escaleras(wb); hoja_columnas(wb); hoja_cimentacion(wb); hoja_muros(wb)
    wb.save("Herramienta-Diseno-Estructural.xlsx")
    print("OK -> Herramienta-Diseno-Estructural.xlsx (10 hojas)")

if __name__=="__main__":
    main()
