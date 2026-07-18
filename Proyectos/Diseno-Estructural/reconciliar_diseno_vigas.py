#!/usr/bin/env python3
"""Reconcilia el libro de vigas con los exportes SAP y recalcula sus fórmulas.

El flujo es atómico: trabaja sobre un archivo temporal en el mismo directorio,
lo valida y solo entonces reemplaza el libro de destino. Los armados existentes
no se modifican para forzar resultados favorables.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent
DESIGN = BASE / "Diseño de vigas proyecto diseño DEF.xlsx"
SAP = BASE / "resultados sap.xlsx"
GEOMETRY = BASE / "geomatria sap.xlsx"
DETAIL_SHEET = "Todas las Vigas (592)"
SUMMARY_SHEETS = ("Vigas de Carga (7)", "Vigas de Rigidez (5)")
EXPECTED_FRAMES = 592
EXPECTED_GROUPS = {
    "VC1", "VC2", "VC3", "VC4", "VC5", "VC6", "VC7",
    "VR1", "VR N1", "VR2", "VR3", "VRAUX",
}
EXPECTED_FORMULAS = 21_744
ALIASES = {"VR1 N1": "VR N1", "VR AUX": "VRAUX"}

PARAMETERS = (
    (2, "f'c", 28, "MPa"),
    (3, "fy longitudinal", 420, "MPa"),
    (4, "fyt transversal", 420, "MPa"),
    (5, "φ flexión", 0.9, "-"),
    (6, "φ cortante/torsión", 0.75, "-"),
    (7, "Recubrimiento libre", 40, "mm"),
    (8, "Área barra Nº5", 199, "mm²"),
    (9, "db estribo de referencia", 9.5, "mm; las fórmulas usan BA por fila"),
    (10, "Separación libre entre capas", 25, "mm"),
    (11, "Reservado", None, "-"),
    (12, "db barra Nº5", 15.9, "mm"),
)


def normalize_group(value: object) -> str:
    text = str(value).strip()
    return ALIASES.get(text, text)


def sap_rows(workbook, sheet_name: str):
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[2]]
    for values in ws.iter_rows(min_row=4, values_only=True):
        if values and values[0] is not None:
            yield dict(zip(headers, values))


def numeric_frame(value: str):
    return int(value) if value.isdigit() else value


def load_source_records():
    sap = load_workbook(SAP, read_only=True, data_only=True)
    geometry = load_workbook(GEOMETRY, read_only=True, data_only=True)

    assignments: dict[str, str] = {}
    groups = defaultdict(set)
    for row in sap_rows(sap, "Groups 2 - Assignments"):
        if str(row.get("ObjectType")) != "Frame":
            continue
        group = normalize_group(row.get("GroupName"))
        if group not in EXPECTED_GROUPS:
            continue
        frame = str(row["ObjectLabel"])
        if frame in assignments and assignments[frame] != group:
            raise ValueError(f"Frame {frame} asignado a dos grupos de vigas")
        assignments[frame] = group
        groups[group].add(frame)

    if len(assignments) != EXPECTED_FRAMES or set(groups) != EXPECTED_GROUPS:
        raise ValueError(
            f"Asignaciones SAP inesperadas: {len(assignments)} frames, "
            f"grupos={sorted(groups)}"
        )

    connectivity = {
        str(row["Frame"]): row for row in sap_rows(geometry, "Connectivity - Frame")
    }
    sections = {
        str(row["Frame"]): str(row["AnalSect"])
        for row in sap_rows(geometry, "Frame Section Assignments")
    }
    section_dimensions = {
        str(row["SectionName"]): (float(row["t2"]) * 1000, float(row["t3"]) * 1000)
        for row in sap_rows(sap, "Frame Props 01 - General")
    }

    forces = defaultdict(lambda: defaultdict(list))
    for row in sap_rows(sap, "Element Forces - Frames"):
        frame = str(row["Frame"])
        case = str(row["OutputCase"])
        if frame in assignments and case in {"ENVFLEX", "ENVCORT"}:
            forces[frame][case].append(row)

    records = {}
    for frame, group in assignments.items():
        if frame not in connectivity or frame not in sections:
            raise ValueError(f"Falta geometría para el frame {frame}")
        section = sections[frame]
        if section not in section_dimensions:
            raise ValueError(f"Falta definición de sección {section} para frame {frame}")
        flex = forces[frame]["ENVFLEX"]
        shear = forces[frame]["ENVCORT"]
        if not flex or not shear:
            raise ValueError(f"Faltan envolventes ENVFLEX/ENVCORT para frame {frame}")
        m3 = [float(row["M3"]) for row in flex]
        v2_flex = [float(row["V2"]) for row in flex]
        v2_shear = [float(row["V2"]) for row in shear]
        torsion = [float(row["T"]) for row in shear]
        b_model, h_model = section_dimensions[section]
        records[frame] = {
            "group": group,
            "frame": numeric_frame(frame),
            "length": float(connectivity[frame]["Length"]),
            "b_model": b_model,
            "h_model": h_model,
            "mu_neg": max(0.0, -min(m3)),
            "mu_pos": max(0.0, max(m3)),
            "vu_flex": max(abs(value) for value in v2_flex),
            "vu_shear": max(abs(value) for value in v2_shear),
            "torsion": max(abs(value) for value in torsion),
        }

    sap.close()
    geometry.close()
    return records


def formula_count(workbook) -> int:
    return sum(
        1
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def add_parameters_sheet(workbook):
    if "Parametros" in workbook.sheetnames:
        del workbook["Parametros"]
    ws = workbook.create_sheet("Parametros", 0)
    ws["A1"] = "PARÁMETROS INTERNOS — DISEÑO DE VIGAS"
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    for row, label, value, unit in PARAMETERS:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 3, unit)
    ws["D2"] = "Internalizados para eliminar vínculos externos."
    ws["D3"] = "El diámetro de estribo efectivo se toma de BA en cada fila."
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 55
    ws.freeze_panes = "A2"


def internalize_formulas(workbook):
    replacements = 0
    row_diameter_replacements = 0
    for ws in workbook.worksheets:
        if ws.title == "Parametros":
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                formula = cell.value
                updated = formula.replace("[1]Parametros!", "Parametros!")
                updated = updated.replace("'[1]Todas las Vigas (592)'!", "'Todas las Vigas (592)'!")
                if "Parametros!$B$9" in updated:
                    count = updated.count("Parametros!$B$9")
                    updated = updated.replace("Parametros!$B$9", f"$BA{cell.row}")
                    row_diameter_replacements += count
                if updated != formula:
                    cell.value = updated
                    replacements += 1
    return replacements, row_diameter_replacements


def adopted_dimensions(workbook):
    result = {}
    for sheet_name in SUMMARY_SHEETS:
        ws = workbook[sheet_name]
        for row in range(4, ws.max_row + 1):
            group = normalize_group(ws.cell(row, 1).value)
            result[group] = (float(ws.cell(row, 8).value), float(ws.cell(row, 9).value))
    if set(result) != EXPECTED_GROUPS:
        raise ValueError(f"Dimensiones adoptadas incompletas: {sorted(result)}")
    return result


def update_detail_sheet(workbook, records, adopted):
    ws = workbook[DETAIL_SHEET]
    seen = set()
    for row in range(4, 596):
        frame = str(ws.cell(row, 2).value)
        if frame in seen or frame not in records:
            raise ValueError(f"Frame duplicado o desconocido en {DETAIL_SHEET}: {frame}")
        seen.add(frame)
        source = records[frame]
        existing_group = normalize_group(ws.cell(row, 1).value)
        if existing_group != source["group"]:
            raise ValueError(
                f"El frame {frame} cambió de grupo: libro={existing_group}, SAP={source['group']}"
            )
        b_adopted, h_adopted = adopted[source["group"]]
        values = (
            source["group"], source["frame"], 1,
            source["length"], source["length"],
            source["b_model"], source["h_model"], b_adopted, h_adopted,
            source["mu_neg"], source["frame"],
            source["mu_pos"], source["frame"],
            source["vu_flex"], source["vu_shear"], source["frame"],
            source["torsion"],
        )
        for column, value in enumerate(values, 1):
            ws.cell(row, column, value)
    if seen != set(records):
        raise ValueError(f"No se reconciliaron los {EXPECTED_FRAMES} frames")


def max_record(records, group: str, key: str):
    candidates = [item for item in records.values() if item["group"] == group]
    return max(candidates, key=lambda item: (item[key], numeric_frame(str(item["frame"]))))


def update_summary_sheets(workbook, records, adopted):
    by_group = defaultdict(list)
    for record in records.values():
        by_group[record["group"]].append(record)

    summary_formula = (
        '=IF(AND(X{r}="CUMPLE",AE{r}="CUMPLE",AS{r}="CUMPLE",'
        'OR(AY{r}="CUMPLE",AY{r}="DESPRECIABLE"),BH{r}="CUMPLE",BJ{r}="CUMPLE",'
        'COUNTIF(\'{detail}\'!$A$4:$A$595,$A{r})>0,'
        'COUNTIFS(\'{detail}\'!$A$4:$A$595,$A{r},'
        '\'{detail}\'!$AZ$4:$AZ$595,"<>CUMPLE")=0),"CUMPLE","NO CUMPLE")'
    )

    for sheet_name in SUMMARY_SHEETS:
        ws = workbook[sheet_name]
        for row in range(4, ws.max_row + 1):
            group = normalize_group(ws.cell(row, 1).value)
            items = by_group[group]
            if not items:
                raise ValueError(f"Grupo sin frames: {group}")
            model_dimensions = {(item["b_model"], item["h_model"]) for item in items}
            if len(model_dimensions) != 1:
                raise ValueError(f"El grupo {group} tiene varias secciones de modelo")
            b_model, h_model = model_dimensions.pop()
            b_adopted, h_adopted = adopted[group]
            mu_neg = max_record(records, group, "mu_neg")
            mu_pos = max_record(records, group, "mu_pos")
            vu_flex = max_record(records, group, "vu_flex")
            vu_shear = max_record(records, group, "vu_shear")
            torsion = max_record(records, group, "torsion")
            values = (
                group, torsion["frame"], len(items),
                min(item["length"] for item in items),
                max(item["length"] for item in items),
                b_model, h_model, b_adopted, h_adopted,
                mu_neg["mu_neg"], mu_neg["frame"],
                mu_pos["mu_pos"], mu_pos["frame"],
                vu_flex["vu_flex"], vu_shear["vu_shear"], vu_shear["frame"],
                torsion["torsion"],
            )
            for column, value in enumerate(values, 1):
                ws.cell(row, column, value)
            ws.cell(row, 52, summary_formula.format(r=row, detail=DETAIL_SHEET))


def recalculate_with_poi(path: Path):
    pom = """<project xmlns="http://maven.apache.org/POM/4.0.0">
  <modelVersion>4.0.0</modelVersion><groupId>local</groupId><artifactId>excel-recalc</artifactId><version>1</version>
  <properties><maven.compiler.release>17</maven.compiler.release></properties>
  <dependencies><dependency><groupId>org.apache.poi</groupId><artifactId>poi-ooxml</artifactId><version>5.4.1</version></dependency></dependencies>
  <build><plugins>
    <plugin><groupId>org.apache.maven.plugins</groupId><artifactId>maven-compiler-plugin</artifactId><version>3.14.0</version></plugin>
    <plugin><groupId>org.codehaus.mojo</groupId><artifactId>exec-maven-plugin</artifactId><version>3.5.0</version></plugin>
  </plugins></build>
</project>"""
    java = r"""import java.io.*;
import java.nio.file.*;
import org.apache.poi.ss.usermodel.*;
public class RecalcularExcel {
  public static void main(String[] args) throws Exception {
    Path path = Paths.get(System.getenv("WORKBOOK_PATH"));
    Workbook workbook;
    try (InputStream input = Files.newInputStream(path)) { workbook = WorkbookFactory.create(input); }
    FormulaEvaluator evaluator = workbook.getCreationHelper().createFormulaEvaluator();
    int formulas = 0;
    int errors = 0;
    for (Sheet sheet : workbook) for (Row row : sheet) for (Cell cell : row) {
      if (cell.getCellType() == CellType.FORMULA) {
        formulas++;
        CellType result = evaluator.evaluateFormulaCell(cell);
        if (result == CellType.ERROR) {
          errors++;
          System.err.println(sheet.getSheetName() + "!" + cell.getAddress() + "=" + FormulaError.forInt(cell.getErrorCellValue()).getString());
        }
      }
    }
    if (errors != 0) throw new IllegalStateException("Errores de fórmula: " + errors);
    workbook.setForceFormulaRecalculation(true);
    try (OutputStream output = Files.newOutputStream(path)) { workbook.write(output); }
    workbook.close();
    System.out.println("POI recalculó " + formulas + " fórmulas sin errores");
  }
}"""
    with tempfile.TemporaryDirectory(prefix="recalculo-poi-") as directory:
        project = Path(directory)
        source = project / "src/main/java"
        source.mkdir(parents=True)
        (project / "pom.xml").write_text(pom, encoding="utf-8")
        (source / "RecalcularExcel.java").write_text(java, encoding="utf-8")
        environment = os.environ.copy()
        environment["WORKBOOK_PATH"] = str(path)
        subprocess.run(
            [
                "mvn", "-q", "compile", "exec:java",
                "-Dexec.mainClass=RecalcularExcel",
            ],
            cwd=project,
            env=environment,
            check=True,
        )


def validate_workbook(path: Path):
    with zipfile.ZipFile(path) as archive:
        external_parts = [name for name in archive.namelist() if "externalLinks" in name]
        if external_parts:
            raise ValueError(f"Persisten partes OOXML externas: {external_parts}")

    formulas = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    count = formula_count(formulas)
    external_formulas = []
    for ws in formulas.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and "[1]" in str(cell.value):
                    external_formulas.append(f"{ws.title}!{cell.coordinate}")
    formulas.close()
    if count != EXPECTED_FORMULAS:
        raise ValueError(f"Fórmulas preservadas={count}; esperadas={EXPECTED_FORMULAS}")
    if external_formulas:
        raise ValueError(f"Persisten fórmulas externas: {external_formulas[:10]}")

    cached = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    errors = []
    for ws in cached.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "e" or (isinstance(cell.value, str) and cell.value.startswith("#")):
                    errors.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    detail_status = Counter(
        cached[DETAIL_SHEET].cell(row, 52).value for row in range(4, 596)
    )
    summary_status = {
        cached[sheet].cell(row, 52).value
        for sheet in SUMMARY_SHEETS
        for row in range(4, cached[sheet].max_row + 1)
    }
    cached.close()
    if errors:
        raise ValueError(f"Errores Excel almacenados: {errors[:10]}")
    if sum(detail_status.values()) != EXPECTED_FRAMES:
        raise ValueError("No se obtuvieron estados para los 592 frames")
    if not set(detail_status).issubset({"CUMPLE", "NO CUMPLE"}):
        raise ValueError(f"Estados inesperados: {detail_status}")
    if not summary_status.issubset({"CUMPLE", "NO CUMPLE"}):
        raise ValueError(f"Estados resumen inesperados: {summary_status}")
    return count, detail_status


def main():
    records = load_source_records()
    with tempfile.NamedTemporaryFile(
        prefix="diseno-vigas-reconciliado-", suffix=".xlsx", dir=BASE, delete=False
    ) as handle:
        temporary = Path(handle.name)
    try:
        shutil.copy2(DESIGN, temporary)
        workbook = load_workbook(temporary, data_only=False, keep_links=False)
        initial_formulas = formula_count(workbook)
        if initial_formulas != EXPECTED_FORMULAS:
            raise ValueError(
                f"El libro de entrada contiene {initial_formulas} fórmulas; "
                f"se esperaban {EXPECTED_FORMULAS}"
            )
        adopted = adopted_dimensions(workbook)
        update_detail_sheet(workbook, records, adopted)
        update_summary_sheets(workbook, records, adopted)
        add_parameters_sheet(workbook)
        replaced, row_diameters = internalize_formulas(workbook)
        if formula_count(workbook) != initial_formulas:
            raise ValueError("La reconciliación alteró la cantidad de fórmulas")
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(temporary)
        workbook.close()

        recalculate_with_poi(temporary)
        count, statuses = validate_workbook(temporary)
        os.replace(temporary, DESIGN)
        print(f"OK {DESIGN.name}: {len(records)} frames, {count} fórmulas")
        print(f"Vínculos internalizados en {replaced} fórmulas; usos de db por fila: {row_diameters}")
        print("Estados recalculados:", dict(statuses))
        print("ADVERTENCIA: los NO CUMPLE se conservan; no constituyen aprobación estructural.")
    finally:
        temporary.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
