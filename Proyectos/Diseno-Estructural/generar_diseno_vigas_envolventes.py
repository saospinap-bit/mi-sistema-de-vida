#!/usr/bin/env python3
"""Genera la verificación seccional conservadora de 592 vigas.

ENVFLEX aporta Mu-, Mu+ y Vu reducido; ENVCORT aporta la demanda directa de
cortante. Se adopta Vu_diseno = max(Vu_ENVCORT, Ve_DMO). La torsión de diseño
es el máximo absoluto entre ambas envolventes, sin recortarla automáticamente
a phi*Tcr: el export no demuestra cuáles torsiones son de compatibilidad ni la
redistribución hacia elementos adyacentes. El acero longitudinal de torsión se
provee adicional al acero de flexión. Todo incumplimiento se conserva visible.
"""
from __future__ import annotations

import hashlib
import math
import shutil
import tempfile
import zipfile
from collections import defaultdict
from copy import copy
from pathlib import Path

from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
SAP_FILE = BASE / "resultados sap.xlsx"
OUTPUT_FILE = BASE / "DISENO-VIGAS-ENTREGA-FINAL-DMO.xlsx"

# Materiales y detallado
FC = 28.0
FY = 420.0
FYT = 420.0
PHI_M = 0.90
PHI_V = 0.75
COVER = 40.0
DB_LONG = 15.9
AB_LONG = 199.0
DB_ST = 12.7
AB_ST = 129.0
CLEAR_BAR = 25.0
CLEAR_SPAN_DEDUCTION = 0.60  # m; hipótesis conservadora total entre ejes y caras de apoyos
SYSTEM = "DMO"
R0 = 5.0
PHI_A = 1.0
PHI_P = 1.0
PHI_R = 1.0
R_DESIGN = R0 * PHI_A * PHI_P * PHI_R

# Secciones efectivamente asignadas y reanalizadas en SAP.
# VC1–VC7 y VRAUX: 450×550 mm. VR1, VR N1, VR2 y VR3: 500×550 mm.
EXPECTED_SECTIONS = {
    "VC": (450.0, 550.0),
    "VR": (500.0, 550.0),
    "VRAUX": (450.0, 550.0),
}

GROUP_ORDER = ["VC1", "VC2", "VC3", "VC4", "VC5", "VC6", "VC7",
               "VR1", "VR N1", "VR2", "VR3", "VRAUX"]

COLORS = {
    "VR1": "DDEBF7", "VR N1": "BDD7EE", "VR2": "9DC3E6", "VR3": "5B9BD5",
    "VRAUX": "A9D18E", "VC1": "F4B183", "VC2": "FFD966", "VC3": "FFF2CC",
    "VC4": "E2F0D9", "VC5": "C6E0B4", "VC6": "A5A5A5", "VC7": "D9EAD3",
}

# Paleta sencilla de hoja académica: tonos estándar de Excel, sin apariencia corporativa.
NAVY = "5B9BD5"
BLUE = "DDEBF7"
GREEN = "E2F0D9"
RED = "F4CCCC"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "262626"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def iter_table(path: Path, sheet: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    headers = list(next(rows))
    next(rows)
    for values in rows:
        if values and values[0] is not None:
            yield dict(zip(headers, values))


def normalize_group(name: str) -> str:
    aliases = {"VR AUX": "VRAUX", "VR1 N1": "VR N1"}
    return aliases.get(str(name).strip(), str(name).strip())


def extract_inputs():
    """Extrae fuerzas, grupos, longitudes y secciones del export actualizado de SAP."""
    wb = load_workbook(SAP_FILE, read_only=True, data_only=True)

    def rows(sheet):
        ws = wb[sheet]
        iterator = ws.iter_rows(values_only=True)
        next(iterator)
        headers = list(next(iterator))
        next(iterator)
        return [dict(zip(headers, values)) for values in iterator
                if values and values[0] is not None]

    program = rows("Program Control")
    if not program or str(program[0].get("CurrUnits")) != "KN, m, C":
        raise ValueError("El export SAP debe estar en unidades KN, m, C")

    # Pertenencia exacta de los 12 grupos de vigas.
    group_frames = {group: set() for group in GROUP_ORDER}
    frame_group = {}
    for row in rows("Groups 2 - Assignments"):
        if str(row.get("ObjectType")) != "Frame":
            continue
        group = normalize_group(row.get("GroupName"))
        if group not in group_frames:
            continue
        frame = str(row["ObjectLabel"])
        if frame in frame_group and frame_group[frame] != group:
            raise ValueError(f"Frame {frame} repetido en {frame_group[frame]} y {group}")
        frame_group[frame] = group
        group_frames[group].add(frame)
    if len(frame_group) != 592:
        raise ValueError(f"Se esperaban 592 vigas agrupadas y se encontraron {len(frame_group)}")
    if any(not group_frames[g] for g in GROUP_ORDER):
        raise ValueError("Falta al menos uno de los 12 grupos de vigas")

    # Propiedades y asignaciones de sección; t2=b y t3=h para rectángulos SAP.
    properties = {}
    for row in rows("Frame Props 01 - General"):
        if str(row.get("Shape")) == "Rectangular":
            properties[str(row["SectionName"])] = (
                float(row["t2"]) * 1000,
                float(row["t3"]) * 1000,
            )
    section_by_frame = {}
    for row in rows("Frame Section Assignments"):
        frame = str(row["Frame"])
        section = str(row["AnalSect"])
        if frame in frame_group:
            if section not in properties:
                raise ValueError(f"No se encontró la propiedad {section} del frame {frame}")
            section_by_frame[frame] = (section, *properties[section])

    connectivity = {str(row["Frame"]): row for row in rows("Connectivity - Frame")}
    missing_geometry = set(frame_group) - set(connectivity)
    missing_sections = set(frame_group) - set(section_by_frame)
    if missing_geometry or missing_sections:
        raise ValueError(f"Faltan geometría/sección: {len(missing_geometry)}/{len(missing_sections)}")

    raw = defaultdict(lambda: defaultdict(list))
    for row in rows("Element Forces - Frames"):
        frame = str(row["Frame"])
        if frame in frame_group and str(row["OutputCase"]) in ("ENVFLEX", "ENVCORT"):
            raw[frame][str(row["OutputCase"])].append(row)
    wb.close()
    if len(raw) != 592:
        raise ValueError(f"Se esperaban fuerzas para 592 vigas y se encontraron {len(raw)}")

    records = []
    for frame, group in frame_group.items():
        flex = raw[frame].get("ENVFLEX", [])
        shear = raw[frame].get("ENVCORT", [])
        if not flex or not shear:
            raise ValueError(f"El frame {frame} no tiene ambas envolventes")

        def vals(source, key):
            return [float(row[key] or 0.0) for row in source]

        geometry = connectivity[frame]
        section_name, b_sap, h_sap = section_by_frame[frame]
        stations = vals(flex + shear, "Station")
        length = float(geometry["Length"])
        if abs(max(stations) - length) > 1e-3:
            raise ValueError(f"Longitud inconsistente en frame {frame}: {max(stations)} vs {length}")
        m3 = vals(flex, "M3")
        torsion_flex = vals(flex, "T")
        torsion_shear = vals(shear, "T")
        tu_flex = max(abs(x) for x in torsion_flex)
        tu_cort = max(abs(x) for x in torsion_shear)
        tu_raw = max(tu_flex, tu_cort)
        acp = b_sap * h_sap
        pcp = 2 * (b_sap + h_sap)
        phi_tcr_reference = PHI_V * 0.33 * math.sqrt(FC) * acp * acp / pcp / 1e6
        # No se reduce Tu a phi*Tcr: no se dispone de una clasificación de
        # compatibilidad ni de un reanálisis que reciba los esfuerzos redistribuidos.
        tu_design = tu_raw
        records.append({
            "frame": frame, "group": group, "L": length,
            "Mu_neg": max(0.0, -min(m3)), "Mu_pos": max(0.0, max(m3)),
            "Vu_flex": max(abs(x) for x in vals(flex, "V2")),
            "Vu_elastic": max(abs(x) for x in vals(shear, "V2")),
            "Tu_flex": tu_flex, "Tu_cort": tu_cort, "Tu_raw": tu_raw,
            "phi_tcr_compat": phi_tcr_reference, "Tu": tu_design,
            "requires_compatibility_assessment": tu_raw > phi_tcr_reference + 1e-9,
            "torsion_redistributed": False,
            "section_name": section_name, "b_sap": b_sap, "h_sap": h_sap,
            "joint_i": str(geometry["JointI"]), "joint_j": str(geometry["JointJ"]),
            "centroid_x": float(geometry["CentroidX"]),
            "centroid_y": float(geometry["CentroidY"]),
            "centroid_z": float(geometry["CentroidZ"]),
        })

    # Validación de las secciones aceptadas por el usuario.
    for record in records:
        key = "VRAUX" if record["group"] == "VRAUX" else ("VC" if record["group"].startswith("VC") else "VR")
        expected = EXPECTED_SECTIONS[key]
        actual = (record["b_sap"], record["h_sap"])
        if actual != expected:
            raise ValueError(f"Sección inesperada en {record['group']}/{record['frame']}: {actual} != {expected}")
    return records, group_frames


def base_section(record: dict):
    b = float(record["b_sap"])
    h = float(record["h_sap"])
    return b, h, b, h


def beta1():
    return 0.85 if FC <= 28 else max(0.65, 0.85 - 0.05 * (FC - 28) / 7)


def nominal_as_required(mu: float, b: float, d: float):
    as_min = max(1.4 / FY, 0.25 * math.sqrt(FC) / FY) * b * d
    rn = mu * 1e6 / (PHI_M * b * d * d) if mu > 0 else 0.0
    disc = 1 - 2 * rn / (0.85 * FC)
    if disc <= 0:
        return math.inf, as_min
    rho = (0.85 * FC / FY) * (1 - math.sqrt(disc))
    return max(rho * b * d, as_min), as_min


def bar_layout(n: int, b: float, h: float):
    available = b - 2 * COVER - 2 * DB_ST
    per_layer = max(2, math.floor((available + CLEAR_BAR) / (DB_LONG + CLEAR_BAR)))
    if n <= per_layer:
        layers, n1, n2 = 1, n, 0
    elif n <= 2 * per_layer:
        layers, n1, n2 = 2, math.ceil(n / 2), math.floor(n / 2)
    else:
        return {"fits": False, "layers": 3, "d": 0.0, "per_layer": per_layer}
    y1 = COVER + DB_ST + DB_LONG / 2
    y2 = y1 + DB_LONG + CLEAR_BAR
    centroid = y1 if layers == 1 else (n1 * y1 + n2 * y2) / n
    return {"fits": True, "layers": layers, "d": h - centroid, "per_layer": per_layer}


def flexural_design(mu: float, b: float, h: float, minimum_n: int = 2):
    d_nom = h - COVER - DB_ST - DB_LONG / 2
    as_req_nom, _ = nominal_as_required(mu, b, d_nom)
    n = max(minimum_n, math.ceil(as_req_nom / AB_LONG))
    for _ in range(30):
        layout = bar_layout(n, b, h)
        if not layout["fits"]:
            return {"ok": False, "n": n, "layers": layout["layers"], "d": 0.0,
                    "as_req": math.inf, "as_min": math.inf, "as_prov": n * AB_LONG,
                    "phi_mn": 0.0, "mpr": 0.0}
        d = layout["d"]
        as_req, as_min = nominal_as_required(mu, b, d)
        as_prov = n * AB_LONG
        a = as_prov * FY / (0.85 * FC * b)
        phi_mn = PHI_M * as_prov * FY * (d - a / 2) / 1e6
        rho = as_prov / (b * d)
        rho_max = 0.85 * beta1() * FC / FY * 3 / 8
        if as_prov + 1e-9 >= as_req and phi_mn + 1e-9 >= mu and rho <= rho_max:
            a_pr = as_prov * 1.25 * FY / (0.85 * FC * b)
            mpr = as_prov * 1.25 * FY * (d - a_pr / 2) / 1e6
            return {"ok": True, "n": n, "layers": layout["layers"], "d": d,
                    "as_req": as_req, "as_min": as_min, "as_prov": as_prov,
                    "phi_mn": phi_mn, "mpr": mpr, "rho": rho,
                    "per_layer": layout["per_layer"]}
        n += 1
    raise RuntimeError("No convergió el diseño a flexión")


def floor10(value: float):
    return max(10.0, math.floor(value / 10.0) * 10.0)


def clear_span(station_length: float):
    """Adopta una luz libre conservadora mientras se confirman caras de apoyos en SAP."""
    return max(station_length - CLEAR_SPAN_DEDUCTION, 0.5 * station_length)


def design_record(record: dict, group_mode: bool = False):
    group = record["group"]
    b0, h0, b, h = base_section(record)

    neg = flexural_design(record["Mu_neg"], b, h, 2)
    if not neg["ok"]:
        raise ValueError(f"No cabe el acero superior de {group}/{record.get('frame','grupo')}")
    # Requisito de continuidad DMO: al menos 2 barras y no menos de 1/3 del acero superior.
    min_pos = max(2, math.ceil(neg["n"] / 3))
    pos = flexural_design(record["Mu_pos"], b, h, min_pos)
    if not pos["ok"]:
        raise ValueError(f"No cabe el acero inferior de {group}/{record.get('frame','grupo')}")

    d = min(neg["d"], pos["d"])
    station_length = record["L_min"] if group_mode else record["L"]
    # Hipótesis conservadora provisional: luz libre = longitud de estación menos 0.60 m.
    length = clear_span(station_length)
    # Se adopta directamente ENVCORT y se conserva Ve,DMO como verificación adicional.
    vcap = record["Vu_flex"] + (neg["mpr"] + pos["mpr"]) / length
    vdesign = max(record["Vu_elastic"], vcap)
    vc = 0.17 * math.sqrt(FC) * b * d / 1000
    vs_req = max(vdesign / PHI_V - vc, 0.0)
    vs_max = 0.66 * math.sqrt(FC) * b * d / 1000
    shear_section_ok = vs_req <= vs_max + 1e-9

    acp = b * h
    pcp = 2 * (b + h)
    phi_tth = PHI_V * 0.083 * math.sqrt(FC) * acp * acp / pcp / 1e6
    centerline_cover = COVER + DB_ST / 2
    aoh = (b - 2 * centerline_cover) * (h - 2 * centerline_cover)
    ao = 0.85 * aoh
    ph = 2 * ((b - 2 * centerline_cover) + (h - 2 * centerline_cover))
    at_s = 0.0 if record["Tu"] <= phi_tth else record["Tu"] * 1e6 / (PHI_V * 2 * ao * FYT)
    al = at_s * ph * FYT / FY
    # Refuerzo longitudinal de torsión dedicado: no se acredita el acero ya
    # requerido por flexión. Se exige además una distribución perimetral con
    # separación máxima aproximada de 300 mm y se redondea a múltiplos de 4.
    horizontal_segments = math.ceil((b - 2 * centerline_cover) / 300)
    vertical_segments = math.ceil((h - 2 * centerline_cover) / 300)
    min_perimeter_bars = 2 * horizontal_segments + 2 * vertical_segments
    area_bars = math.ceil(al / AB_LONG) if at_s > 0 else 0
    n_torsion = 0 if at_s <= 0 else max(min_perimeter_bars, 4 * math.ceil(area_bars / 4))
    al_prov = n_torsion * AB_LONG
    longitudinal_torsion_ok = al_prov + 1e-9 >= al

    # Interacción de esfuerzos en la sección por cortante y torsión.
    interaction_vt = math.sqrt(
        (vdesign * 1000 / (b * d)) ** 2
        + (record["Tu"] * 1e6 * ph / (1.7 * aoh ** 2)) ** 2
    )
    interaction_limit = PHI_V * (vc * 1000 / (b * d) + 0.66 * math.sqrt(FC))
    interaction_ok = interaction_vt <= interaction_limit + 1e-9

    # El acero transversal combina Av/s por cortante + 2At/s por torsión.
    # Refuerzo mínimo transversal: mayor de las dos expresiones reglamentarias.
    av_s_min = max(0.062 * math.sqrt(FC) * b / FYT, 0.35 * b / FYT)
    av_s_strength = vs_req * 1000 / (FYT * d) if vs_req > 0 else 0.0
    av_s_required = max(av_s_strength, av_s_min)
    combined_s_required = av_s_required + 2 * at_s
    intense_shear = vs_req > 0.33 * math.sqrt(FC) * b * d / 1000
    candidates = []
    for stirrup_label, db_st, ab_branch in (("Nº3 cerrado", 9.5, 71.0), ("Nº4 cerrado", 12.7, 129.0)):
        for legs in (2, 4):
            s_resistance = math.inf if combined_s_required <= 1e-12 else legs * ab_branch / combined_s_required
            torsion_spacing = ph / 8 if at_s > 0 else math.inf
            s_end = floor10(min(s_resistance, d / 4, 8 * DB_LONG, 24 * db_st, 300, torsion_spacing))
            s_center_limit = d / 4 if intense_shear else d / 2
            s_center = floor10(min(s_resistance, s_center_limit, 600, torsion_spacing))
            av_s_provided = legs * ab_branch / s_end
            av_s_effective = max(av_s_provided - 2 * at_s, 0.0)
            phi_vn = PHI_V * (vc + av_s_effective * FYT * d / 1000)
            transverse_torsion_ok = at_s <= 1e-12 or ab_branch / s_end + 1e-12 >= at_s
            if s_end >= 80 and phi_vn + 1e-9 >= vdesign and transverse_torsion_ok:
                candidates.append((legs * ab_branch, -s_end, stirrup_label, db_st, ab_branch,
                                   legs, s_end, s_center, phi_vn, av_s_provided))
    practical_stirrup_found = bool(candidates)
    if candidates:
        _, _, stirrup_label, db_st, ab_branch, legs, s_end, s_center, phi_vn, av_s_provided = min(candidates)
    else:
        # Se conserva una alternativa de referencia para poder reportar la fila,
        # pero nunca se considera conforme ni se emite como detalle aprobado.
        stirrup_label, db_st, ab_branch, legs = "Nº4 cerrado (referencia)", 12.7, 129.0, 4
        s_resistance = math.inf if combined_s_required <= 1e-12 else legs * ab_branch / combined_s_required
        torsion_spacing = ph / 8 if at_s > 0 else math.inf
        s_end = floor10(min(s_resistance, d / 4, 8 * DB_LONG, 24 * db_st, 300, torsion_spacing))
        s_center_limit = d / 4 if intense_shear else d / 2
        s_center = floor10(min(s_resistance, s_center_limit, 600, torsion_spacing))
        av_s_provided = legs * ab_branch / s_end
        av_s_effective = max(av_s_provided - 2 * at_s, 0.0)
        phi_vn = PHI_V * (vc + av_s_effective * FYT * d / 1000)
    transverse_torsion_ok = (
        practical_stirrup_found
        and (at_s <= 1e-12 or ab_branch / s_end + 1e-12 >= at_s)
    )
    shear_ok = practical_stirrup_found and shear_section_ok and phi_vn + 1e-9 >= vdesign
    torsion_ok = transverse_torsion_ok and longitudinal_torsion_ok and interaction_ok
    section_changed = False
    overall = neg["ok"] and pos["ok"] and shear_ok and torsion_ok

    return {**record, "b0": b0, "h0": h0, "b": b, "h": h, "Ln": length,
            "neg": neg, "pos": pos, "d": d, "Vcap": vcap, "Vdesign": vdesign,
            "Vc": vc, "Vs_req": vs_req, "Vs_max": vs_max,
            "shear_section_ok": shear_section_ok,
            "phi_vmax": PHI_V * (vc + vs_max),
            "direct_elastic_ok": PHI_V * (vc + vs_max) + 1e-9 >= record["Vu_elastic"],
            "stirrup_label": stirrup_label, "db_st": db_st, "ab_branch": ab_branch,
            "legs": legs, "s_end": s_end, "s_center": s_center,
            "av_s_required": av_s_required, "av_s_min": av_s_min,
            "combined_s_required": combined_s_required,
            "av_s_provided": av_s_provided, "phi_vn": phi_vn,
            "practical_stirrup_found": practical_stirrup_found,
            "transverse_torsion_ok": transverse_torsion_ok,
            "shear_ok": shear_ok, "phi_tth": phi_tth, "at_s": at_s,
            "Al": al, "Al_prov": al_prov, "n_torsion": n_torsion,
            "interaction_vt": interaction_vt, "interaction_limit": interaction_limit,
            "interaction_ok": interaction_ok, "torsion_ok": torsion_ok,
            "section_changed": section_changed, "overall": overall}


def build_group_records(records, all_results=None):
    """Resume cada grupo con el frame de mayor utilización real, sin mezclar máximos incompatibles."""
    if all_results is None:
        all_results = [design_record(record) for record in records]
    raw_by_group = defaultdict(list)
    result_by_group = defaultdict(list)
    for record in records:
        raw_by_group[record["group"]].append(record)
    for result in all_results:
        result_by_group[result["group"]].append(result)

    output = []
    for group in GROUP_ORDER:
        items = raw_by_group[group]
        designs = result_by_group[group]

        def utilization(result):
            ratios = [
                result["Mu_neg"] / result["neg"]["phi_mn"] if result["neg"]["phi_mn"] else math.inf,
                result["Mu_pos"] / result["pos"]["phi_mn"] if result["pos"]["phi_mn"] else math.inf,
                result["Vdesign"] / result["phi_vn"] if result["phi_vn"] else math.inf,
                result["interaction_vt"] / result["interaction_limit"] if result["interaction_limit"] else math.inf,
                result["Al"] / result["Al_prov"] if result["Al_prov"] else (0.0 if result["Al"] <= 1e-12 else math.inf),
            ]
            return max(ratios)

        governing = dict(max(designs, key=utilization))
        crit_neg = max(items, key=lambda x: x["Mu_neg"])
        crit_pos = max(items, key=lambda x: x["Mu_pos"])
        crit_v = max(items, key=lambda x: x["Vu_elastic"])
        crit_t = max(items, key=lambda x: x["Tu_raw"])
        governing.update({
            "frame": governing["frame"], "count": len(items),
            "L_min": min(x["L"] for x in items), "L_max": max(x["L"] for x in items),
            "frame_Mu_neg": crit_neg["frame"], "frame_Mu_pos": crit_pos["frame"],
            "frame_Vu": crit_v["frame"], "frame_Tu": crit_t["frame"],
            "max_Mu_neg": crit_neg["Mu_neg"], "max_Mu_pos": crit_pos["Mu_pos"],
            "max_Vu_flex": max(x["Vu_flex"] for x in items),
            "max_Vu_elastic": crit_v["Vu_elastic"],
            "max_Tu_raw": crit_t["Tu_raw"],
            "group_all_ok": all(x["overall"] for x in designs),
            "governing_utilization": utilization(governing),
        })
        output.append(governing)
    return output


def sort_frame(frame: str):
    try:
        return (0, int(frame))
    except ValueError:
        return (1, frame)


def sha256(path: Path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


HEADERS = [
    "Grupo", "Frame", "N vigas", "Ln adopt. mín (m)", "Ln adopt. máx (m)",
    "b modelo (mm)", "h modelo (mm)", "b adopt. (mm)", "h adopt. (mm)",
    "Mu− ENVFLEX (kN·m)", "Frame Mu−", "Mu+ ENVFLEX (kN·m)", "Frame Mu+",
    "Vu ENVFLEX (kN)", "Vu ENVCORT (kN)", "Frame Vu", "Tu diseño compat. (kN·m)",
    "As req. sup (mm²)", "Nº5 sup", "Capas sup", "d sup (mm)", "As prov. sup (mm²)",
    "φMn− (kN·m)", "Flexión −", "As req. inf (mm²)", "Nº5 inf", "Capas inf",
    "d inf (mm)", "As prov. inf (mm²)", "φMn+ (kN·m)", "Flexión +",
    "Mpr− (kN·m)", "Mpr+ (kN·m)", "Vcap (kN)", "Vu diseño (kN)",
    "Vc (kN)", "Vs req. (kN)", "Vs máx. (kN)", "Sección cortante", "Estribo",
    "Ramas", "s extremo DMO (mm; zona 2h)", "s centro (mm)", "φVn (kN)", "Cortante",
    "φTth (kN·m)", "Torsión", "At/s (mm²/mm)", "Al adicional req. (mm²)", "Nº5 torsión dedicadas", "Chequeo torsión", "ESTADO",
    "db estribo (mm)", "Área rama (mm²)", "Av/s req. cortante incl. mín.", "2At/s req. torsión",
    "Av/s prov. extremo", "Interacción V-T", "Límite V-T", "Chequeo V-T",
    "Al provisto (mm²)", "Chequeo Al"
]


def add_formula(ws, coord: str, formula: str, cached, cache):
    ws[coord] = formula
    cache[(ws.title, coord)] = cached


def write_design_sheet(wb, name: str, rows, grouped: bool, cache):
    ws = wb.create_sheet(name)
    last_col = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Diseño de vigas - {name}"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = CENTER
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Flexión: ENVFLEX. Cortante: ENVCORT directo. Torsión: máximo de ambas envolventes. Verificación DMO: Ve. Vu adoptado = máx(ENVCORT; Ve)."
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A2"].alignment = LEFT
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[3].height = 64

    for r_idx, result in enumerate(rows, 4):
        group = result["group"]
        color = COLORS[group]
        length_min = result["Ln"]
        length_max = clear_span(result["L_max"]) if grouped else result["Ln"]
        frame = result["frame"]
        count = result.get("count", 1)
        crit_neg = result.get("frame_Mu_neg", frame) if grouped else frame
        crit_pos = result.get("frame_Mu_pos", frame) if grouped else frame
        crit_v = result.get("frame_Vu", frame) if grouped else frame
        overall_status = result["group_all_ok"] if grouped else result["overall"]
        values = [group, frame, count, length_min, length_max, result["b0"], result["h0"],
                  result["b"], result["h"], result["Mu_neg"], crit_neg,
                  result["Mu_pos"], crit_pos, result["Vu_flex"], result["Vu_elastic"],
                  crit_v, result["Tu"]]
        for c_idx, value in enumerate(values, 1):
            ws.cell(r_idx, c_idx, value)

        # Adopted bar counts and detailing choices are explicit design decisions; capacities remain formulas.
        ws.cell(r_idx, 19, result["neg"]["n"])
        ws.cell(r_idx, 20, result["neg"]["layers"])
        ws.cell(r_idx, 26, result["pos"]["n"])
        ws.cell(r_idx, 27, result["pos"]["layers"])
        ws.cell(r_idx, 40, result["stirrup_label"])
        ws.cell(r_idx, 41, result["legs"])
        ws.cell(r_idx, 50, result["n_torsion"])
        ws.cell(r_idx, 53, result["db_st"])
        ws.cell(r_idx, 54, result["ab_branch"])

        # Formulas. Las referencias a Parametros mantienen el libro editable.
        p = "'Parametros'!"
        group_gate = ""
        if grouped:
            group_gate = (
                f",COUNTIF('Todas las Vigas (592)'!$A$4:$A$595,A{r_idx})="
                f"COUNTIFS('Todas las Vigas (592)'!$A$4:$A$595,A{r_idx},"
                f"'Todas las Vigas (592)'!$AZ$4:$AZ$595,\"CUMPLE\")"
            )
        formulas = {
            18: (f"=MAX(MAX(1.4/{p}$B$3,0.25*SQRT({p}$B$2)/{p}$B$3)*H{r_idx}*U{r_idx},"
                 f"(0.85*{p}$B$2/{p}$B$3)*(1-SQRT(1-2*(J{r_idx}*1000000/({p}$B$5*H{r_idx}*U{r_idx}^2))/(0.85*{p}$B$2)))*H{r_idx}*U{r_idx})", result["neg"]["as_req"]),
            21: (f"=IF(T{r_idx}=1,I{r_idx}-({p}$B$7+{p}$B$9+{p}$B$12/2),I{r_idx}-(ROUNDUP(S{r_idx}/2,0)*({p}$B$7+{p}$B$9+{p}$B$12/2)+ROUNDDOWN(S{r_idx}/2,0)*({p}$B$7+{p}$B$9+{p}$B$12/2+{p}$B$12+{p}$B$10))/S{r_idx})", result["neg"]["d"]),
            22: (f"=S{r_idx}*{p}$B$8", result["neg"]["as_prov"]),
            23: (f"={p}$B$5*V{r_idx}*{p}$B$3*(U{r_idx}-(V{r_idx}*{p}$B$3/(0.85*{p}$B$2*H{r_idx}))/2)/1000000", result["neg"]["phi_mn"]),
            24: (f'=IF(W{r_idx}>=J{r_idx},"CUMPLE","NO CUMPLE")', "CUMPLE" if result["neg"]["phi_mn"] >= result["Mu_neg"] else "NO CUMPLE"),
            25: (f"=MAX(MAX(1.4/{p}$B$3,0.25*SQRT({p}$B$2)/{p}$B$3)*H{r_idx}*AB{r_idx},"
                 f"(0.85*{p}$B$2/{p}$B$3)*(1-SQRT(1-2*(L{r_idx}*1000000/({p}$B$5*H{r_idx}*AB{r_idx}^2))/(0.85*{p}$B$2)))*H{r_idx}*AB{r_idx})", result["pos"]["as_req"]),
            28: (f"=IF(AA{r_idx}=1,I{r_idx}-({p}$B$7+{p}$B$9+{p}$B$12/2),I{r_idx}-(ROUNDUP(Z{r_idx}/2,0)*({p}$B$7+{p}$B$9+{p}$B$12/2)+ROUNDDOWN(Z{r_idx}/2,0)*({p}$B$7+{p}$B$9+{p}$B$12/2+{p}$B$12+{p}$B$10))/Z{r_idx})", result["pos"]["d"]),
            29: (f"=Z{r_idx}*{p}$B$8", result["pos"]["as_prov"]),
            30: (f"={p}$B$5*AC{r_idx}*{p}$B$3*(AB{r_idx}-(AC{r_idx}*{p}$B$3/(0.85*{p}$B$2*H{r_idx}))/2)/1000000", result["pos"]["phi_mn"]),
            31: (f'=IF(AD{r_idx}>=L{r_idx},"CUMPLE","NO CUMPLE")', "CUMPLE" if result["pos"]["phi_mn"] >= result["Mu_pos"] else "NO CUMPLE"),
            32: (f"=V{r_idx}*1.25*{p}$B$3*(U{r_idx}-(V{r_idx}*1.25*{p}$B$3/(0.85*{p}$B$2*H{r_idx}))/2)/1000000", result["neg"]["mpr"]),
            33: (f"=AC{r_idx}*1.25*{p}$B$3*(AB{r_idx}-(AC{r_idx}*1.25*{p}$B$3/(0.85*{p}$B$2*H{r_idx}))/2)/1000000", result["pos"]["mpr"]),
            34: (f"=N{r_idx}+(AF{r_idx}+AG{r_idx})/D{r_idx}", result["Vcap"]),
            35: (f"=MAX(O{r_idx},AH{r_idx})", result["Vdesign"]),
            36: (f"=0.17*SQRT({p}$B$2)*H{r_idx}*MIN(U{r_idx},AB{r_idx})/1000", result["Vc"]),
            37: (f"=MAX(AI{r_idx}/{p}$B$6-AJ{r_idx},0)", result["Vs_req"]),
            38: (f"=0.66*SQRT({p}$B$2)*H{r_idx}*MIN(U{r_idx},AB{r_idx})/1000", result["Vs_max"]),
            39: (f'=IF(AK{r_idx}<=AL{r_idx},"CUMPLE","NO CUMPLE")', "CUMPLE" if result["shear_section_ok"] else "NO CUMPLE"),
            42: (f"=ROUNDDOWN(MIN(IF(BC{r_idx}+BD{r_idx}=0,1000000000,AO{r_idx}*BB{r_idx}/(BC{r_idx}+BD{r_idx})),MIN(U{r_idx},AB{r_idx})/4,8*{p}$B$12,24*BA{r_idx},300,IF(AV{r_idx}=0,1000000000,2*((H{r_idx}-2*({p}$B$7+{p}$B$9/2))+(I{r_idx}-2*({p}$B$7+{p}$B$9/2)))/8))/10,0)*10", result["s_end"]),
            43: (f"=ROUNDDOWN(MIN(IF(BC{r_idx}+BD{r_idx}=0,1000000000,AO{r_idx}*BB{r_idx}/(BC{r_idx}+BD{r_idx})),IF(AK{r_idx}>0.33*SQRT({p}$B$2)*H{r_idx}*MIN(U{r_idx},AB{r_idx})/1000,MIN(U{r_idx},AB{r_idx})/4,MIN(U{r_idx},AB{r_idx})/2),600,IF(AV{r_idx}=0,1000000000,2*((H{r_idx}-2*({p}$B$7+{p}$B$9/2))+(I{r_idx}-2*({p}$B$7+{p}$B$9/2)))/8))/10,0)*10", result["s_center"]),
            44: (f"={p}$B$6*(AJ{r_idx}+MAX(BE{r_idx}-BD{r_idx},0)*{p}$B$4*MIN(U{r_idx},AB{r_idx})/1000)", result["phi_vn"]),
            45: (f'=IF(AND(ISERROR(SEARCH("referencia",AN{r_idx})),AM{r_idx}="CUMPLE",AR{r_idx}>=AI{r_idx}),"CUMPLE","NO CUMPLE")', "CUMPLE" if result["shear_ok"] else "NO CUMPLE"),
            46: (f"={p}$B$6*0.083*SQRT({p}$B$2)*(H{r_idx}*I{r_idx})^2/(2*(H{r_idx}+I{r_idx}))/1000000", result["phi_tth"]),
            47: (f'=IF(Q{r_idx}<=AT{r_idx},"DESPRECIABLE","DISEÑAR")', "DESPRECIABLE" if result["at_s"] == 0 else "DISEÑAR"),
            48: (f"=IF(Q{r_idx}<=AT{r_idx},0,Q{r_idx}*1000000/({p}$B$6*2*0.85*(H{r_idx}-2*({p}$B$7+{p}$B$9/2))*(I{r_idx}-2*({p}$B$7+{p}$B$9/2))*{p}$B$4))", result["at_s"]),
            49: (f"=AV{r_idx}*2*((H{r_idx}-2*({p}$B$7+{p}$B$9/2))+(I{r_idx}-2*({p}$B$7+{p}$B$9/2)))*{p}$B$4/{p}$B$3", result["Al"]),
            51: (f'=IF(AV{r_idx}=0,"DESPRECIABLE",IF(AND(ISERROR(SEARCH("referencia",AN{r_idx})),BB{r_idx}/AP{r_idx}>=AV{r_idx},BI{r_idx}>=AW{r_idx},BH{r_idx}="CUMPLE"),"CUMPLE","NO CUMPLE"))', "DESPRECIABLE" if result["at_s"] == 0 else ("CUMPLE" if result["torsion_ok"] else "NO CUMPLE")),
            52: (f'=IF(AND(X{r_idx}="CUMPLE",AE{r_idx}="CUMPLE",AS{r_idx}="CUMPLE",OR(AY{r_idx}="CUMPLE",AY{r_idx}="DESPRECIABLE"),BH{r_idx}="CUMPLE",BJ{r_idx}="CUMPLE"{group_gate}),"CUMPLE","NO CUMPLE")',
                 "CUMPLE" if overall_status else "NO CUMPLE"),
            55: (f"=MAX(IF(AK{r_idx}=0,0,AK{r_idx}*1000/({p}$B$4*MIN(U{r_idx},AB{r_idx}))),MAX(0.062*SQRT({p}$B$2)*H{r_idx}/{p}$B$4,0.35*H{r_idx}/{p}$B$4))", result["av_s_required"]),
            56: (f"=2*AV{r_idx}", 2 * result["at_s"]),
            57: (f"=AO{r_idx}*BB{r_idx}/AP{r_idx}", result["av_s_provided"]),
            58: (f"=SQRT((AI{r_idx}*1000/(H{r_idx}*MIN(U{r_idx},AB{r_idx})))^2+(Q{r_idx}*1000000*2*((H{r_idx}-2*({p}$B$7+{p}$B$9/2))+(I{r_idx}-2*({p}$B$7+{p}$B$9/2)))/(1.7*((H{r_idx}-2*({p}$B$7+{p}$B$9/2))*(I{r_idx}-2*({p}$B$7+{p}$B$9/2)))^2))^2)", result["interaction_vt"]),
            59: (f"={p}$B$6*(AJ{r_idx}*1000/(H{r_idx}*MIN(U{r_idx},AB{r_idx}))+0.66*SQRT({p}$B$2))", result["interaction_limit"]),
            60: (f'=IF(BF{r_idx}<=BG{r_idx},"CUMPLE","NO CUMPLE")', "CUMPLE" if result["interaction_ok"] else "NO CUMPLE"),
            61: (f"=AX{r_idx}*{p}$B$8", result["Al_prov"]),
            62: (f'=IF(BI{r_idx}>=AW{r_idx},"CUMPLE","NO CUMPLE")', "CUMPLE" if result["Al_prov"] >= result["Al"] else "NO CUMPLE"),
        }
        for col_idx, (formula, cached) in formulas.items():
            add_formula(ws, f"{get_column_letter(col_idx)}{r_idx}", formula, cached, cache)

        for c_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(r_idx, c_idx)
            cell.border = BORDER
            cell.alignment = CENTER
            if c_idx in (1, 2):
                cell.fill = PatternFill("solid", fgColor=color)
            if c_idx in (8, 9, 19, 26, 40, 41, 50):
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            if c_idx >= 10 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        if result["section_changed"]:
            ws.cell(r_idx, 8).fill = PatternFill("solid", fgColor=ORANGE)
            ws.cell(r_idx, 9).fill = PatternFill("solid", fgColor=ORANGE)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{3 + len(rows)}"
    ws.sheet_view.showGridLines = False
    widths = {1: 12, 2: 12, 3: 10, 4: 11, 5: 11, 6: 12, 7: 12, 8: 12, 9: 12}
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 14)
    ws.conditional_formatting.add(f"X4:X{3+len(rows)}", CellIsRule(operator="equal", formula=['"CUMPLE"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"AE4:AE{3+len(rows)}", CellIsRule(operator="equal", formula=['"CUMPLE"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"AS4:AS{3+len(rows)}", CellIsRule(operator="equal", formula=['"CUMPLE"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"AZ4:AZ{3+len(rows)}", CellIsRule(operator="equal", formula=['"NO CUMPLE"'], fill=PatternFill("solid", fgColor=RED)))
    return ws


def write_parameters(wb):
    ws = wb.create_sheet("Parametros")
    ws["A1"] = "Parámetros generales"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:D1")
    data = [
        ("f'c", FC, "MPa", "Concreto del proyecto"),
        ("fy", FY, "MPa", "Acero longitudinal"),
        ("fyt", FYT, "MPa", "Acero transversal"),
        ("φ flexión", PHI_M, "—", "Secciones controladas por tracción"),
        ("φ cortante/torsión", PHI_V, "—", "NSR-10 Título C"),
        ("Recubrimiento", COVER, "mm", "Vigas interiores"),
        ("Área barra longitudinal", AB_LONG, "mm²", "Nº5"),
        ("db estribo para d", DB_ST, "mm", "Nº4, conservador; el estribo adoptado se calcula por fila"),
        ("Separación libre vertical", CLEAR_BAR, "mm", "Entre capas"),
        ("Área rama de referencia", AB_ST, "mm²", "Nº4; cada fila muestra el área realmente adoptada"),
        ("db longitudinal", DB_LONG, "mm", "Barra Nº5"),
        ("R₀ del sistema DMO", R0, "—", "Valor básico adoptado"),
        ("φa", PHI_A, "—", "Irregularidad en altura"),
        ("φp", PHI_P, "—", "Irregularidad en planta"),
        ("φr", PHI_R, "—", "Ausencia de redundancia"),
        ("R = R₀·φa·φp·φr", R_DESIGN, "—", "Coeficiente usado para reducir Ex y Ey en ENVFLEX"),
    ]
    for r, row in enumerate(data, 2):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = CENTER if c != 4 else LEFT
            if c == 1:
                ws.cell(r, c).font = Font(bold=True)
                ws.cell(r, c).fill = PatternFill("solid", fgColor=BLUE)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 48
    ws.sheet_view.showGridLines = False


def write_cover(wb, group_results, all_results):
    ws = wb.active
    ws.title = "Portada y criterio"
    ws.merge_cells("A1:H1")
    ws["A1"] = "Verificación de vigas - resumen del cálculo"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = LEFT
    ws.merge_cells("A2:H2")
    ws["A2"] = "Hoja de cálculo del proyecto de Diseño Estructural - Grupo 6"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=TEXT)
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A2"].alignment = LEFT
    lines = [
        (3, "Alcance", "592 vigas de los 12 grupos SAP; diseño individual y resumen por grupo. No incluye viguetas."),
        (4, "Fuente", "resultados sap.xlsx, SAP2000 v26, unidades kN, m, C; ENVFLEX y ENVCORT con máximos y mínimos."),
        (5, "Flexión y cortante", "Mu se toma de ENVFLEX. Vu diseño = máx(Vu ENVCORT; Ve DMO), sin volver a dividir las fuerzas por R."),
        (6, "Verificación DMO", "Ve = Vu_ENVFLEX + (Mpr−+Mpr+)/Ln, con Ln conservadora = máx(L estación−0.60 m; 0.50L)."),
        (7, "Torsión", "Tu = máx(|T| ENVFLEX; |T| ENVCORT), sin reducción automática a φTcr. Los 100 casos que exceden esa referencia requieren demostrar compatibilidad y reanalizar la redistribución antes de cualquier reducción."),
        (8, "Acero de torsión", "Al se provee con barras Nº5 dedicadas, adicionales al acero requerido por flexión y distribuidas alrededor del perímetro; no se acredita doblemente As superior/inferior."),
        (9, "Agrupación", "Cada grupo conserva color propio. El resumen usa el frame de mayor utilización real y evita combinar máximos incompatibles de vigas diferentes."),
        (10, "Secciones SAP", "VC1–VC7 y VRAUX: 0.45×0.55 m. VR1, VR N1, VR2 y VR3: 0.50×0.55 m."),
        (11, "Resultado", f"{sum(x['overall'] for x in all_results)} de 592 vigas cumplen con torsión bruta; {sum(not x['overall'] for x in all_results)} quedan NO CUMPLE y requieren rediseño/reanálisis SAP. El expediente completo es NO EMITIR PARA CONSTRUCCIÓN."),
    ]
    for row, title, text in lines:
        ws.cell(row, 1, title).font = Font(bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 2, text).alignment = LEFT
        for col in range(1, 9):
            ws.cell(row, col).border = BORDER
    ws["A12"] = "Resultados de control"
    ws["A12"].font = Font(bold=True, color=WHITE)
    ws["A12"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A12:H12")
    summary = [
        ("Frames exportados/diseñados", 592), ("Vigas diseñadas", len(all_results)),
        ("Frames que exceden referencia φTcr", sum(x["requires_compatibility_assessment"] for x in all_results)), ("Grupos", len(group_results)),
        ("Grupos que cumplen", sum(x["group_all_ok"] for x in group_results)),
        ("Vigas que cumplen", sum(x["overall"] for x in all_results)),
        ("Grupos con observación", sum(not x["group_all_ok"] for x in group_results)),
        ("Vigas con observación", sum(not x["overall"] for x in all_results)),
    ]
    for idx, (label, value) in enumerate(summary, 13):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)
        ws.cell(idx, 2).font = Font(bold=True)
        ws.cell(idx, 2).fill = PatternFill("solid", fgColor=GREEN if "cumplen" in label else YELLOW)
        for col in (1, 2): ws.cell(idx, col).border = BORDER
    ws.column_dimensions["A"].width = 32
    for col in range(2, 9): ws.column_dimensions[get_column_letter(col)].width = 19
    ws.sheet_view.showGridLines = False


def write_justification(wb, group_results, all_results):
    ws = wb.create_sheet("Justificacion DMO")
    ws.merge_cells("A1:L1")
    ws["A1"] = "Revisión de cortante - sistema DMO (R = 5)"
    ws["A1"].font = Font(bold=True, size=15, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = CENTER

    explanations = [
        (3, "Decisión de diseño", "Se adopta directamente ENVCORT y, adicionalmente, se verifica el cortante por capacidad DMO. No se reduce ni se descarta ningún valor de la envolvente de cortante."),
        (4, "Reducción sísmica", "R = R₀·φa·φp·φr = 5·1·1·1 = 5, incorporado por el equipo en las combinaciones suministradas."),
        (5, "Flexión y torsión", "Mu− y Mu+ se obtienen de ENVFLEX. Tu de diseño es el mayor |T| de ENVFLEX y ENVCORT, sin recorte a φTcr porque no se aportó clasificación ni reanálisis de redistribución."),
        (6, "Cortante directo", "Vu,ENVCORT = máximo |V2| de ENV CORTANTE y se considera íntegramente como demanda."),
        (7, "Cortante por capacidad", "Ve,DMO = Vg + (Mpr− + Mpr+)/Ln. Se adopta provisionalmente Ln = máx(L estación−0.60 m; 0.50L), hipótesis conservadora hasta confirmar las caras de apoyo en SAP. |V2| de ENVFLEX aproxima Vg."),
        (8, "Demanda adoptada", "Vu,diseño = máx(Vu,ENVCORT; Ve,DMO). Por tanto, el diseño nunca queda por debajo del cortante directo suministrado."),
        (9, "Cortante y torsión", "El estribo se diseña con máx(Av/s por resistencia; Av/s mínimo) + 2At/s; se verifican interacción V-T y límites de separación. Al se provee con barras dedicadas adicionales a flexión, distribuidas en el perímetro."),
        (10, "Secciones SAP", "VC1–VC7 y VRAUX usan 450×550 mm; VR1, VR N1, VR2 y VR3 usan 500×550 mm. Son las secciones efectivamente reanalizadas en resultados sap.xlsx."),
        (11, "Resultado", f"{sum(x['overall'] for x in all_results)}/592 vigas cumplen con torsión bruta y {sum(not x['overall'] for x in all_results)} quedan NO CUMPLE. Todo el expediente se marca NO EMITIR PARA CONSTRUCCIÓN hasta rediseño y reanálisis SAP."),
    ]
    for row, title, text in explanations:
        ws.cell(row, 1, title).font = Font(bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        ws.cell(row, 2, text).alignment = LEFT
        for col in range(1, 13):
            ws.cell(row, col).border = BORDER

    headers = ["Grupo", "N", "Vu ENVFLEX", "Ve DMO conserv.", "Vu ENVCORT",
               "Vu adoptado", "φVn provisto", "φVn máx. sección", "ENVCORT directo",
               "Chequeo DMO", "Sección adoptada", "Conclusión"]
    header_row = 14
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 42

    for row, result in enumerate(group_results, header_row + 1):
        direct = "CUMPLE DIRECTO" if result["direct_elastic_ok"] else "EXCEDE DIRECTO"
        failed_count = sum(1 for x in all_results if x["group"] == result["group"] and not x["overall"])
        dmo = "CUMPLE" if result["group_all_ok"] else "NO CUMPLE"
        conclusion = "CUMPLE" if result["group_all_ok"] else f"{failed_count} FRAMES NO CUMPLEN"
        values = [result["group"], result["count"], result["Vu_flex"], result["Vcap"],
                  result["Vu_elastic"], result["Vdesign"], result["phi_vn"], result["phi_vmax"],
                  direct, dmo, f"{int(result['b'])}×{int(result['h'])} mm", conclusion]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = CENTER
            if isinstance(value, float): cell.number_format = "0.0"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS[result["group"]])
        ws.cell(row, 9).fill = PatternFill("solid", fgColor=GREEN if result["direct_elastic_ok"] else ORANGE)
        ws.cell(row, 10).fill = PatternFill("solid", fgColor=GREEN if result["group_all_ok"] else RED)
        ws.cell(row, 12).fill = PatternFill("solid", fgColor=GREEN if result["group_all_ok"] else RED)

    widths = [14, 8, 14, 17, 14, 15, 15, 18, 19, 15, 19, 22]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A15"
    ws.sheet_view.showGridLines = False


def write_torsion_compatibility(wb, all_results):
    """Traza la demanda bruta y señala casos que requieren estudio de compatibilidad."""
    ws = wb.create_sheet("Torsion demanda bruta")
    ws.merge_cells("A1:J1")
    ws["A1"] = "Revisión de torsión sin redistribución"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = CENTER
    ws.merge_cells("A2:J2")
    ws["A2"] = ("Tu diseño = máx(|T| ENVFLEX, |T| ENVCORT), sin limitar a φTcr. Los casos que exceden "
                  "φTcr quedan marcados para demostrar compatibilidad y reanalizar la redistribución antes de reducir Tu.")
    ws["A2"].alignment = LEFT
    ws["A2"].fill = PatternFill("solid", fgColor=YELLOW)
    headers = ["Grupo", "Frame", "Tu ENVFLEX", "Tu ENVCORT", "Tu bruto", "φTcr referencia",
               "Tu diseño", "Requiere estudio compat.", "Sección (mm)", "Estado seccional"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER
        cell.border = BORDER
    for row, result in enumerate(all_results, 5):
        values = [result["group"], result["frame"], result["Tu_flex"], result["Tu_cort"],
                  result["Tu_raw"], result["phi_tcr_compat"], result["Tu"],
                  "SÍ" if result["requires_compatibility_assessment"] else "NO",
                  f"{int(result['b'])}×{int(result['h'])}",
                  "CUMPLE" if result["overall"] else "NO CUMPLE"]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = CENTER
            if isinstance(value, float):
                cell.number_format = "0.00"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS[result["group"]])
        if result["requires_compatibility_assessment"]:
            ws.cell(row, 8).fill = PatternFill("solid", fgColor=ORANGE)
        ws.cell(row, 10).fill = PatternFill("solid", fgColor=GREEN if result["overall"] else RED)
    for col, width in enumerate([13, 11, 15, 15, 13, 16, 13, 14, 17, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{4+len(all_results)}"
    ws.sheet_view.showGridLines = False


def write_frame_traceability(wb, all_results):
    """Añade localización tabular por frame y un registro explícito de pendientes."""
    ws = wb.create_sheet("Trazabilidad frames")
    headers = [
        "Grupo", "Frame", "Joint I", "Joint J", "L eje-eje (m)",
        "Centroide X (m)", "Centroide Y (m)", "Nivel Z (m)", "Sección (mm)",
        "Nº5 sup.", "Nº5 inf.", "Nº5 torsión dedicadas", "Estribo", "Ramas",
        "s extremo (mm)", "s centro (mm)", "Interacción V-T", "Estribo práctico",
        "Estado", "Acción requerida",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CENTER
        cell.border = BORDER
    for row, result in enumerate(all_results, 2):
        reasons = []
        if not result["shear_section_ok"]:
            reasons.append("aumentar sección por cortante")
        if not result["interaction_ok"]:
            reasons.append("aumentar sección por interacción V-T")
        if not result["practical_stirrup_found"]:
            reasons.append("rediseñar sección/estribo")
        if result["requires_compatibility_assessment"]:
            reasons.append("demostrar compatibilidad antes de reducir Tu")
        if not reasons:
            reasons.append("coordinar con planta y reanálisis global")
        values = [
            result["group"], result["frame"], result["joint_i"], result["joint_j"], result["L"],
            result["centroid_x"], result["centroid_y"], result["centroid_z"],
            f"{int(result['b'])}×{int(result['h'])}", result["neg"]["n"], result["pos"]["n"],
            result["n_torsion"], result["stirrup_label"], result["legs"], result["s_end"],
            result["s_center"], result["interaction_vt"] / result["interaction_limit"],
            "SÍ" if result["practical_stirrup_found"] else "NO",
            "CUMPLE" if result["overall"] else "NO CUMPLE", "; ".join(reasons),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = BORDER
            cell.alignment = CENTER if col != 20 else LEFT
            if isinstance(value, float):
                cell.number_format = "0.000"
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS[result["group"]])
        ws.cell(row, 19).fill = PatternFill("solid", fgColor=GREEN if result["overall"] else RED)
    for col, width in enumerate([12, 10, 12, 12, 13, 14, 14, 12, 16, 10, 10, 18, 25, 9, 13, 13, 15, 15, 14, 46], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{1 + len(all_results)}"
    ws.sheet_view.showGridLines = False


def write_audit(wb, group_results, all_results, group_frames):
    ws = wb.create_sheet("Auditoria y fuentes")
    ws.merge_cells("A1:J1")
    ws["A1"] = "Revisión de datos y resultados"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = CENTER
    audit = [
        ("Fuente SAP", "resultados sap.xlsx — SAP2000 v26, unidades kN, m, C"),
        ("Filas Element Forces - Frames", "20.656"),
        ("Vigas incluidas en grupos", "592"),
        ("Envolventes", "ENVFLEX y ENVCORT, máximos y mínimos"),
        ("Grupos con estado NO CUMPLE", str(sum(not x["group_all_ok"] for x in group_results))),
        ("Vigas con estado NO CUMPLE", str(sum(not x["overall"] for x in all_results))),
        ("Criterio de torsión", "Tu bruto de ambas envolventes, sin recorte automático a φTcr."),
        ("Criterio de Al", "Barras Nº5 dedicadas, adicionales al acero requerido por flexión."),
        ("Condición de emisión", "NO EMITIR PARA CONSTRUCCIÓN: existen incumplimientos y falta reanálisis/plantas de localización."),
        ("Criterio de secciones", "VC1–VC7 y VRAUX: 450×550 mm; VR1, VR N1, VR2 y VR3: 500×550 mm, ya reanalizadas en SAP."),
    ]
    for r, (a, b) in enumerate(audit, 3):
        ws.cell(r, 1, a); ws.cell(r, 2, b)
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(r, 1).border = ws.cell(r, 2).border = BORDER
        ws.cell(r, 2).alignment = LEFT

    start = 12
    headers = ["Archivo", "SHA-256"]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h); ws.cell(start, c).font = Font(bold=True, color=WHITE); ws.cell(start, c).fill = PatternFill("solid", fgColor=NAVY)
    for i, path in enumerate((SAP_FILE,), start + 1):
        ws.cell(i, 1, path.name); ws.cell(i, 2, sha256(path))
        ws.cell(i, 1).border = ws.cell(i, 2).border = BORDER

    start = 18
    headers = ["Grupo", "N", "L estación mín", "L estación máx", "Frame Mu−", "Frame Mu+", "Frame Vu", "Estado", "Sección adoptada", "Nota"]
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h); ws.cell(start, c).font = Font(bold=True, color=WHITE); ws.cell(start, c).fill = PatternFill("solid", fgColor=NAVY); ws.cell(start, c).border = BORDER
    for r, result in enumerate(group_results, start + 1):
        values = [result["group"], result["count"], result["L_min"], result["L_max"], result["frame_Mu_neg"], result["frame_Mu_pos"], result["frame_Vu"],
                  "CUMPLE" if result["group_all_ok"] else "NO CUMPLE", f"{int(result['b'])}×{int(result['h'])} mm",
                  "NO EMITIR: contiene frames no conformes" if not result["group_all_ok"] else "Verificado, sujeto a coordinación global"]
        for c, v in enumerate(values, 1):
            ws.cell(r, c, v); ws.cell(r, c).border = BORDER; ws.cell(r, c).alignment = CENTER
            if c == 1: ws.cell(r, c).fill = PatternFill("solid", fgColor=COLORS[result["group"]])
    for col, width in enumerate([18, 10, 12, 12, 14, 14, 14, 16, 20, 26], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False


def apply_personal_format(wb):
    """Deja el libro sin colores decorativos ni filtros; conserva solo colores de grupo."""
    design_sheets = {"Vigas de Carga (7)", "Vigas de Rigidez (5)", "Todas las Vigas (592)"}
    large_sheets = {"Todas las Vigas (592)", "Torsion demanda bruta", "Trazabilidad frames"}

    for ws in wb.worksheets:
        # Sin colores en pestañas ni flechas desplegables de AutoFilter.
        ws.sheet_properties.tabColor = None
        ws.sheet_properties.filterMode = False
        ws.auto_filter.ref = None
        ws.sheet_view.showGridLines = True
        ws.sheet_view.zoomScale = 80 if ws.title in large_sheets else 90
        ws.sheet_view.zoomScaleNormal = 100
        ws.sheet_format.defaultRowHeight = 18
        ws.conditional_formatting._cf_rules.clear()

        # Fondo blanco y texto negro en todo el libro. Se conservan bordes,
        # alineaciones, fórmulas, formatos numéricos y tamaños de columnas.
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = PatternFill()
                if cell.value is None:
                    continue
                font = copy(cell.font)
                font.name = "Arial"
                font.color = TEXT
                if font.sz is None:
                    font.sz = 10
                cell.font = font

        # Títulos y encabezados sencillos: negrita negra sobre fondo blanco.
        merged_title = any(
            rng.min_row == 1 and rng.max_row == 1 and rng.min_col == 1
            for rng in ws.merged_cells.ranges
        )
        if merged_title:
            title = ws["A1"]
            title.font = Font(name="Arial", bold=True, size=14, color=TEXT)
            title.alignment = LEFT
            ws.row_dimensions[1].height = 24

        # Único color permitido: identificación del grupo. En las hojas de
        # diseño se colorean Grupo y Frame; en las demás, solamente Grupo.
        for row_index in range(1, ws.max_row + 1):
            group = ws.cell(row_index, 1).value
            if group not in COLORS:
                continue
            ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=COLORS[group])
            if ws.title in design_sheets:
                ws.cell(row_index, 2).fill = PatternFill("solid", fgColor=COLORS[group])

        if ws.title in design_sheets:
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 52
        elif ws.title == "Torsion demanda bruta":
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[4].height = 34
        elif ws.title == "Trazabilidad frames":
            ws.row_dimensions[1].height = 36
        elif ws.title == "Portada y criterio":
            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 22



def inject_cached_values(path: Path, cache: dict):
    """Inserta resultados precalculados sin eliminar las fórmulas."""
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
          "pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        with zipfile.ZipFile(path, "r") as zin:
            zin.extractall(tmp_path)
        wb_tree = etree.parse(str(tmp_path / "xl/workbook.xml"))
        rel_tree = etree.parse(str(tmp_path / "xl/_rels/workbook.xml.rels"))
        rel_targets = {r.get("Id"): r.get("Target") for r in rel_tree.xpath("//pr:Relationship", namespaces=ns)}
        sheet_paths = {}
        for sheet in wb_tree.xpath("//m:sheets/m:sheet", namespaces=ns):
            rid = sheet.get(f"{{{ns['r']}}}id")
            target = rel_targets[rid].lstrip("/")
            sheet_paths[sheet.get("name")] = tmp_path / target
        for sheet_name, sheet_path in sheet_paths.items():
            items = {coord: val for (s, coord), val in cache.items() if s == sheet_name}
            if not items:
                continue
            tree = etree.parse(str(sheet_path))
            for cell in tree.xpath("//m:c[m:f]", namespaces=ns):
                coord = cell.get("r")
                if coord not in items:
                    continue
                value = items[coord]
                v_nodes = cell.xpath("m:v", namespaces=ns)
                v = v_nodes[0] if v_nodes else etree.SubElement(cell, f"{{{ns['m']}}}v")
                if isinstance(value, str):
                    cell.set("t", "str")
                    v.text = value
                else:
                    cell.attrib.pop("t", None)
                    v.text = f"{float(value):.15g}"
            tree.write(str(sheet_path), xml_declaration=True, encoding="UTF-8", standalone=True)
        calc = wb_tree.xpath("//m:calcPr", namespaces=ns)
        if calc:
            calc[0].set("calcMode", "auto")
            calc[0].set("fullCalcOnLoad", "1")
            calc[0].set("forceFullCalc", "1")
        wb_tree.write(str(tmp_path / "xl/workbook.xml"), xml_declaration=True, encoding="UTF-8", standalone=True)
        new_path = path.with_suffix(".tmp.xlsx")
        with zipfile.ZipFile(new_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for file in tmp_path.rglob("*"):
                if file.is_file(): zout.write(file, file.relative_to(tmp_path))
        shutil.move(new_path, path)


def main():
    records, group_frames = extract_inputs()
    all_results = []
    group_order = GROUP_ORDER
    for record in sorted(records, key=lambda x: (group_order.index(x["group"]), sort_frame(x["frame"]))):
        all_results.append(design_record(record))
    group_results = build_group_records(records, all_results)

    cache = {}
    wb = Workbook()
    write_cover(wb, group_results, all_results)
    write_parameters(wb)
    write_justification(wb, group_results, all_results)
    write_design_sheet(wb, "Vigas de Carga (7)", [x for x in group_results if x["group"].startswith("VC")], True, cache)
    write_design_sheet(wb, "Vigas de Rigidez (5)", [x for x in group_results if x["group"].startswith("VR")], True, cache)
    write_design_sheet(wb, "Todas las Vigas (592)", all_results, False, cache)
    write_torsion_compatibility(wb, all_results)
    write_frame_traceability(wb, all_results)
    write_audit(wb, group_results, all_results, group_frames)
    apply_personal_format(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT_FILE)
    inject_cached_values(OUTPUT_FILE, cache)

    print(f"OK: {OUTPUT_FILE}")
    print(f"Grupos: {len(group_results)}; vigas: {len(all_results)}")
    print(f"Grupos que cumplen: {sum(x['group_all_ok'] for x in group_results)}/{len(group_results)}")
    print(f"Vigas que cumplen: {sum(x['overall'] for x in all_results)}/{len(all_results)}")
    print(f"Celdas con cache: {len(cache)}")


if __name__ == "__main__":
    main()
