#!/usr/bin/env python3
"""Genera dos archivos CAD B2 con las vigas completas por ejes y tramos.

La geometría y continuidad se reconstruyen desde resultados sap.xlsx y
geomatria sap.xlsx. Los armados por grupo se leen del Excel final del proyecto.
Cada familia se entrega en un único DXF con una presentación B2 por grupo:
VC1–VC7 en V-01 y VR1, VR N1, VR2, VR3 y VRAUX en V-02.
"""
from __future__ import annotations

import math
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import ezdxf
from ezdxf.enums import TextEntityAlignment
from ezdxf.math import Matrix44
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
SAP = BASE / "resultados sap.xlsx"
GEOMETRY = BASE / "geomatria sap.xlsx"
DESIGN = BASE / "Diseño de vigas proyecto diseño DEF.xlsx"
OUT = BASE / "Planos-Autocad-B2-Vigas-Continuas"
ZIP = OUT / "PLANOS-AUTOCAD-B2-VIGAS-CONTINUAS.zip"

SHEET_W, SHEET_H = 707.0, 500.0
MARGIN = 7.0
FC, FY = 28.0, 420.0
DB_LONG = 15.9
COVER = 40.0
LD_INF, LD_SUP, LDH = 370.0, 480.0, 305.0
TAIL_90, BEND_90 = 195.0, 100.0
DATE = "18-07-2026"

VC_GROUPS = ["VC1", "VC2", "VC3", "VC4", "VC5", "VC6", "VC7"]
VR_GROUPS = ["VR1", "VR N1", "VR2", "VR3", "VRAUX"]
GROUPS = VC_GROUPS + VR_GROUPS
# Cuatro láminas conservan la legibilidad B2: cada grupo aparece una sola vez,
# separado por familia estructural (carga y rigidez).
VC_SHEETS = [VC_GROUPS[:4], VC_GROUPS[4:]]
VR_SHEETS = [VR_GROUPS[:3], VR_GROUPS[3:]]
ALIASES = {"VR1 N1": "VR N1", "VR AUX": "VRAUX"}

# AutoCAD ACI colors.
BLACK = 7
RED = 1
YELLOW = 2
GREEN = 3
CYAN = 4
BLUE = 5
MAGENTA = 6
GRAY = 8
BROWN = 30
LIGHT = 9


def normalize_group(value) -> str:
    text = str(value).strip()
    return ALIASES.get(text, text)


def read_sap_table(workbook, sheet: str):
    ws = workbook[sheet]
    headers = [cell.value for cell in ws[2]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=4, values_only=True)
            if row and row[0] is not None]


def load_sources():
    sap = load_workbook(SAP, read_only=True, data_only=True)
    geometry = load_workbook(GEOMETRY, read_only=True, data_only=True)

    grids = [row for row in read_sap_table(sap, "Grid Lines")
             if row.get("CoordSys") == "GLOBAL"]
    # SAP no exportó la línea 15 en la tabla, pero aparece en las plantas y en
    # la geometría a Y=37.70 m.
    grids.append({"CoordSys": "GLOBAL", "AxisDir": "Y", "GridID": "15", "XRYZCoord": 37.70})

    joints = {
        str(row["Joint"]): (
            float(row["GlobalX"]), float(row["GlobalY"]), float(row["GlobalZ"])
        )
        for row in read_sap_table(geometry, "Joint Coordinates")
    }
    frames = {str(row["Frame"]): row for row in read_sap_table(geometry, "Connectivity - Frame")}
    assignments = defaultdict(set)
    for row in read_sap_table(sap, "Groups 2 - Assignments"):
        if str(row.get("ObjectType")) != "Frame":
            continue
        group = normalize_group(row.get("GroupName"))
        if group in GROUPS:
            assignments[group].add(str(row["ObjectLabel"]))
    sap.close()
    geometry.close()

    if sum(len(assignments[group]) for group in GROUPS) != 592:
        raise ValueError("La asignación de grupos no contiene las 592 vigas esperadas")
    if any(not assignments[group] for group in GROUPS):
        raise ValueError("Falta al menos un grupo de vigas en el export SAP")

    design = load_workbook(DESIGN, read_only=True, data_only=True)
    demands = {}
    detail_ws = design["Todas las Vigas (592)"]
    detail_headers = [cell.value for cell in detail_ws[3]]
    for values in detail_ws.iter_rows(min_row=4, values_only=True):
        if values[0] is None:
            continue
        row = dict(zip(detail_headers, values))
        frame = str(row["Frame"])
        if frame in demands:
            raise ValueError(f"Frame duplicado en el Excel de diseño: {frame}")
        demands[frame] = {
            "group": normalize_group(row["Grupo"]),
            "mu_neg": float(row["Mu− ENVFLEX (kN·m)"]),
            "mu_pos": float(row["Mu+ ENVFLEX (kN·m)"]),
            "vu_flex": float(row["Vu ENVFLEX (kN)"]),
            "vu_shear": float(row["Vu ENVCORT (kN)"]),
            "torsion": float(row["Tu diseño compat. (kN·m)"]),
            "status": str(row["ESTADO"]),
        }

    expected_frames = set().union(*(assignments[group] for group in GROUPS))
    if set(demands) != expected_frames:
        missing = sorted(expected_frames - set(demands), key=int)
        extra = sorted(set(demands) - expected_frames, key=int)
        raise ValueError(f"Excel/SAP no coinciden; faltan={missing[:10]}, sobran={extra[:10]}")
    for frame, demand in demands.items():
        expected_group = next(group for group in GROUPS if frame in assignments[group])
        if demand["group"] != expected_group:
            raise ValueError(
                f"Grupo inconsistente para frame {frame}: Excel={demand['group']}, SAP={expected_group}"
            )
        if demand["status"] not in {"CUMPLE", "NO CUMPLE"}:
            raise ValueError(f"Estado no calculado para frame {frame}: {demand['status']}")

    details = {}
    for sheet in ("Vigas de Carga (7)", "Vigas de Rigidez (5)"):
        ws = design[sheet]
        headers = [cell.value for cell in ws[3]]
        for values in ws.iter_rows(min_row=4, values_only=True):
            if values[0] is None:
                continue
            row = dict(zip(headers, values))
            group = normalize_group(row["Grupo"])
            group_demands = [item for item in demands.values() if item["group"] == group]
            details[group] = {
                "group": group,
                "b": float(row["b adopt. (mm)"]),
                "h": float(row["h adopt. (mm)"]),
                "n_sup": int(row["Nº5 sup"]),
                "n_inf": int(row["Nº5 inf"]),
                "n_tor": int(row["Nº5 torsión dedicadas"]),
                "stirrup": str(row["Estribo"]).replace(" (referencia)", ""),
                "legs": int(row["Ramas"]),
                "s_end": int(round(float(row["s extremo DMO (mm; zona 2h)"]))),
                "s_center": int(round(float(row["s centro (mm)"]))),
                "db_st": float(row["db estribo (mm)"]),
                "count": int(row["N vigas"]),
                "status": str(row["ESTADO"]),
                "noncompliant": sum(item["status"] == "NO CUMPLE" for item in group_demands),
            }
    design.close()
    if len(demands) != 592:
        raise ValueError(f"El Excel final contiene {len(demands)} demandas; se esperaban 592")
    if set(details) != set(GROUPS):
        raise ValueError(f"Los grupos del Excel final no coinciden: {sorted(details)}")
    return grids, joints, frames, assignments, details, demands


def nearest_axis(grids, direction: str, coordinate: float, tolerance=0.22) -> str:
    candidates = [
        (abs(float(row["XRYZCoord"]) - coordinate), str(row["GridID"]))
        for row in grids if row.get("AxisDir") == direction
    ]
    if not candidates:
        return f"{coordinate:.2f}"
    distance, label = min(candidates)
    return label if distance <= tolerance else f"{coordinate:.2f}"


def frame_components(group, assignments, frames, joints, grids, demands):
    ids = [frame for frame in assignments[group] if frame in frames]
    z_values = sorted({round(float(frames[frame]["CentroidZ"]), 3) for frame in ids})
    z = z_values[0]
    level = [frame for frame in ids if abs(float(frames[frame]["CentroidZ"]) - z) < 0.02]
    by_joint = defaultdict(list)
    for frame in level:
        by_joint[str(frames[frame]["JointI"])].append(frame)
        by_joint[str(frames[frame]["JointJ"])].append(frame)

    seen, components = set(), []
    for initial in sorted(level, key=lambda value: int(value)):
        if initial in seen:
            continue
        stack, found = [initial], []
        seen.add(initial)
        while stack:
            frame = stack.pop()
            found.append(frame)
            for joint in (str(frames[frame]["JointI"]), str(frames[frame]["JointJ"])):
                for neighbor in by_joint[joint]:
                    if neighbor not in seen:
                        seen.add(neighbor)
                        stack.append(neighbor)
        components.append(found)

    chains = []
    for component in components:
        local = defaultdict(list)
        degree = Counter()
        for frame in component:
            i, j = str(frames[frame]["JointI"]), str(frames[frame]["JointJ"])
            local[i].append(frame); local[j].append(frame)
            degree[i] += 1; degree[j] += 1
        if max(degree.values()) > 2:
            raise ValueError(f"El grupo {group} forma una red ramificada en el nivel {z}")
        ends = [joint for joint, value in degree.items() if value == 1]
        start = min(ends or degree, key=lambda joint: (joints[joint][0], joints[joint][1]))
        used, ordered, current = set(), [], start
        while len(used) < len(component):
            options = [frame for frame in local[current] if frame not in used]
            if not options:
                break
            frame = options[0]
            used.add(frame)
            i, j = str(frames[frame]["JointI"]), str(frames[frame]["JointJ"])
            following = j if i == current else i
            ordered.append({
                "frame": frame,
                "joint_i": current,
                "joint_j": following,
                "length": float(frames[frame]["Length"]) * 1000.0,
                "demand": demands[frame],
            })
            current = following
        coordinates = [joints[ordered[0]["joint_i"]]] + [joints[item["joint_j"]] for item in ordered]
        dx = max(point[0] for point in coordinates) - min(point[0] for point in coordinates)
        dy = max(point[1] for point in coordinates) - min(point[1] for point in coordinates)
        direction = "X" if dx >= dy else "Y"
        axis_labels = [nearest_axis(grids, direction, point[0] if direction == "X" else point[1])
                       for point in coordinates]
        constant_direction = "Y" if direction == "X" else "X"
        constant_coordinate = sum(point[1] if direction == "X" else point[0] for point in coordinates) / len(coordinates)
        constant_axis = nearest_axis(grids, constant_direction, constant_coordinate)
        chains.append({
            "group": group,
            "z": z,
            "all_levels": z_values,
            "segments": ordered,
            "axes": axis_labels,
            "line_axis": constant_axis,
            "direction": direction,
            "coordinates": coordinates,
        })

    # Las vigas repetidas en ejes paralelos se muestran juntas cuando tienen
    # exactamente los mismos tramos. VRAUX se reduce a cuatro variantes de tramo.
    if group == "VRAUX":
        unique = {}
        for chain in chains:
            key = (tuple(round(item["length"]) for item in chain["segments"]), tuple(chain["axes"]))
            if key not in unique:
                unique[key] = chain
            else:
                unique[key]["line_axis"] += f"/{chain['line_axis']}"
        chains = list(unique.values())
    elif group in ("VR1", "VR N1"):
        unique = {}
        for chain in chains:
            key = (tuple(round(item["length"]) for item in chain["segments"]), tuple(chain["axes"]))
            if key not in unique:
                unique[key] = chain
            else:
                unique[key]["line_axis"] += f"/{chain['line_axis']}"
        chains = list(unique.values())
    return sorted(chains, key=lambda chain: (chain["axes"][0], chain["line_axis"]))


def add_layers(doc):
    colors = {
        "MARCO": BLACK, "CONTORNO": BLACK, "TEXTO": BLACK, "EJES": RED,
        "APOYOS": LIGHT, "ACERO_SUP_CONT": CYAN, "ACERO_SUP_APOYO": BLUE,
        "ACERO_INF": RED, "ACERO_TORSION": BROWN, "ESTRIBOS": GRAY,
        "COTAS": GREEN, "CORTE": MAGENTA, "NOTAS": BLACK,
    }
    for name, color in colors.items():
        if name not in doc.layers:
            doc.layers.add(name, color=color)
    if "DASHED" not in doc.linetypes:
        doc.linetypes.add("DASHED", pattern=[0.6, 0.3, -0.3], description="Dashed")


def add_text(space, text, point, height=2.0, layer="TEXTO", color=None,
             align=TextEntityAlignment.LEFT, rotation=0.0):
    attribs = {"height": height, "layer": layer, "rotation": rotation}
    if color is not None:
        attribs["color"] = color
    entity = space.add_text(str(text), dxfattribs=attribs)
    entity.set_placement(point, align=align)
    return entity


def line(space, start, end, layer="CONTORNO", color=None, weight=18, linetype=None):
    attribs = {"layer": layer, "lineweight": weight}
    if color is not None:
        attribs["color"] = color
    if linetype:
        attribs["linetype"] = linetype
    return space.add_line(start, end, dxfattribs=attribs)


def polyline(space, points, layer="CONTORNO", color=None, weight=25, close=False):
    attribs = {"layer": layer, "lineweight": weight}
    if color is not None:
        attribs["color"] = color
    return space.add_lwpolyline(points, close=close, dxfattribs=attribs)


def rectangle(space, x, y, width, height, layer="CONTORNO", color=None, weight=18):
    return polyline(space, [(x, y), (x + width, y), (x + width, y + height), (x, y + height)],
                    layer, color, weight, True)


def dimension(space, x0, x1, y, text, height=1.8):
    line(space, (x0, y), (x1, y), "COTAS", GREEN, 13)
    line(space, (x0, y - 1.2), (x0, y + 1.2), "COTAS", GREEN, 13)
    line(space, (x1, y - 1.2), (x1, y + 1.2), "COTAS", GREEN, 13)
    add_text(space, text, ((x0 + x1) / 2, y + 1.0), height, "COTAS", GREEN,
             TextEntityAlignment.MIDDLE_CENTER)


def choose_scale(total_length: float, max_width=600.0):
    for scale in (10, 20, 25, 30, 40, 50, 60, 75, 100):
        if total_length / scale <= max_width:
            return scale
    return math.ceil(total_length / max_width / 25) * 25


def draw_info_box(space, chain, detail, x, y, width=40.0, height=50.0):
    rectangle(space, x, y, width, height, "MARCO", BLACK, 25)
    add_text(space, f"VIGA {detail['group']}", (x + width / 2, y + height - 6), 2.8,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"EJE {chain['line_axis']}", (x + width / 2, y + height - 13), 2.5,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"{int(detail['b'])}x{int(detail['h'])} mm", (x + width / 2, y + height - 20), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"SUP {detail['n_sup']}#5", (x + 3, y + height - 28), 2.0)
    add_text(space, f"INF {detail['n_inf']}#5", (x + 3, y + height - 34), 2.0)
    add_text(space, f"TOR {detail['n_tor']}#5", (x + 3, y + height - 40), 2.0)
    add_text(space, f"z={chain['z']:.1f} m", (x + 3, y + 4), 1.8)


def draw_axis(space, x, beam_y, beam_h, label, bubble_level=0):
    """Dibuja un eje; bubble_level escalona burbujas cuando los apoyos están muy juntos."""
    bubble_y = beam_y + beam_h + 12 + bubble_level * 8
    line(space, (x, beam_y - 11), (x, bubble_y + 4), "EJES", RED, 9, "DASHED")
    space.add_circle((x, bubble_y), 3.5,
                     dxfattribs={"layer": "EJES", "color": RED, "lineweight": 18})
    add_text(space, label, (x, bubble_y), 2.5, "EJES", RED,
             TextEntityAlignment.MIDDLE_CENTER)


def stirrup_positions(span_start, span_end, scale, s_end, s_center, zone_mm):
    length_mm = (span_end - span_start) * scale
    zone = min(zone_mm, length_mm / 2)
    values = []
    # Primer estribo a 50 mm desde cada cara; zonas 2h en ambos extremos.
    left = 50.0
    while left < zone - 1e-6:
        values.append(span_start + left / scale)
        left += s_end
    center = zone + s_center
    while center < length_mm - zone - 1e-6:
        values.append(span_start + center / scale)
        center += s_center
    right = length_mm - 50.0
    while right > length_mm - zone + 1e-6:
        values.append(span_start + right / scale)
        right -= s_end
    return sorted(set(round(value, 4) for value in values))


def clear_span(length_mm):
    return max(length_mm - 600.0, 0.5 * length_mm)


def demand_texts(segment):
    """Rótulos compactos y auditables de demanda/estado para un frame SAP."""
    demand = segment["demand"]
    first = (
        f"F{segment['frame']} | M- {demand['mu_neg']:.2f} | "
        f"M+ {demand['mu_pos']:.2f} kN·m"
    )
    second = (
        f"VF {demand['vu_flex']:.2f} | VC {demand['vu_shear']:.2f} kN | "
        f"T {demand['torsion']:.2f} kN·m | {demand['status']}"
    )
    return first, second


def draw_chain(space, chain, detail, row_y, row_height=110.0):
    info_x, info_width = 10.0, 40.0
    anchor_gap = 4.0
    total = sum(item["length"] for item in chain["segments"])
    # La escala reserva explícitamente los dos desarrollos extremos M2: ningún
    # gancho puede entrar al cuadro de datos ni al cajetín lateral.
    available = 530.0 - (info_x + info_width + anchor_gap)
    scale = choose_scale(total + 2 * LDH, available)
    x0 = info_x + info_width + anchor_gap + LDH / scale
    drawn = total / scale
    x1 = x0 + drawn
    vertical_scale = 18.0 if row_height >= 190.0 else 22.0
    beam_h = max(10.0, detail["h"] / vertical_scale)
    beam_y = row_y + row_height * 0.48
    draw_info_box(space, chain, detail, info_x, beam_y - 15.0, info_width, 50.0)

    nodes = [x0]
    for segment in chain["segments"]:
        nodes.append(nodes[-1] + segment["length"] / scale)

    # Contorno y apoyos en cada eje. Las burbujas se escalonan en cadenas
    # densas para que los identificadores sigan siendo legibles.
    rectangle(space, x0, beam_y, drawn, beam_h, "CONTORNO", BLACK, 22)
    minimum_gap = min((nodes[index + 1] - nodes[index] for index in range(len(nodes) - 1)),
                      default=999.0)
    stagger_bubbles = minimum_gap < 18.0
    for index, (x, label) in enumerate(zip(nodes, chain["axes"])):
        support_width = max(4.0, 500.0 / scale)
        rectangle(space, x - support_width / 2, beam_y - 4.0, support_width,
                  beam_h + 8.0, "APOYOS", LIGHT, 15)
        draw_axis(space, x, beam_y, beam_h, label,
                  bubble_level=index % 2 if stagger_bubbles else 0)

    # Estribos reales por cada tramo.
    for idx, segment in enumerate(chain["segments"]):
        start, end = nodes[idx], nodes[idx + 1]
        for x in stirrup_positions(start, end, scale, detail["s_end"], detail["s_center"],
                                   2 * detail["h"]):
            line(space, (x, beam_y + 1.0), (x, beam_y + beam_h - 1.0),
                 "ESTRIBOS", GRAY, 9)
        add_text(space,
                 f"{detail['stirrup'].replace(' cerrado','')} {detail['legs']}R @{detail['s_end']}/@{detail['s_center']}",
                 ((start + end) / 2, beam_y + beam_h / 2 - 0.8), 1.6,
                 "ESTRIBOS", GRAY, TextEntityAlignment.MIDDLE_CENTER)

    # M1: dos barras superiores continuas con ganchos en extremos.
    y_top = beam_y + beam_h - 2.0
    polyline(space, [(x0, y_top - 6.0), (x0, y_top), (x1, y_top), (x1, y_top - 6.0)],
             "ACERO_SUP_CONT", CYAN, 35)
    continuous_cut = clear_span(total) + 2 * LDH
    add_text(space, f"M1  2#5 CONT.  Lc={continuous_cut:.0f} mm", ((x0 + x1) / 2, y_top + 3.0),
             2.2, "ACERO_SUP_CONT", CYAN, TextEntityAlignment.MIDDLE_CENTER)

    # M2: barras negativas sobre cada apoyo, alternadas verticalmente para evitar superposición.
    extra = max(detail["n_sup"] - 2, 0)
    if extra:
        for index, x in enumerate(nodes):
            if index == 0:
                right = nodes[1]
                extent_r = (right - x) / 4
                start, end = x - LDH / scale, x + extent_r
                cut = clear_span(chain["segments"][0]["length"]) / 4 + LDH
            elif index == len(nodes) - 1:
                left = nodes[-2]
                extent_l = (x - left) / 4
                start, end = x - extent_l, x + LDH / scale
                cut = clear_span(chain["segments"][-1]["length"]) / 4 + LDH
            else:
                extent_l = (x - nodes[index - 1]) / 4
                extent_r = (nodes[index + 1] - x) / 4
                start, end = x - extent_l, x + extent_r
                cut = (chain["segments"][index - 1]["length"] / 4
                       + chain["segments"][index]["length"] / 4 + 500.0)
            y = beam_y + beam_h - 4.5 - (index % 2) * 2.2
            points = [(start, y)]
            if index == 0:
                points = [(start, y - 5.0), (start, y)]
            points.append((end, y))
            if index == len(nodes) - 1:
                points.append((end, y - 5.0))
            polyline(space, points, "ACERO_SUP_APOYO", BLUE, 30)
            add_text(space, f"M2 {extra}#5 L={cut:.0f}", (x, y + 1.5), 1.7,
                     "ACERO_SUP_APOYO", BLUE, TextEntityAlignment.MIDDLE_CENTER)

    # M3: barras inferiores independientes por tramo, con ganchos en ambas caras.
    for idx, segment in enumerate(chain["segments"]):
        start, end = nodes[idx], nodes[idx + 1]
        y = beam_y + 2.0 + (idx % 2) * 1.8
        support_shift = min(2.0, 150.0 / scale)
        polyline(space, [(start + support_shift, y + 5.0), (start + support_shift, y),
                         (end - support_shift, y), (end - support_shift, y + 5.0)],
                 "ACERO_INF", RED, 35)
        cut = clear_span(segment["length"]) + 2 * LDH
        add_text(space, f"M3 {detail['n_inf']}#5 L={cut:.0f}", ((start + end) / 2, y - 2.7),
                 1.7, "ACERO_INF", RED, TextEntityAlignment.MIDDLE_CENTER)

    # M4: acero torsional dedicado continuo, separado de flexión.
    if detail["n_tor"]:
        y_tor = beam_y + beam_h / 2 + 2.4
        polyline(space, [(x0 + 1.0, y_tor), (x1 - 1.0, y_tor)],
                 "ACERO_TORSION", BROWN, 26)
        add_text(space, f"M4 {detail['n_tor']}#5 TORSIÓN DEDICADA", ((x0 + x1) / 2, y_tor + 1.6),
                 1.7, "ACERO_TORSION", BROWN, TextEntityAlignment.MIDDLE_CENTER)

    # Cortes en el interior de los tramos, nunca sobre el centro de un apoyo.
    # B se desplaza del centro geométrico para no atravesar los rótulos M1/M3/M4.
    segment_count = len(chain["segments"])
    if segment_count == 1:
        cut_locations = [
            (x0 + drawn * 0.18, "A"),
            (x0 + drawn * 0.68, "B"),
            (x0 + drawn * 0.84, "C"),
        ]
    else:
        middle = segment_count // 2
        cut_locations = [
            (nodes[0] + (nodes[1] - nodes[0]) * 0.18, "A"),
            (nodes[middle] + (nodes[middle + 1] - nodes[middle]) * 0.68, "B"),
            (nodes[-2] + (nodes[-1] - nodes[-2]) * 0.82, "C"),
        ]
    for x, label in cut_locations:
        line(space, (x, beam_y - 2.5), (x, beam_y + beam_h + 2.5), "CORTE", MAGENTA, 13)
        add_text(space, label, (x, beam_y + beam_h + 7.5), 2.0, "CORTE", MAGENTA,
                 TextEntityAlignment.MIDDLE_CENTER)

    # En cadenas con tramos cortos, cotas y números de frame alternan en dos
    # niveles. La cota total queda en un tercer nivel independiente.
    dim_y = beam_y - 10.0
    dense_dimensions = minimum_gap < 20.0
    for idx, segment in enumerate(chain["segments"]):
        local_dim_y = dim_y - (idx % 2) * 6.0 if dense_dimensions else dim_y
        dimension(space, nodes[idx], nodes[idx + 1], local_dim_y,
                  f"{segment['length'] / 1000:.3f} m", 1.8)
        first, second = demand_texts(segment)
        midpoint = (nodes[idx] + nodes[idx + 1]) / 2
        status_color = RED if segment["demand"]["status"] == "NO CUMPLE" else BLACK
        add_text(space, first, (midpoint, local_dim_y - 3.0),
                 1.05, "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
        add_text(space, second, (midpoint, local_dim_y - 5.2),
                 1.05, "TEXTO", status_color, TextEntityAlignment.MIDDLE_CENTER)
    total_dim_y = dim_y - (18.0 if dense_dimensions else 11.0)
    dimension(space, x0, x1, total_dim_y, f"TOTAL SAP I-J = {total / 1000:.3f} m", 2.0)
    add_text(space, f"ESC. H 1:{scale} | V 1:22 | EJES Y TRAMOS SAP",
             (x1, row_y + 4.0), 1.9, "TEXTO", BLACK, TextEntityAlignment.RIGHT)
    return {"scale": scale, "total": total, "nodes": nodes, "beam_y": beam_y, "beam_h": beam_h}


def flex_positions(count, width, y, margin=65.0):
    if count <= 0:
        return []
    if count == 1:
        return [(width / 2, y)]
    usable = width - 2 * margin
    return [(margin + index * usable / (count - 1), y) for index in range(count)]


def torsion_positions(count, width, height):
    if count <= 0:
        return []
    # Franja independiente de las filas superior/inferior para evitar colisiones.
    per_side = max(1, count // 4)
    left, right = 90.0, width - 90.0
    bottom, top = 110.0, height - 110.0
    points = []
    for index in range(per_side):
        fraction = (index + 1) / (per_side + 1)
        points.append((left + fraction * (right - left), bottom))
        points.append((left + fraction * (right - left), top))
    remaining = count - len(points)
    per_vertical = max(0, remaining // 2)
    for index in range(per_vertical):
        fraction = (index + 1) / (per_vertical + 1)
        points.append((left, bottom + fraction * (top - bottom)))
        points.append((right, bottom + fraction * (top - bottom)))
    return points[:count]


def check_section_clearance(detail):
    b, h = detail["b"], detail["h"]
    y_bottom, y_top = 60.65, h - 60.65
    points = [("SUP", point) for point in flex_positions(detail["n_sup"], b, y_top)]
    points += [("INF", point) for point in flex_positions(detail["n_inf"], b, y_bottom)]
    points += [("TOR", point) for point in torsion_positions(detail["n_tor"], b, h)]
    minimum = DB_LONG + 25.0
    for idx, (kind_a, a) in enumerate(points):
        for kind_b, bpoint in points[idx + 1:]:
            if math.dist(a, bpoint) < minimum - 1e-6:
                raise ValueError(f"Colisión {detail['group']} {kind_a}/{kind_b}: {a} {bpoint}")


def draw_section(space, cx, cy, detail, label, center=False):
    scale = 20.0
    width, height = detail["b"] / scale, detail["h"] / scale
    left, bottom = cx - width / 2, cy - height / 2
    rectangle(space, left, bottom, width, height, "CONTORNO", BLACK, 20)
    inset = COVER / scale
    rectangle(space, left + inset, bottom + inset, width - 2 * inset, height - 2 * inset,
              "ESTRIBOS", GRAY, 18)
    if detail["legs"] == 4:
        line(space, (left + width / 3, bottom + inset), (left + width / 3, bottom + height - inset),
             "ESTRIBOS", GRAY, 13)
        line(space, (left + 2 * width / 3, bottom + inset), (left + 2 * width / 3, bottom + height - inset),
             "ESTRIBOS", GRAY, 13)

    n_top = 2 if center else detail["n_sup"]
    for x, y in flex_positions(n_top, detail["b"], detail["h"] - 60.65):
        space.add_circle((left + x / scale, bottom + y / scale), 0.55,
                         dxfattribs={"layer": "ACERO_SUP_CONT", "color": CYAN})
    for x, y in flex_positions(detail["n_inf"], detail["b"], 60.65):
        space.add_circle((left + x / scale, bottom + y / scale), 0.55,
                         dxfattribs={"layer": "ACERO_INF", "color": RED})
    for x, y in torsion_positions(detail["n_tor"], detail["b"], detail["h"]):
        space.add_circle((left + x / scale, bottom + y / scale), 0.48,
                         dxfattribs={"layer": "ACERO_TORSION", "color": BROWN})
    add_text(space, label, (cx, bottom - 3.6), 1.9, "TEXTO", BLACK,
             TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"{int(detail['b'])}x{int(detail['h'])} | 1:20", (cx, bottom - 6.4), 1.55,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)


def draw_side_details(space, detail, sheet_id, chains):
    """Cajetín lateral y detalles, como en la lámina guía del usuario."""
    x0, x1 = 538.0, 697.0
    rectangle(space, x0, 7.0, x1 - x0, 486.0, "MARCO", BLACK, 30)
    for y in (458.0, 405.0, 325.0, 230.0, 150.0, 75.0):
        line(space, (x0, y), (x1, y), "MARCO", BLACK, 20)

    # Encabezado del proyecto.
    add_text(space, "UNIVERSIDAD NACIONAL DE COLOMBIA", ((x0 + x1) / 2, 484), 2.5,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, "DISEÑO ESTRUCTURAL — GRUPO 6", ((x0 + x1) / 2, 476), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, "EDIFICIO RESIDENCIAL — SANTA MARTA", ((x0 + x1) / 2, 468), 2.0,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)

    # Identificación de la lámina.
    add_text(space, "DESPIECE DE VIGA CONTINUA", ((x0 + x1) / 2, 448), 3.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, detail["group"], ((x0 + x1) / 2, 435), 6.0,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"SECCIÓN {int(detail['b'])} x {int(detail['h'])} mm", ((x0 + x1) / 2, 423), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"{detail['count']} FRAMES SAP — NIVEL TIPO", ((x0 + x1) / 2, 414), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)

    # Cortes agrupados y alineados con el detalle longitudinal.
    add_text(space, "CORTES TRANSVERSALES", ((x0 + x1) / 2, 396), 2.6,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    draw_section(space, 564, 363, detail, "A-A APOYO", False)
    draw_section(space, 617.5, 363, detail, "B-B CENTRO", True)
    draw_section(space, 671, 363, detail, "C-C APOYO", False)

    # Gancho de anclaje a 90 grados.
    add_text(space, "DETALLE DE ANCLAJE #5 — ESC. 1:10", ((x0 + x1) / 2, 315), 2.5,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    hx, hy = 557.0, 285.0
    polyline(space, [(hx, hy - TAIL_90 / 10), (hx, hy), (hx + LDH / 10, hy)],
             "ACERO_SUP_APOYO", BLUE, 50)
    dimension(space, hx, hx + LDH / 10, hy + 7.0, f"ldh = {LDH:.0f} mm", 1.8)
    add_text(space, f"COLA 90° = {TAIL_90:.0f} mm", (596, 287), 2.0,
             "ACERO_SUP_APOYO", BLUE)
    add_text(space, f"Dint >= {BEND_90:.0f} mm", (596, 278), 2.0)
    add_text(space, f"ld SUP = {LD_SUP:.0f} mm", (596, 269), 2.0)
    add_text(space, f"ld INF = {LD_INF:.0f} mm", (596, 260), 2.0)
    add_text(space, "USAR GANCHO SI NO CABE EL ld RECTO", ((x0 + x1) / 2, 240), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)

    # Cuadro de marcas; evita repetir notas pequeñas sobre el dibujo.
    add_text(space, "CUADRO DE ARMADURAS", ((x0 + x1) / 2, 220), 2.6,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    rows = [
        ("M1", "2#5", "SUPERIOR CONTINUA"),
        ("M2", f"{max(detail['n_sup'] - 2, 0)}#5", "NEGATIVA DE APOYO"),
        ("M3", f"{detail['n_inf']}#5", "INFERIOR POR TRAMO"),
        ("M4", f"{detail['n_tor']}#5", "TORSIÓN DEDICADA"),
        ("E1", detail["stirrup"].replace(" cerrado", ""), f"{detail['legs']}R @{detail['s_end']}/@{detail['s_center']}"),
    ]
    for index, (mark, steel, use) in enumerate(rows):
        y = 208 - index * 10.5
        add_text(space, mark, (545, y), 2.2)
        add_text(space, steel, (565, y), 2.2)
        add_text(space, use, (595, y), 2.2)

    # Localización resumida en el mismo orden descendente usado en el plano SAP.
    add_text(space, "LOCALIZACIÓN / RECORRIDO SAP", ((x0 + x1) / 2, 140), 2.5,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    for index, chain in enumerate(chains[:4]):
        axes = "-".join(chain["axes"])
        add_text(space, f"{index + 1}. EJE {chain['line_axis']} : {axes}", (545, 128 - index * 12), 2.2)

    # Cajetín final lateral.
    add_text(space, "FORMATO ISO B2 — 707 x 500 mm", ((x0 + x1) / 2, 65), 2.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, "COTAS EN mm — ESCALAS INDICADAS", ((x0 + x1) / 2, 55), 2.0,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"LÁMINA {sheet_id}", ((x0 + x1) / 2, 42), 3.2,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, DATE, ((x0 + x1) / 2, 31), 2.0,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, "DETALLE ACADÉMICO", ((x0 + x1) / 2, 20), 2.2,
             "TEXTO", RED, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, "NO EMITIR PARA CONSTRUCCIÓN", ((x0 + x1) / 2, 12), 2.2,
             "TEXTO", RED, TextEntityAlignment.MIDDLE_CENTER)


def configure_layout(layout):
    layout.page_setup(size=(SHEET_W, SHEET_H), margins=(0, 0, 0, 0), units="mm",
                      rotation=0, scale=(1, 1), name="ISO B2 707x500 mm 1:1",
                      device="DWG To PDF.pc3")
    layout.dxf.plot_layout_flags = 0


def representative_chain(chains):
    """Selecciona una sola viga tipo: la continua más larga y con más tramos."""
    return max(
        chains,
        key=lambda chain: (
            sum(segment["length"] for segment in chain["segments"]),
            len(chain["segments"]),
        ),
    )


def draw_group_side_panel(space, group, chain, detail, row_y, row_height):
    """Resumen de sección y armado leído del Excel para una viga tipo."""
    x0, x1 = 538.0, 697.0
    rectangle(space, x0, row_y, x1 - x0, row_height, "MARCO", BLACK, 22)
    top = row_y + row_height
    add_text(space, f"VIGA TIPO {group}", ((x0 + x1) / 2, top - 7.0), 3.4,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(space, f"{int(detail['b'])}x{int(detail['h'])} mm — DATOS EXCEL",
             ((x0 + x1) / 2, top - 14.0), 2.0,
             "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)

    section_x = 562.0
    section_y = row_y + row_height * 0.48
    draw_section(space, section_x, section_y, detail, "SECCIÓN APOYO", False)

    text_x = 585.0
    first_y = top - 27.0
    lines = [
        f"SUP: {detail['n_sup']}#5",
        f"INF: {detail['n_inf']}#5",
        f"TORSIÓN: {detail['n_tor']}#5",
        f"EST: {detail['stirrup'].replace(' cerrado', '')} — {detail['legs']} RAMAS",
        f"s EXT/CENTRO: {detail['s_end']}/{detail['s_center']} mm",
        f"FRAMES SAP: {detail['count']}",
        f"ESTADO GRUPO: {detail['status']}",
        f"FRAMES NO CUMPLE: {detail['noncompliant']}",
        f"EJE {chain['line_axis']}: {'-'.join(chain['axes'])}",
    ]
    for index, text in enumerate(lines):
        color = RED if "NO CUMPLE" in text and not text.endswith(": 0") else BLACK
        add_text(space, text, (text_x, first_y - index * 7.0), 1.8, color=color)


def draw_grouped_layout(layout, groups, chains_by_group, details, sheet_id, family_label):
    """Agrupa varias vigas tipo en una lámina, una representación por grupo."""
    configure_layout(layout)
    rectangle(layout, MARGIN, MARGIN, SHEET_W - 2 * MARGIN, SHEET_H - 2 * MARGIN,
              "MARCO", BLACK, 30)
    add_text(layout, f"DESPIECE DE VIGAS DE {family_label} — {sheet_id}", (353.5, 488.0),
             5.0, "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    add_text(layout, "UNA VIGA TIPO POR GRUPO | DEMANDAS/ESTADOS POR FRAME DESDE EXCEL RECONCILIADO",
             (353.5, 478.0), 2.5, "TEXTO", BLACK, TextEntityAlignment.MIDDLE_CENTER)
    line(layout, (10, 470), (697, 470), "MARCO", BLACK, 25)

    top, bottom = 470.0, 15.0
    row_height = (top - bottom) / len(groups)
    if row_height < 105.0:
        raise ValueError(f"No caben los grupos {groups} en una lámina B2 legible")
    for index, group in enumerate(groups):
        row_y = top - (index + 1) * row_height
        chain = representative_chain(chains_by_group[group])
        draw_chain(layout, chain, details[group], row_y, row_height)
        draw_group_side_panel(layout, group, chain, details[group], row_y, row_height)
        if index < len(groups) - 1:
            line(layout, (10, row_y), (697, row_y), "MARCO", GRAY, 12)

    add_text(layout, "DETALLE ACADÉMICO — NO EMITIR PARA CONSTRUCCIÓN",
             (353.5, 9.5), 1.8, "TEXTO", RED, TextEntityAlignment.MIDDLE_CENTER)


def populate_model_overview(doc, layout_names):
    """Copia todas las láminas al Model para que nunca se abra un espacio vacío."""
    model = doc.modelspace()
    gap = 25.0
    for index, layout_name in enumerate(layout_names):
        column, row = index % 2, index // 2
        offset_x = column * (SHEET_W + gap)
        offset_y = -row * (SHEET_H + gap)
        transform = Matrix44.translate(offset_x, offset_y, 0.0)
        for entity in doc.layouts.get(layout_name):
            if entity.dxftype() == "VIEWPORT":
                continue
            duplicate = entity.copy()
            try:
                duplicate.transform(transform)
            except NotImplementedError:
                continue
            model.add_entity(duplicate)


def create_family(path: Path, sheet_groups, chains_by_group, details, prefix, family_label):
    doc = ezdxf.new("R2018", setup=True)
    add_layers(doc)
    layout_names = []
    for index, groups in enumerate(sheet_groups, 1):
        layout_name = f"{prefix}_{index:02d}"
        layout_names.append(layout_name)
        if index == 1:
            doc.layouts.rename("Layout1", layout_name)
            layout = doc.layouts.get(layout_name)
        else:
            layout = doc.layouts.new(layout_name)
        draw_grouped_layout(
            layout, groups, chains_by_group, details,
            f"{prefix}-{index:02d}", family_label,
        )

    # El Model contiene también las dos láminas para visores que ignoran layouts.
    doc.layouts.set_active_layout(layout_names[0])
    doc.header["$TILEMODE"] = 0
    populate_model_overview(doc, layout_names)
    doc.units = 4
    doc.header["$INSUNITS"] = 4
    doc.header["$MEASUREMENT"] = 1
    doc.saveas(path)
    audit = doc.audit()
    if audit.has_errors:
        raise ValueError(f"Auditoría DXF con errores en {path.name}: {len(audit.errors)}")


def audit_family_against_design(path, sheet_groups, details, chains_by_group):
    """Comprueba que cada sección dibujada reproduce las cantidades del Excel."""
    doc = ezdxf.readfile(path)
    layouts = [layout for layout in doc.layouts if layout.name != "Model"]
    if len(layouts) != len(sheet_groups):
        raise ValueError(f"Cantidad de láminas inesperada en {path.name}")

    seen = []
    for layout, groups in zip(layouts, sheet_groups):
        texts = [entity.dxf.text for entity in layout.query("TEXT")]
        top, bottom = 470.0, 15.0
        row_height = (top - bottom) / len(groups)
        for index, group in enumerate(groups):
            detail = details[group]
            seen.append(group)
            required = [
                f"VIGA TIPO {group}",
                f"SUP: {detail['n_sup']}#5",
                f"INF: {detail['n_inf']}#5",
                f"TORSIÓN: {detail['n_tor']}#5",
                f"EST: {detail['stirrup'].replace(' cerrado', '')} — {detail['legs']} RAMAS",
                f"s EXT/CENTRO: {detail['s_end']}/{detail['s_center']} mm",
                f"ESTADO GRUPO: {detail['status']}",
                f"FRAMES NO CUMPLE: {detail['noncompliant']}",
            ]
            chain = representative_chain(chains_by_group[group])
            for segment in chain["segments"]:
                required.extend(demand_texts(segment))
            missing = [text for text in required if text not in texts]
            if missing:
                raise ValueError(f"{group}: datos Excel ausentes del DXF: {missing}")

            row_y = top - (index + 1) * row_height
            row_top = row_y + row_height
            section_circles = [
                entity for entity in layout.query("CIRCLE")
                if entity.dxf.center.x >= 538.0
                and row_y <= entity.dxf.center.y <= row_top
            ]
            counts = Counter(entity.dxf.layer for entity in section_circles)
            expected = {
                "ACERO_SUP_CONT": detail["n_sup"],
                "ACERO_INF": detail["n_inf"],
                "ACERO_TORSION": detail["n_tor"],
            }
            for layer, quantity in expected.items():
                if counts[layer] != quantity:
                    raise ValueError(
                        f"{group}: {layer} dibujado={counts[layer]}, Excel={quantity}"
                    )
    expected_groups = [group for sheet in sheet_groups for group in sheet]
    if seen != expected_groups or len(seen) != len(set(seen)):
        raise ValueError(f"Grupos repetidos o faltantes en {path.name}: {seen}")


def render_previews(dxf_paths):
    """Crea PDF y PNG visibles sin necesidad de AutoCAD."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import font_manager as mpl_font_manager
    from matplotlib.backends.backend_pdf import PdfPages
    from PIL import Image, ImageDraw
    from ezdxf.addons.drawing import Frontend, RenderContext
    from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
    from ezdxf.fonts import fonts
    from ezdxf.fonts.font_manager import get_ttf_font_face

    # ezdxf no descubre siempre las fuentes incluidas con Matplotlib. Se
    # registra DejaVu Sans explícitamente para que los rótulos sean visibles.
    font_path = Path(mpl_font_manager.findfont("DejaVu Sans"))
    font_face = get_ttf_font_face(font_path)
    fonts.font_manager._font_cache.add_entry(font_path, font_face)
    fonts.font_manager._fallback_font_name = font_path.name

    preview_dir = OUT / "Vistas-Previas-PNG"
    preview_dir.mkdir(parents=True, exist_ok=True)
    # Evita mezclar las antiguas doce láminas individuales con las cuatro
    # láminas agrupadas de esta entrega.
    for old_preview in preview_dir.glob("*.png"):
        old_preview.unlink()
    rendered = []
    for dxf_path in dxf_paths:
        doc = ezdxf.readfile(dxf_path)
        family = "V-01-CARGA" if dxf_path.name.startswith("V-01") else "V-02-RIGIDEZ"
        pdf_path = OUT / f"{family}-VISTA-PREVIA.pdf"
        with PdfPages(pdf_path) as pdf:
            for layout in doc.layouts:
                if layout.name == "Model":
                    continue
                fig = plt.figure(figsize=(14.14, 10.0), dpi=120, facecolor="white")
                axis = fig.add_axes([0, 0, 1, 1])
                axis.set_facecolor("white")
                Frontend(RenderContext(doc), MatplotlibBackend(axis)).draw_layout(layout, finalize=True)
                axis.set_xlim(0, SHEET_W)
                axis.set_ylim(0, SHEET_H)
                axis.set_aspect("equal", adjustable="box")
                axis.axis("off")
                png_path = preview_dir / f"{family}-{layout.name}.png"
                fig.savefig(png_path, dpi=120, facecolor="white", bbox_inches=None, pad_inches=0)
                pdf.savefig(fig, facecolor="white", bbox_inches=None, pad_inches=0)
                plt.close(fig)
                rendered.append((layout.name, png_path))

    thumb_w, thumb_h, label_h = 707, 500, 28
    columns = 2
    rows = math.ceil(len(rendered) / columns)
    contact = Image.new("RGB", (columns * thumb_w, rows * (thumb_h + label_h)), "white")
    draw = ImageDraw.Draw(contact)
    for index, (layout_name, image_path) in enumerate(rendered):
        image = Image.open(image_path).convert("RGB")
        image.thumbnail((thumb_w, thumb_h))
        column, row = index % columns, index // columns
        x = column * thumb_w + (thumb_w - image.width) // 2
        y = row * (thumb_h + label_h)
        contact.paste(image, (x, y))
        draw.text((column * thumb_w + 8, y + thumb_h + 5), layout_name, fill="black")
    contact.save(OUT / "VISTA-PREVIA-TODAS-LAS-LAMINAS.png")


def audit_inputs(chains_by_group, details):
    if set(chains_by_group) != set(GROUPS):
        raise ValueError("No se reconstruyeron los doce grupos")
    for group in GROUPS:
        if not chains_by_group[group]:
            raise ValueError(f"El grupo {group} no tiene cadenas")
        check_section_clearance(details[group])
        for chain in chains_by_group[group]:
            if len(chain["axes"]) != len(chain["segments"]) + 1:
                raise ValueError(f"Ejes/tramos inconsistentes en {group}")
            if len({item["frame"] for item in chain["segments"]}) != len(chain["segments"]):
                raise ValueError(f"Frames duplicados en la cadena {group}")
            if any(item["length"] <= 0 for item in chain["segments"]):
                raise ValueError(f"Longitud no positiva en {group}")
    if (LD_INF, LD_SUP, LDH, TAIL_90, BEND_90) != (370.0, 480.0, 305.0, 195.0, 100.0):
        raise ValueError("Las longitudes de desarrollo/ganchos no son las aprobadas")


def main():
    grids, joints, frames, assignments, details, demands = load_sources()
    chains_by_group = {
        group: frame_components(group, assignments, frames, joints, grids, demands)
        for group in GROUPS
    }
    audit_inputs(chains_by_group, details)
    OUT.mkdir(parents=True, exist_ok=True)
    vc_path = OUT / "V-01-VIGAS-DE-CARGA-CONTINUAS-B2.dxf"
    vr_path = OUT / "V-02-VIGAS-DE-RIGIDEZ-CONTINUAS-B2.dxf"
    create_family(vc_path, VC_SHEETS, chains_by_group, details, "CARGA", "CARGA")
    create_family(vr_path, VR_SHEETS, chains_by_group, details, "RIGIDEZ", "RIGIDEZ")
    audit_family_against_design(vc_path, VC_SHEETS, details, chains_by_group)
    audit_family_against_design(vr_path, VR_SHEETS, details, chains_by_group)
    render_previews([vc_path, vr_path])
    preview_files = [
        OUT / "V-01-CARGA-VISTA-PREVIA.pdf",
        OUT / "V-02-RIGIDEZ-VISTA-PREVIA.pdf",
        OUT / "VISTA-PREVIA-TODAS-LAS-LAMINAS.png",
    ]
    with zipfile.ZipFile(ZIP, "w", zipfile.ZIP_DEFLATED) as archive:
        for deliverable in [vc_path, vr_path, *preview_files]:
            archive.write(deliverable, deliverable.name)
    print(f"OK {vc_path}")
    print(f"OK {vr_path}")
    for preview in preview_files:
        print(f"OK {preview}")
    print(f"OK {ZIP}")
    for group in GROUPS:
        chain = representative_chain(chains_by_group[group])
        signature = "-".join(chain["axes"])
        print(
            f"{group}: 1 viga tipo dibujada | recorrido {signature} | "
            f"SUP {details[group]['n_sup']}#5, INF {details[group]['n_inf']}#5, "
            f"TOR {details[group]['n_tor']}#5 | ESTADO {details[group]['status']} | "
            f"{details[group]['noncompliant']} frames NO CUMPLE"
        )


if __name__ == "__main__":
    main()
