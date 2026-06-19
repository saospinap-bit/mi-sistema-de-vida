#!/usr/bin/env python3
"""
Herramienta de diseno a FLEXION, CORTANTE y TORSION (NSR-10 / ACI 318) en SI.
Genera: Herramienta-Diseno-Viguetas-Vigas.xlsx
Hojas: Viguetas | Vigas | Riostras
Unidades: f'c, fy en MPa ; dimensiones en mm ; Mu/Tu en kN*m ; Vu en kN ; As en mm2.
Metes los datos de entrada (azul) y la hoja calcula sola con formulas.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="1A237E"; GRIS="37474F"; CLARO="ECEFF1"; AZUL_IN="BBDEFB"; VERDE="C8E6C9"; AMAR="FFF9C4"
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def titulo(ws,a,t,span,color=NAVY):
    ws.merge_cells(f"{a}:{span}"); c=ws[a]; c.value=t
    c.font=Font(bold=True,size=13,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER

def seccion(ws,row,txt,color=GRIS):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    c=ws.cell(row=row,column=1,value=txt); c.font=Font(bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=LEFT

def inp(ws,row,label,val,unit="",nota="",fmt=None):
    ws.cell(row=row,column=1,value=label).alignment=LEFT
    c=ws.cell(row=row,column=2,value=val); c.alignment=CENTER; c.fill=PatternFill("solid",fgColor=AZUL_IN)
    if fmt: c.number_format=fmt
    ws.cell(row=row,column=3,value=unit).alignment=CENTER
    ws.cell(row=row,column=4,value=nota).alignment=LEFT
    for j in range(1,5): ws.cell(row=row,column=j).border=BORDER

def out(ws,row,label,formula,unit="",fmt="0.00",bold=False,hl=None):
    ws.cell(row=row,column=1,value=label).alignment=LEFT
    c=ws.cell(row=row,column=2,value=formula); c.alignment=CENTER; c.number_format=fmt
    if bold: c.font=Font(bold=True)
    if hl: c.fill=PatternFill("solid",fgColor=hl)
    ws.cell(row=row,column=3,value=unit).alignment=CENTER
    for j in range(1,5): ws.cell(row=row,column=j).border=BORDER

def anchos(ws):
    ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=16
    ws.column_dimensions["C"].width=10; ws.column_dimensions["D"].width=46

# ============================================================ VIGUETAS
def hoja_viguetas(wb):
    ws=wb.active; ws.title="Viguetas"; anchos(ws)
    titulo(ws,"A1","DISENO DE VIGUETA (seccion T) - NSR-10 / ACI 318","D1")
    seccion(ws,3,"DATOS DE ENTRADA (celdas azules)")
    inp(ws,4,"f'c (resist. concreto)",21,"MPa")
    inp(ws,5,"fy (acero longitudinal)",420,"MPa")
    inp(ws,6,"fyt (acero transversal)",420,"MPa")
    inp(ws,7,"bw (ancho del nervio)",100,"mm")
    inp(ws,8,"h (altura total)",400,"mm")
    inp(ws,9,"d (altura efectiva)",360,"mm","d = h - recubrimiento - estribo - db/2")
    inp(ws,10,"hf (espesor loseta)",50,"mm")
    inp(ws,11,"bf (ancho efectivo del ala)",850,"mm","min(L/4, bw+16hf, sep. entre nervios)")
    inp(ws,12,"L (luz libre)",5000,"mm")
    inp(ws,13,"Factor de apoyo",18.5,"-","16=simple, 18.5=1 extremo cont., 21=ambos cont., 8=voladizo")
    inp(ws,14,"Mu+ (momento positivo)",25,"kN*m")
    inp(ws,15,"Mu- (momento negativo)",30,"kN*m")
    inp(ws,16,"Vu (cortante ultimo)",35,"kN")
    inp(ws,17,"Tu (torsion ultima)",0.3,"kN*m")
    inp(ws,18,"Av (area estribo, 2 ramas)",64,"mm2","#2 2 ramas ~= 64 mm2")
    inp(ws,19,"phi flexion",0.9,"-")
    inp(ws,20,"phi cortante/torsion",0.75,"-")

    seccion(ws,23,"1) FLEXION POSITIVA (+)  -> acero inferior, comportamiento de viga T")
    out(ws,24,"Rn = Mu+/(phi*bf*d^2)","=B14*10^6/(B19*B11*B9^2)","MPa","0.000")
    out(ws,25,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B24/(0.85*B4)))","-","0.00000")
    out(ws,26,"As+ requerido","=B25*B11*B9","mm2","0")
    out(ws,27,"a (profundidad bloque)","=B26*B5/(0.85*B4*B11)","mm","0.0")
    out(ws,28,"Comportamiento","=IF(B27<=B10,\"Rectangular OK (a<=hf)\",\"Es viga T real: revisar\")","","General")
    out(ws,29,"As minimo","=MAX(1.4/B5,0.25*SQRT(B4)/B5)*B7*B9","mm2","0")
    out(ws,30,"As+ ADOPTADO","=MAX(B26,B29)","mm2","0",bold=True,hl=VERDE)

    seccion(ws,32,"2) FLEXION NEGATIVA (-)  -> acero superior, seccion rectangular (bw)")
    out(ws,33,"Rn = Mu-/(phi*bw*d^2)","=B15*10^6/(B19*B7*B9^2)","MPa","0.000")
    out(ws,34,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B33/(0.85*B4)))","-","0.00000")
    out(ws,35,"As- requerido","=B34*B7*B9","mm2","0")
    out(ws,36,"beta1","=IF(B4<=28,0.85,MAX(0.65,0.85-0.05*(B4-28)/7))","-","0.000")
    out(ws,37,"rho maximo (ductil, et=0.005)","=0.85*B36*B4/B5*(3/8)","-","0.00000")
    out(ws,38,"Chequeo ductilidad","=IF(B34<=B37,\"OK ductil\",\"Aumentar h o f'c\")","","General")
    out(ws,39,"As- ADOPTADO","=MAX(B35,B29)","mm2","0",bold=True,hl=VERDE)

    seccion(ws,41,"3) CORTANTE  (las viguetas admiten Vc x 1.1, NSR-10 C.8.13)")
    out(ws,42,"Vc = 0.17*raiz(f'c)*bw*d","=0.17*SQRT(B4)*B7*B9/1000","kN","0.00")
    out(ws,43,"Vc vigueta = 1.1*Vc","=1.1*B42","kN","0.00")
    out(ws,44,"phi*Vc","=B20*B43","kN","0.00")
    out(ws,45,"Vs requerido = Vu/phi - Vc","=MAX(B16/B20-B43,0)","kN","0.00")
    out(ws,46,"s calculado = Av*fyt*d/Vs","=IF(B45<=0,9999,B18*B6*B9/(B45*1000))","mm","0")
    out(ws,47,"s maximo","=MIN(B9/2,600)","mm","0")
    out(ws,48,"RESULTADO","=IF(B16<=B44,\"Concreto resiste: estribos de montaje\",\"Estribos #2 @ \"&ROUND(MIN(B46,B47),0)&\" mm\")","","General",bold=True,hl=AMAR)

    seccion(ws,50,"4) TORSION  (umbral: si Tu < phi*Tth se desprecia)")
    out(ws,51,"Acp = bw*h","=B7*B8","mm2","0")
    out(ws,52,"pcp = 2*(bw+h)","=2*(B7+B8)","mm","0")
    out(ws,53,"phi*Tth = phi*0.083*raiz(f'c)*Acp^2/pcp","=B20*0.083*SQRT(B4)*(B51^2/B52)/10^6","kN*m","0.000")
    out(ws,54,"RESULTADO","=IF(B17<B53,\"Torsion despreciable\",\"Disenar torsion (ver hoja Vigas)\")","","General",bold=True,hl=AMAR)

    seccion(ws,56,"5) DEFLEXIONES (NSR-10 C.9.5)")
    out(ws,57,"h minimo = L/factor","=B12/B13","mm","0")
    out(ws,58,"RESULTADO","=IF(B8>=B57,\"h>=hmin: no obligatorio calcular deflexiones\",\"Calcular deflexion a largo plazo\")","","General",bold=True,hl=AMAR)
    out(ws,59,"Multiplicador largo plazo (xi=2, rho'=0)","=2.0/(1+50*0)","-","0.0")
    ws.cell(row=61,column=1,value="Nota: As+ usa el ala (bf) a compresion; As- usa solo bw. Verifica que a<=hf para usar la formula rectangular.").font=Font(italic=True,size=9)

# ============================================================ VIGAS
def hoja_vigas(wb):
    ws=wb.create_sheet("Vigas"); anchos(ws)
    titulo(ws,"A1","DISENO DE VIGA rectangular - flexion, cortante y torsion","D1")
    seccion(ws,3,"DATOS DE ENTRADA (celdas azules)")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa"); inp(ws,6,"fyt",420,"MPa")
    inp(ws,7,"b (ancho)",300,"mm"); inp(ws,8,"h (altura)",500,"mm"); inp(ws,9,"d (efectiva)",450,"mm")
    inp(ws,10,"recubrimiento libre cc",40,"mm")
    inp(ws,11,"Mu+ ",120,"kN*m"); inp(ws,12,"Mu- ",150,"kN*m")
    inp(ws,13,"Vu ",140,"kN"); inp(ws,14,"Tu ",12,"kN*m")
    inp(ws,15,"Av (estribo, 2 ramas)",142,"mm2","#3 2 ramas ~= 142 mm2")
    inp(ws,16,"phi flexion",0.9,"-"); inp(ws,17,"phi cortante/torsion",0.75,"-")

    seccion(ws,20,"1) FLEXION (rectangular)")
    out(ws,21,"Rn+ ","=B11*10^6/(B16*B7*B9^2)","MPa","0.000")
    out(ws,22,"rho+","=(0.85*B4/B5)*(1-SQRT(1-2*B21/(0.85*B4)))","-","0.00000")
    out(ws,23,"As+ ","=B22*B7*B9","mm2","0")
    out(ws,24,"Rn- ","=B12*10^6/(B16*B7*B9^2)","MPa","0.000")
    out(ws,25,"rho-","=(0.85*B4/B5)*(1-SQRT(1-2*B24/(0.85*B4)))","-","0.00000")
    out(ws,26,"As- ","=B25*B7*B9","mm2","0")
    out(ws,27,"As minimo","=MAX(1.4/B5,0.25*SQRT(B4)/B5)*B7*B9","mm2","0")
    out(ws,28,"As+ ADOPTADO","=MAX(B23,B27)","mm2","0",bold=True,hl=VERDE)
    out(ws,29,"As- ADOPTADO","=MAX(B26,B27)","mm2","0",bold=True,hl=VERDE)

    seccion(ws,31,"2) CORTANTE")
    out(ws,32,"Vc = 0.17*raiz(f'c)*b*d","=0.17*SQRT(B4)*B7*B9/1000","kN","0.00")
    out(ws,33,"phi*Vc","=B17*B32","kN","0.00")
    out(ws,34,"phi*Vc/2","=B17*B32/2","kN","0.00")
    out(ws,35,"Vs requerido = Vu/phi - Vc","=MAX(B13/B17-B32,0)","kN","0.00")
    out(ws,36,"Limite 0.33*raiz(f'c)*b*d","=0.33*SQRT(B4)*B7*B9/1000","kN","0.00")
    out(ws,37,"s max (depende de Vs)","=IF(B35<=B36,MIN(B9/2,600),MIN(B9/4,300))","mm","0")
    out(ws,38,"s por resistencia = Av*fyt*d/Vs","=IF(B35<=0,9999,B15*B6*B9/(B35*1000))","mm","0")
    out(ws,39,"s por minimo (Av,min)","=B15*B6/MAX(0.062*SQRT(B4)*B7,0.35*B7)","mm","0")
    out(ws,40,"RESULTADO cortante","=IF(B13<=B34,\"No requiere estribos\",\"Estribos @ \"&ROUND(MIN(B37,B38,B39),0)&\" mm\")","","General",bold=True,hl=AMAR)

    seccion(ws,42,"3) TORSION")
    out(ws,43,"Acp = b*h","=B7*B8","mm2","0")
    out(ws,44,"pcp = 2*(b+h)","=2*(B7+B8)","mm","0")
    out(ws,45,"phi*Tth (umbral)","=B17*0.083*SQRT(B4)*(B43^2/B44)/10^6","kN*m","0.00")
    out(ws,46,"Aoh = (b-2cc)*(h-2cc)","=(B7-2*B10)*(B8-2*B10)","mm2","0")
    out(ws,47,"ph = 2*((b-2cc)+(h-2cc))","=2*((B7-2*B10)+(B8-2*B10))","mm","0")
    out(ws,48,"Ao = 0.85*Aoh","=0.85*B46","mm2","0")
    out(ws,49,"At/s (1 rama) = Tu/(phi*2*Ao*fyt) ","=IF(B14<B45,0,B14*10^6/(B17*2*B48*B6))","mm2/mm","0.000")
    out(ws,50,"Al (acero long. por torsion) = (At/s)*ph*(fyt/fy)","=B49*B47*(B6/B5)","mm2","0")
    out(ws,51,"RESULTADO torsion","=IF(B14<B45,\"Despreciable\",\"Sumar At/s a estribos y repartir Al en el perimetro\")","","General",bold=True,hl=AMAR)
    ws.cell(row=53,column=1,value="Estribo total por cortante+torsion: (Av/2)/s [cortante, por rama] + At/s [torsion]. Theta=45 (cot=1).").font=Font(italic=True,size=9)
    ws.cell(row=54,column=1,value="Si el sistema es DMO/DES, ademas verifica diseno por capacidad (cortante a partir de Mpr).").font=Font(italic=True,size=9)

# ============================================================ RIOSTRAS
def hoja_riostras(wb):
    ws=wb.create_sheet("Riostras"); anchos(ws)
    titulo(ws,"A1","DISENO DE RIOSTRA - flexion y cortante simple","D1")
    seccion(ws,3,"DATOS DE ENTRADA")
    inp(ws,4,"f'c",21,"MPa"); inp(ws,5,"fy",420,"MPa"); inp(ws,6,"fyt",420,"MPa")
    inp(ws,7,"b",100,"mm"); inp(ws,8,"h",400,"mm"); inp(ws,9,"d",360,"mm")
    inp(ws,10,"Mu",18,"kN*m"); inp(ws,11,"Vu",20,"kN")
    inp(ws,12,"Av (estribo 2 ramas)",64,"mm2")
    inp(ws,13,"phi flexion",0.9,"-"); inp(ws,14,"phi cortante",0.75,"-")
    seccion(ws,17,"FLEXION")
    out(ws,18,"Rn","=B10*10^6/(B13*B7*B9^2)","MPa","0.000")
    out(ws,19,"rho","=(0.85*B4/B5)*(1-SQRT(1-2*B18/(0.85*B4)))","-","0.00000")
    out(ws,20,"As requerido","=B19*B7*B9","mm2","0")
    out(ws,21,"As minimo","=MAX(1.4/B5,0.25*SQRT(B4)/B5)*B7*B9","mm2","0")
    out(ws,22,"As ADOPTADO","=MAX(B20,B21)","mm2","0",bold=True,hl=VERDE)
    seccion(ws,24,"CORTANTE")
    out(ws,25,"Vc","=0.17*SQRT(B4)*B7*B9/1000","kN","0.00")
    out(ws,26,"phi*Vc","=B14*B25","kN","0.00")
    out(ws,27,"Vs requerido","=MAX(B11/B14-B25,0)","kN","0.00")
    out(ws,28,"s calculado","=IF(B27<=0,9999,B12*B6*B9/(B27*1000))","mm","0")
    out(ws,29,"RESULTADO","=IF(B11<=B26,\"Concreto resiste\",\"Estribos @ \"&ROUND(MIN(B28,B9/2),0)&\" mm\")","","General",bold=True,hl=AMAR)

def main():
    wb=Workbook()
    hoja_viguetas(wb); hoja_vigas(wb); hoja_riostras(wb)
    wb.save("Herramienta-Diseno-Viguetas-Vigas.xlsx")
    print("OK -> Herramienta-Diseno-Viguetas-Vigas.xlsx")

if __name__=="__main__":
    main()
