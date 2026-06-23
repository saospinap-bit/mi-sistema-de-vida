# -*- coding: utf-8 -*-
"""
Modelo financiero del Proyecto PEPI - Acueducto de Tado (Choco) en Excel.
IMPORTANTE: para que NINGUNA celda quede vacia en cualquier visor, este script
CALCULA todos los valores en Python y los escribe como NUMEROS (no formulas).
Los numeros coinciden exactamente con modelo_financiero.py.

Hojas: Supuestos | Diseno Caudal | Presupuesto | Amortizacion Deuda |
       Flujos de Caja | Indicadores | Riesgos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 1) CALCULOS (identicos a modelo_financiero.py)
# ============================================================
POB_URBANA = 11917
TASA_CREC = 0.008
HOR = 25
POB_DIS = POB_URBANA * (1 + TASA_CREC) ** HOR
DOT_NETA = 140.0
PERDIDAS = 0.25
DOT_BRUTA = DOT_NETA / (1 - PERDIDAS)
K1, K2 = 1.30, 1.60
Q_MED = DOT_BRUTA * POB_DIS / 86400.0
QMD = Q_MED * K1
QMH = QMD * K2
PTAP = 45.0

presupuesto = [
    ("Preliminares y obras provisionales", 350),
    ("Captacion (bocatoma rio San Juan) + desarenador", 950),
    ("Aduccion y conduccion (PVC/HD ~4 km)", 1900),
    ("PTAP compacta 45 L/s (civil+equipos+dosif.+lab)", 3800),
    ("Tanque de almacenamiento (1.000 m3)", 1200),
    ("Estacion de bombeo (bombas + variadores)", 750),
    ("Redes de distribucion y sectorizacion", 2700),
    ("Micromedicion (~3.700 medidores) + macromedicion", 720),
    ("Estudios, disenios e interventoria", 1100),
]
CAPEX = sum(v for _, v in presupuesto)

PCT_PUB, PCT_DEUDA, PCT_EQ = 0.90, 0.05, 0.05
APORTE = CAPEX * PCT_PUB
DEUDA = CAPEX * PCT_DEUDA
EQUITY = CAPEX * PCT_EQ
KD, KE, IMP, PLAZO = 0.11, 0.14, 0.35, 12
WACC = 0.5 * KE + 0.5 * KD * (1 - IMP)
DEP = CAPEX / HOR

def cuota_fija(p, i, n):
    return p * i / (1 - (1 + i) ** (-n))
CUOTA = cuota_fija(DEUDA, KD, PLAZO)

# Operacion
SUS1, CSUS = 3400, 0.012
FAC1, CFAC = 28000.0, 0.035
OPEX1, COPEX = 1050.0, 0.04
TSOC, BENH = 0.09, 75000.0

ingresos, opex = [], []
for t in range(1, HOR + 1):
    ingresos.append(SUS1 * (1 + CSUS) ** (t - 1) * FAC1 * (1 + CFAC) ** (t - 1) * 12 / 1e6)
    opex.append(OPEX1 * (1 + COPEX) ** (t - 1))

# Amortizacion
saldo = DEUDA
amort_tabla = []  # (saldo_ini, interes, abono, cuota, saldo_fin)
for t in range(1, HOR + 1):
    if t <= PLAZO:
        si = saldo
        it = si * KD
        ab = CUOTA - it
        sf = si - ab
        saldo = sf
        amort_tabla.append((si, it, ab, CUOTA, sf))
    else:
        amort_tabla.append((0.0, 0.0, 0.0, 0.0, 0.0))

# Flujos
ebitda, ebit, fclp, fcla, fcbanco = [], [], [], [], []
for t in range(HOR):
    e = ingresos[t] - opex[t]; ebitda.append(e)
    eb = e - DEP; ebit.append(eb)
    fclp.append(eb * (1 - IMP) + DEP)
    it = amort_tabla[t][1]; ab = amort_tabla[t][2]
    uai = eb - it
    fcla.append(uai * (1 - IMP) + DEP - ab)
    fcbanco.append(-(it + ab))

flujo_proy = [-CAPEX] + fclp
flujo_inv = [-EQUITY] + fcla

def vpn(tasa, fl):
    return sum(f / (1 + tasa) ** i for i, f in enumerate(fl))

def tir(fl):
    lo, hi = -0.95, 5.0
    flo = vpn(lo, fl); fhi = vpn(hi, fl)
    if flo * fhi > 0:
        return None
    for _ in range(300):
        mid = (lo + hi) / 2; fm = vpn(mid, fl)
        if abs(fm) < 1e-7:
            return mid
        if flo * fm < 0:
            hi = mid
        else:
            lo = mid; flo = fm
    return (lo + hi) / 2

VPN_PROY = vpn(WACC, flujo_proy); TIR_PROY = tir(flujo_proy)
VPN_INV = vpn(KE, flujo_inv); TIR_INV = tir(flujo_inv)

# Socioeconomico
beneficios = [SUS1 * (1 + CSUS) ** t * BENH * 12 / 1e6 for t in range(HOR)]
flujo_econ = [-CAPEX] + [beneficios[t] - opex[t] for t in range(HOR)]
VPN_ECON = vpn(TSOC, flujo_econ); TIR_ECON = tir(flujo_econ)
RBC = (sum(beneficios[t] / (1 + TSOC) ** (t + 1) for t in range(HOR)) /
       (CAPEX + sum(opex[t] / (1 + TSOC) ** (t + 1) for t in range(HOR))))

# ============================================================
# 2) ESTILOS
# ============================================================
AZUL = "1F4E79"; CLARO = "D6E4F0"; GRIS = "F2F2F2"
f_tit = Font(bold=True, color="FFFFFF", size=12)
f_hdr = Font(bold=True, color="FFFFFF", size=10)
f_b = Font(bold=True)
fa = PatternFill("solid", fgColor=AZUL); fc = PatternFill("solid", fgColor=CLARO)
fg = PatternFill("solid", fgColor=GRIS)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
rgt = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="BFBFBF")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)

def title(ws, rng, text):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = text; c.font = f_tit; c.fill = fa; c.alignment = ctr
    ws.row_dimensions[1].height = 26

def header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h); c.font = f_hdr; c.fill = fa; c.alignment = ctr; c.border = bd

wb = Workbook()

# ============================================================
# HOJA 1: SUPUESTOS
# ============================================================
ws = wb.active; ws.title = "Supuestos"; ws.sheet_view.showGridLines = False
title(ws, "A1:C1", "SUPUESTOS - PROYECTO PEPI ACUEDUCTO DE TADO (CHOCO)")
ws["A2"] = "Cifras en COP millones (salvo % y unidades). Fuente poblacion: DANE CNPV 2018."
ws["A2"].font = Font(italic=True, size=8)
header(ws, 3, ["Parametro", "Valor", "Unidad / fuente"])
sup = [
    ("Poblacion urbana 2025 (cabecera)", POB_URBANA, "hab (DANE)", "#,##0"),
    ("Tasa de crecimiento", TASA_CREC, "% anual", "0.00%"),
    ("Horizonte de diseno/evaluacion", HOR, "anios (Res 0330)", "#,##0"),
    ("Poblacion de diseno (anio 25)", POB_DIS, "hab", "#,##0"),
    ("Dotacion neta (alt<1000 m)", DOT_NETA, "L/hab.dia (Res 0330)", "#,##0.0"),
    ("Perdidas maximas admisibles", PERDIDAS, "% (Res 0330)", "0%"),
    ("Dotacion bruta", DOT_BRUTA, "L/hab.dia", "#,##0.0"),
    ("Coef. maximo diario k1", K1, "Res 0330", "#,##0.00"),
    ("Coef. maximo horario k2", K2, "Res 0330", "#,##0.00"),
    ("Caudal medio diario", Q_MED, "L/s", "#,##0.0"),
    ("Caudal maximo diario (PTAP)", QMD, "L/s", "#,##0.0"),
    ("Caudal maximo horario", QMH, "L/s", "#,##0.0"),
    ("Capacidad PTAP adoptada", PTAP, "L/s", "#,##0"),
    ("CAPEX total", CAPEX, "COP MM", "#,##0"),
    ("% Aporte publico (SGR/PDA/SGP)", PCT_PUB, "%", "0%"),
    ("% Deuda", PCT_DEUDA, "%", "0%"),
    ("% Equity", PCT_EQ, "%", "0%"),
    ("Aporte publico", APORTE, "COP MM", "#,##0"),
    ("Deuda", DEUDA, "COP MM", "#,##0.0"),
    ("Equity", EQUITY, "COP MM", "#,##0.0"),
    ("Kd (costo deuda)", KD, "%", "0.0%"),
    ("Ke (costo equity)", KE, "%", "0.0%"),
    ("Tasa de impuestos", IMP, "%", "0%"),
    ("Plazo de la deuda", PLAZO, "anios", "#,##0"),
    ("Depreciacion anual", DEP, "COP MM", "#,##0.0"),
    ("Cuota anual deuda (frances)", CUOTA, "COP MM", "#,##0.0"),
    ("WACC (capital remunerado 50/50)", WACC, "%", "0.00%"),
    ("Suscriptores anio 1", SUS1, "suscriptores", "#,##0"),
    ("Crecimiento suscriptores", CSUS, "% anual", "0.0%"),
    ("Factura media acueducto anio 1", FAC1, "COP/mes", "#,##0"),
    ("Crecimiento tarifa", CFAC, "% anual", "0.0%"),
    ("OPEX anio 1", OPEX1, "COP MM", "#,##0"),
    ("Crecimiento OPEX", COPEX, "% anual", "0.0%"),
    ("Tasa social de descuento (DNP)", TSOC, "%", "0%"),
    ("Beneficio social por hogar", BENH, "COP/mes", "#,##0"),
]
r = 4
for nombre, val, unit, fmt in sup:
    ws.cell(r, 1, nombre).font = f_b
    ws.cell(r, 1).fill = fc
    c = ws.cell(r, 2, round(val, 4) if isinstance(val, float) else val)
    c.number_format = fmt; c.alignment = rgt
    ws.cell(r, 3, unit).alignment = lft
    for j in range(1, 4):
        ws.cell(r, j).border = bd
    r += 1
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 20

# ============================================================
# HOJA 2: DISENO CAUDAL
# ============================================================
wd = wb.create_sheet("Diseno Caudal"); wd.sheet_view.showGridLines = False
title(wd, "A1:C1", "CAUDAL DE DISENIO (Resolucion 0330 de 2017)")
header(wd, 2, ["Concepto", "Valor", "Unidad"])
dis = [
    ("Poblacion de diseno (anio 25)", POB_DIS, "hab", "#,##0"),
    ("Dotacion neta", DOT_NETA, "L/hab.dia", "#,##0.0"),
    ("Perdidas admisibles", PERDIDAS, "%", "0%"),
    ("Dotacion bruta", DOT_BRUTA, "L/hab.dia", "#,##0.0"),
    ("Caudal medio diario (Qmd)", Q_MED, "L/s", "#,##0.0"),
    ("Caudal maximo diario (QMD = Qmd x k1)", QMD, "L/s -> PTAP", "#,##0.0"),
    ("Caudal maximo horario (QMH = QMD x k2)", QMH, "L/s -> redes", "#,##0.0"),
    ("Capacidad PTAP adoptada", PTAP, "L/s", "#,##0"),
]
r = 3
for nombre, val, unit, fmt in dis:
    wd.cell(r, 1, nombre).font = f_b; wd.cell(r, 1).fill = fc
    c = wd.cell(r, 2, round(val, 2)); c.number_format = fmt; c.alignment = rgt
    wd.cell(r, 3, unit).alignment = lft
    for j in range(1, 4):
        wd.cell(r, j).border = bd
    r += 1
wd.column_dimensions["A"].width = 36; wd.column_dimensions["B"].width = 14; wd.column_dimensions["C"].width = 16

# ============================================================
# HOJA 3: PRESUPUESTO
# ============================================================
wp = wb.create_sheet("Presupuesto"); wp.sheet_view.showGridLines = False
title(wp, "A1:C1", "PRESUPUESTO DE OBRA - CAPEX (precios de referencia; validar con APU)")
header(wp, 2, ["Capitulo", "COP MM", "% del total"])
r = 3
for nombre, val in presupuesto:
    wp.cell(r, 1, nombre).alignment = lft
    wp.cell(r, 2, val).number_format = "#,##0"; wp.cell(r, 2).alignment = rgt
    pc = wp.cell(r, 3, val / CAPEX); pc.number_format = "0.0%"; pc.alignment = rgt
    for j in range(1, 4):
        wp.cell(r, j).border = bd
    if r % 2 == 0:
        for j in range(1, 4):
            wp.cell(r, j).fill = fg
    r += 1
wp.cell(r, 1, "TOTAL CAPEX").font = f_b; wp.cell(r, 1).fill = fc
wp.cell(r, 2, CAPEX).number_format = "#,##0"; wp.cell(r, 2).font = f_b; wp.cell(r, 2).alignment = rgt
wp.cell(r, 3, 1.0).number_format = "0%"; wp.cell(r, 3).font = f_b; wp.cell(r, 3).alignment = rgt
for j in range(1, 4):
    wp.cell(r, j).border = bd
wp.column_dimensions["A"].width = 52; wp.column_dimensions["B"].width = 12; wp.column_dimensions["C"].width = 12

# ============================================================
# HOJA 4: AMORTIZACION
# ============================================================
wa = wb.create_sheet("Amortizacion Deuda"); wa.sheet_view.showGridLines = False
title(wa, "A1:F1", "TABLA DE AMORTIZACION DE LA DEUDA (Sistema Frances - cuota fija)")
header(wa, 2, ["Anio", "Saldo inicial", "Interes", "Abono capital", "Cuota", "Saldo final"])
r = 3
for t in range(HOR):
    si, it, ab, cu, sf = amort_tabla[t]
    wa.cell(r, 1, t + 1).alignment = ctr
    for j, val in zip(range(2, 7), (si, it, ab, cu, sf)):
        c = wa.cell(r, j, round(val, 1)); c.number_format = "#,##0.0"; c.alignment = rgt
    for j in range(1, 7):
        wa.cell(r, j).border = bd
    if (t + 1) % 2 == 0:
        for j in range(1, 7):
            wa.cell(r, j).fill = fg
    r += 1
for col, w in zip("ABCDEF", [8, 15, 13, 15, 13, 15]):
    wa.column_dimensions[col].width = w

# ============================================================
# HOJA 5: FLUJOS DE CAJA
# ============================================================
wf = wb.create_sheet("Flujos de Caja"); wf.sheet_view.showGridLines = False
title(wf, "A1:K1", "FLUJOS DE CAJA - PROYECTO, INVERSIONISTA Y BANCO (COP MM)")
header(wf, 2, ["Anio", "Ingresos", "OPEX", "EBITDA", "Deprec.", "EBIT",
               "Intereses", "FC Proyecto", "Util.neta", "FC Inversionista", "FC Banco"])
# anio 0
wf.cell(3, 1, 0).alignment = ctr
wf.cell(3, 8, round(-CAPEX, 1)); wf.cell(3, 10, round(-EQUITY, 1)); wf.cell(3, 11, round(DEUDA, 1))
for j in range(1, 12):
    cell = wf.cell(3, j); cell.border = bd
    cell.alignment = ctr if j == 1 else rgt
    if j > 1 and cell.value is not None:
        cell.number_format = "#,##0.0"
r = 4
for t in range(HOR):
    it = amort_tabla[t][1]; ab = amort_tabla[t][2]
    uai = ebit[t] - it
    util = uai * (1 - IMP)
    vals = [t + 1, ingresos[t], opex[t], ebitda[t], DEP, ebit[t], it,
            fclp[t], util, fcla[t], fcbanco[t]]
    for j, val in enumerate(vals, 1):
        c = wf.cell(r, j, round(val, 1) if j > 1 else val)
        c.alignment = ctr if j == 1 else rgt
        if j > 1:
            c.number_format = "#,##0.0"
        c.border = bd
    if (t + 1) % 2 == 0:
        for j in range(1, 12):
            wf.cell(r, j).fill = fg
    r += 1
for col, w in zip("ABCDEFGHIJK", [6, 10, 9, 9, 9, 9, 10, 12, 9, 14, 11]):
    wf.column_dimensions[col].width = w

# ============================================================
# HOJA 6: INDICADORES
# ============================================================
wi = wb.create_sheet("Indicadores"); wi.sheet_view.showGridLines = False
title(wi, "A1:C1", "INDICADORES DE RENTABILIDAD Y EVALUACION SOCIOECONOMICA")
header(wi, 2, ["Indicador", "Valor", "Criterio / interpretacion"])
ind = [
    ("WACC (capital remunerado)", WACC, "0.00%", "Costo de capital"),
    ("VPN Proyecto 'puro' (@WACC)", VPN_PROY, "#,##0", "<0: no rentable solo con tarifas"),
    ("TIR Proyecto 'puro'", TIR_PROY, "0.00%", "< WACC -> no rentable solo"),
    ("VPN Inversionista (@Ke)", VPN_INV, "#,##0", ">0: crea valor"),
    ("TIR Inversionista", TIR_INV, "0.00%", "> Ke (14%) -> viable"),
    ("VPN Economico (@tasa social 9%)", VPN_ECON, "#,##0", ">0: socialmente viable"),
    ("TIR Economica (TIRE)", TIR_ECON, "0.00%", "> tasa social (9%)"),
    ("Relacion Beneficio/Costo (B/C)", RBC, "#,##0.00", ">1: socialmente viable"),
]
r = 3
for nombre, val, fmt, crit in ind:
    wi.cell(r, 1, nombre).font = f_b; wi.cell(r, 1).fill = fc
    c = wi.cell(r, 2, round(val, 4) if val is not None else "NA")
    c.number_format = fmt; c.alignment = rgt; c.font = f_b
    wi.cell(r, 3, crit).alignment = lft
    for j in range(1, 4):
        wi.cell(r, j).border = bd
    r += 1
wi.column_dimensions["A"].width = 32; wi.column_dimensions["B"].width = 14; wi.column_dimensions["C"].width = 34

# Sensibilidad
r += 1
wi.merge_cells(f"A{r}:D{r}")
wi.cell(r, 1, "SENSIBILIDAD: % aporte publico vs viabilidad del inversionista")
wi.cell(r, 1).font = f_tit; wi.cell(r, 1).fill = fa; wi.cell(r, 1).alignment = ctr
r += 1
header(wi, r, ["Aporte publico", "Equity (COP MM)", "TIR inversionista", "VPN inv (COP MM)"])
r += 1
for g in [0.70, 0.80, 0.85, 0.90, 0.95]:
    pe = (1 - g) / 2; eq = CAPEX * pe; de = CAPEX * pe
    cu = cuota_fija(de, KD, PLAZO); sa = de
    fl = [-eq]
    for t in range(HOR):
        if t < PLAZO:
            i_ = sa * KD; a_ = cu - i_; sa -= a_
        else:
            i_ = a_ = 0.0
        uai = ebit[t] - i_
        fl.append(uai * (1 - IMP) + DEP - a_)
    ti = tir(fl); vi = vpn(KE, fl)
    wi.cell(r, 1, g).number_format = "0%"; wi.cell(r, 1).alignment = ctr
    wi.cell(r, 2, round(eq, 0)).number_format = "#,##0"; wi.cell(r, 2).alignment = rgt
    wi.cell(r, 3, round(ti, 4)).number_format = "0.00%"; wi.cell(r, 3).alignment = rgt
    wi.cell(r, 4, round(vi, 0)).number_format = "#,##0"; wi.cell(r, 4).alignment = rgt
    for j in range(1, 5):
        wi.cell(r, j).border = bd
        if g == 0.90:
            wi.cell(r, j).fill = fc
    r += 1

# ============================================================
# HOJA 7: RIESGOS
# ============================================================
wr = wb.create_sheet("Riesgos"); wr.sheet_view.showGridLines = False
title(wr, "A1:G1", "MATRIZ DE RIESGOS (20) - Nivel=PxI [Bajo 1-6 / Medio 8-12 / Alto 15-25]")
header(wr, 2, ["#", "Riesgo", "Tipo", "P", "I", "PxI", "Nivel"])
riesgos = [
    [1, "Sobrecostos en la construccion", "Financiero", 4, 4],
    [2, "Retrasos en el cronograma de obra", "Operativo", 4, 3],
    [3, "Demoras en gestion predial y servidumbres", "Legal", 3, 4],
    [4, "Insuficiencia/retraso del aporte publico (SGR/PDA)", "Financiero", 3, 5],
    [5, "Incremento de tasas de interes", "Financiero", 3, 4],
    [6, "Recaudo tarifario inferior al proyectado", "Comercial", 4, 4],
    [7, "Cambios regulatorios marco tarifario (CRA)", "Regulatorio", 2, 4],
    [8, "Alta turbiedad/contaminacion (mercurio) rio San Juan", "Ambiental", 4, 5],
    [9, "Inundaciones por lluvias extremas (Choco)", "Ambiental", 4, 4],
    [10, "Inadecuada disposicion de lodos PTAP", "Ambiental", 3, 4],
    [11, "Vertimiento de aguas de lavado sin tratar", "Ambiental", 3, 4],
    [12, "Afectacion de fauna/flora en obras", "Ambiental", 2, 3],
    [13, "Fallas o paradas de equipos de bombeo", "Operativo", 3, 3],
    [14, "Interrupcion del suministro electrico", "Operativo", 3, 4],
    [15, "Accidentes laborales en obra", "SST", 3, 4],
    [16, "Dificultad logistica y de acceso (transporte)", "Operativo", 4, 3],
    [17, "Orden publico en la region", "Social", 3, 4],
    [18, "Errores en estudios y disenios tecnicos", "Tecnico", 2, 4],
    [19, "Incumplimiento del contratista", "Contractual", 2, 4],
    [20, "Inflacion de costos operativos", "Financiero", 3, 3],
]
def nivel(pi):
    return "Alto" if pi >= 15 else ("Medio" if pi >= 8 else "Bajo")
r = 3
for ri in riesgos:
    pi = ri[3] * ri[4]
    wr.cell(r, 1, ri[0]).alignment = ctr
    wr.cell(r, 2, ri[1]).alignment = lft
    wr.cell(r, 3, ri[2]).alignment = ctr
    wr.cell(r, 4, ri[3]).alignment = ctr
    wr.cell(r, 5, ri[4]).alignment = ctr
    wr.cell(r, 6, pi).alignment = ctr
    wr.cell(r, 7, nivel(pi)).alignment = ctr
    for j in range(1, 8):
        wr.cell(r, j).border = bd
    r += 1
for col, w in zip("ABCDEFG", [5, 44, 14, 5, 5, 7, 9]):
    wr.column_dimensions[col].width = w

base = r + 2
wr.merge_cells(f"A{base}:H{base}")
wr.cell(base, 1, "INTERVENCION DE 10 RIESGOS (5 ambientales) Y RECALIFICACION")
wr.cell(base, 1).font = f_tit; wr.cell(base, 1).fill = fa; wr.cell(base, 1).alignment = ctr
header(wr, base + 1, ["#", "Riesgo", "Tipo", "Intervencion", "P res", "I res", "PxI res", "Nivel res"])
inter = [
    [1, "Sobrecostos de construccion", "Financiero", "Contrato EPC precio fijo + contingencia 10% + control de cambios", 2, 3],
    [6, "Bajo recaudo tarifario", "Comercial", "Cultura de pago, subsidios Ley 142, corte por mora, micromedicion", 2, 3],
    [4, "Insuficiencia del aporte publico", "Financiero", "Cierre financiero previo; convenios SGR/PDA y cooperacion", 2, 3],
    [8, "Turbiedad/mercurio rio San Juan", "Ambiental", "Bocatoma con desarenador, pretratamiento y monitoreo de mercurio", 2, 4],
    [9, "Inundaciones por lluvias extremas", "Ambiental", "Obras elevadas, drenaje, diseno hidrologico y plan de contingencia", 2, 3],
    [10, "Disposicion de lodos", "Ambiental", "Lechos de secado y disposicion final autorizada (CODECHOCO)", 1, 3],
    [11, "Vertimiento aguas de lavado", "Ambiental", "Recirculacion y tratamiento previo, permiso de vertimientos", 1, 3],
    [12, "Afectacion fauna/flora", "Ambiental", "Plan de Manejo Ambiental y compensacion forestal", 1, 2],
    [16, "Dificultad logistica/acceso", "Operativo", "Plan logistico, acopio de materiales y cronograma con holguras", 2, 2],
    [15, "Accidentes laborales", "SST", "SG-SST (Dec 1072/2015), capacitacion, EPP, permisos", 1, 4],
]
r = base + 2
for it in inter:
    pir = it[4] * it[5]
    wr.cell(r, 1, it[0]).alignment = ctr
    wr.cell(r, 2, it[1]).alignment = lft
    wr.cell(r, 3, it[2]).alignment = ctr
    wr.cell(r, 4, it[3]).alignment = lft
    wr.cell(r, 5, it[4]).alignment = ctr
    wr.cell(r, 6, it[5]).alignment = ctr
    wr.cell(r, 7, pir).alignment = ctr
    wr.cell(r, 8, nivel(pir)).alignment = ctr
    for j in range(1, 9):
        wr.cell(r, j).border = bd
    r += 1
wr.column_dimensions["D"].width = 54; wr.column_dimensions["H"].width = 10

# Forzar recalculo al abrir (por si el visor lo soporta) - aunque ya hay valores
wb.calculation.fullCalcOnLoad = True
wb.save("MODELO FINANCIERO PEPI.xlsx")
print("OK -> MODELO FINANCIERO PEPI.xlsx (valores escritos; sin celdas vacias)")
print(f"   CAPEX={CAPEX:,.0f}  WACC={WACC:.2%}  VPN_proy={VPN_PROY:,.0f}  "
      f"VPN_inv={VPN_INV:,.0f}  TIR_inv={TIR_INV:.2%}  VPN_econ={VPN_ECON:,.0f}  BC={RBC:.2f}")
