#!/usr/bin/env python3
"""
Generador de 'Notas-y-Promedios.xlsx'  (UNAL - escala 0.0 a 5.0, aprueba 3.0)
Hojas:
  1) Resumen Semestre  -> notas por materia, estado, PAPA del semestre (ponderado por creditos)
  2) Calculadora Materia -> por cada materia: evaluaciones con % y nota, definitiva y
                            cuanto necesitas en lo que falta para pasar
  3) PAPA Carrera -> promedio acumulado de toda la carrera ponderado por creditos
  4) Prioridades (auto) -> ranking automatico de riesgo academico
Todo con FORMULAS (se recalcula solo al cambiar las notas).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

AZUL = "37474F"; GRIS = "546E7A"; CLARO = "ECEFF1"
ROJO = "FFCDD2"; AMAR = "FFF9C4"; VERDE = "C8E6C9"
APRUEBA = 3.0

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Materias del semestre (editables). credito + nombre.  La critica va primero.
MATERIAS = [
    ("Diseno Estructural (PROYECTO)", 4),
    ("Materia obligatoria 1", 3),
    ("Materia obligatoria 2", 3),
    ("Materia obligatoria 3", 3),
    ("Materia optativa 1", 2),
    ("Materia optativa 2", 2),
]

def titulo(ws, celda, texto, span):
    ws.merge_cells(f"{celda}:{span}")
    c = ws[celda]
    c.value = texto
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = CENTER

def header(ws, row, col, texto, color=GRIS, width=None):
    c = ws.cell(row=row, column=col, value=texto)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = CENTER
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


# ----------------------------------------------------------------------------
# HOJA 1: RESUMEN SEMESTRE
# ----------------------------------------------------------------------------
def hoja_resumen(wb):
    ws = wb.active
    ws.title = "Resumen Semestre"
    titulo(ws, "A1", "RESUMEN DEL SEMESTRE (9o)  -  Aprueba: 3.0", "G1")

    cols = [("Materia", 32), ("Creditos", 9), ("Nota Definitiva", 13),
            ("Aporte (cred x nota)", 14), ("Estado", 13), ("Prioridad", 12), ("Falta %", 9)]
    for j, (txt, w) in enumerate(cols, start=1):
        header(ws, 3, j, txt, width=w)

    first = 4
    for i, (nombre, cred) in enumerate(MATERIAS):
        r = first + i
        ws.cell(row=r, column=1, value=nombre).alignment = LEFT
        ws.cell(row=r, column=2, value=cred).alignment = CENTER
        # Nota definitiva: la trae la hoja Calculadora (se llena alla); por defecto editable
        ws.cell(row=r, column=3, value=f"='Calculadora Materia'!{calc_def_cell(i)}").alignment = CENTER
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").alignment = CENTER
        ws.cell(row=r, column=5, value=f'=IF(C{r}="","-",IF(C{r}>={APRUEBA},"Aprobada","EN RIESGO"))').alignment = CENTER
        ws.cell(row=r, column=6,
                value=f'=IF(C{r}="","-",IF(C{r}<{APRUEBA},"ALTA",IF(C{r}<3.5,"MEDIA","OK")))').alignment = CENTER
        ws.cell(row=r, column=7, value=f"='Calculadora Materia'!{calc_falta_cell(i)}").alignment = CENTER
        for j in range(1, 8):
            ws.cell(row=r, column=j).border = BORDER

    last = first + len(MATERIAS) - 1
    # Totales / PAPA semestre
    tr = last + 1
    ws.cell(row=tr, column=1, value="PAPA SEMESTRE (ponderado por creditos)").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=f"=SUM(B{first}:B{last})").font = Font(bold=True)
    ws.cell(row=tr, column=2).alignment = CENTER
    papa = ws.cell(row=tr, column=3,
                   value=f"=IFERROR(ROUND(SUMPRODUCT(B{first}:B{last},C{first}:C{last})/SUM(B{first}:B{last}),2),0)")
    papa.font = Font(bold=True, size=12, color="1B5E20")
    papa.alignment = CENTER
    for j in range(1, 4):
        ws.cell(row=tr, column=j).border = BORDER
        ws.cell(row=tr, column=j).fill = PatternFill("solid", fgColor=CLARO)

    # Conteo de materias en riesgo
    ws.cell(row=tr + 2, column=1, value="Materias EN RIESGO:").font = Font(bold=True, color="B71C1C")
    ws.cell(row=tr + 2, column=2,
            value=f'=COUNTIF(E{first}:E{last},"EN RIESGO")').font = Font(bold=True, color="B71C1C")

    # Formato condicional por nota (col C)
    rango = f"C{first}:C{last}"
    ws.conditional_formatting.add(rango, CellIsRule(operator="lessThan", formula=[str(APRUEBA)],
                                  fill=PatternFill("solid", fgColor=ROJO)))
    ws.conditional_formatting.add(rango, CellIsRule(operator="between", formula=[str(APRUEBA), "3.49"],
                                  fill=PatternFill("solid", fgColor=AMAR)))
    ws.conditional_formatting.add(rango, CellIsRule(operator="greaterThanOrEqual", formula=["3.5"],
                                  fill=PatternFill("solid", fgColor=VERDE)))
    ws.freeze_panes = "A4"


# ----------------------------------------------------------------------------
# HOJA 2: CALCULADORA POR MATERIA
# ----------------------------------------------------------------------------
N_EVAL = 7          # filas de evaluaciones por materia
BLOCK = N_EVAL + 9  # alto de cada bloque de materia (incluye fila en blanco de separacion)

def block_start(i):
    return 3 + i * BLOCK

def calc_def_cell(i):
    # celda de "Nota definitiva" del bloque i  (rd = block_start + N_EVAL + 4)
    return f"D{block_start(i) + N_EVAL + 4}"

def calc_falta_cell(i):
    # celda de "% que falta por cursar"  (rf = block_start + N_EVAL + 6)
    return f"D{block_start(i) + N_EVAL + 6}"

def hoja_calculadora(wb):
    ws = wb.create_sheet("Calculadora Materia")
    titulo(ws, "A1", "CALCULADORA POR MATERIA  -  pon el % (peso) y la nota (0-5)", "D1")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    for i, (nombre, _) in enumerate(MATERIAS):
        s = block_start(i)
        # Nombre materia
        ws.merge_cells(f"A{s}:D{s}")
        c = ws[f"A{s}"]
        c.value = f"MATERIA: {nombre}"
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GRIS)
        c.alignment = LEFT
        # Encabezados
        for j, txt in enumerate(["Evaluacion", "% (peso)", "Nota (0-5)", "Aporte"], start=1):
            hc = ws.cell(row=s + 1, column=j, value=txt)
            hc.font = Font(bold=True)
            hc.fill = PatternFill("solid", fgColor=CLARO)
            hc.alignment = CENTER
            hc.border = BORDER
        ev0 = s + 2
        for k in range(N_EVAL):
            r = ev0 + k
            ws.cell(row=r, column=1, value=f"Eval {k+1}").alignment = LEFT
            ws.cell(row=r, column=2).alignment = CENTER   # % (vacio)
            ws.cell(row=r, column=3).alignment = CENTER   # nota (vacio)
            ws.cell(row=r, column=4, value=f"=IFERROR(B{r}*C{r}/100,0)").alignment = CENTER
            for j in range(1, 5):
                ws.cell(row=r, column=j).border = BORDER
        evN = ev0 + N_EVAL - 1
        # Suma %
        rs = evN + 1
        ws.cell(row=rs, column=1, value="Suma % (debe dar 100)").font = Font(bold=True)
        ws.cell(row=rs, column=2, value=f"=SUM(B{ev0}:B{evN})").alignment = CENTER
        ws.cell(row=rs, column=3,
                value=f'=IF(B{rs}=100,"OK","REVISAR")').alignment = CENTER
        # Nota acumulada hasta ahora
        ra = evN + 2
        ws.cell(row=ra, column=1, value="Nota acumulada (lo ya calificado)").font = Font(bold=True)
        ws.cell(row=ra, column=4, value=f"=ROUND(SUM(D{ev0}:D{evN}),2)").alignment = CENTER
        # Nota definitiva (si ya estuviera todo el 100%)
        rd = evN + 3
        ws.cell(row=rd, column=1, value="NOTA DEFINITIVA").font = Font(bold=True, color="1B5E20")
        ws.cell(row=rd, column=4, value=f"=ROUND(SUM(D{ev0}:D{evN}),2)").font = Font(bold=True, color="1B5E20")
        ws.cell(row=rd, column=4).alignment = CENTER
        # % cursado y % faltante
        rc = evN + 4
        ws.cell(row=rc, column=1, value="% ya cursado").alignment = LEFT
        ws.cell(row=rc, column=4, value=f'=SUMIF(C{ev0}:C{evN},">0",B{ev0}:B{evN})').alignment = CENTER
        rf = evN + 5
        ws.cell(row=rf, column=1, value="% que falta por cursar").font = Font(bold=True)
        ws.cell(row=rf, column=4, value=f"=100-D{rc}").font = Font(bold=True)
        ws.cell(row=rf, column=4).alignment = CENTER
        # Nota que necesito en lo que falta para pasar (3.0)
        rn = evN + 6
        ws.cell(row=rn, column=1, value=f"Nota que necesito (en lo que falta) para pasar {APRUEBA}").font = Font(bold=True, color="B71C1C")
        ws.cell(row=rn, column=4,
                value=f'=IFERROR(IF(D{rf}=0,"-",ROUND(({APRUEBA}-D{ra})/(D{rf}/100),2)),"-")').font = Font(bold=True, color="B71C1C")
        ws.cell(row=rn, column=4).alignment = CENTER
        for rr in (rs, ra, rd, rc, rf, rn):
            for j in range(1, 5):
                ws.cell(row=rr, column=j).border = BORDER


# ----------------------------------------------------------------------------
# HOJA 3: PAPA CARRERA
# ----------------------------------------------------------------------------
def hoja_carrera(wb):
    ws = wb.create_sheet("PAPA Carrera")
    titulo(ws, "A1", "PROMEDIO ACUMULADO DE CARRERA (PAPA)", "D1")
    for j, (txt, w) in enumerate([("Semestre", 12), ("Creditos", 12),
                                   ("Promedio semestre", 18), ("Aporte", 12)], start=1):
        header(ws, 3, j, txt, width=w)
    first = 4
    for s in range(1, 10):  # semestres 1 a 9
        r = first + s - 1
        ws.cell(row=r, column=1, value=f"Sem {s}").alignment = CENTER
        ws.cell(row=r, column=2).alignment = CENTER     # creditos (a llenar)
        if s == 9:
            ws.cell(row=r, column=3, value="='Resumen Semestre'!C10").alignment = CENTER  # link aproximado
        ws.cell(row=r, column=3).alignment = CENTER     # promedio (a llenar)
        ws.cell(row=r, column=4, value=f"=IFERROR(B{r}*C{r},0)").alignment = CENTER
        for j in range(1, 5):
            ws.cell(row=r, column=j).border = BORDER
    last = first + 8
    tr = last + 1
    ws.cell(row=tr, column=1, value="PAPA CARRERA").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=f"=SUM(B{first}:B{last})").alignment = CENTER
    pc = ws.cell(row=tr, column=3,
                 value=f"=IFERROR(ROUND(SUMPRODUCT(B{first}:B{last},C{first}:C{last})/SUM(B{first}:B{last}),2),0)")
    pc.font = Font(bold=True, size=12, color="1B5E20")
    pc.alignment = CENTER
    for j in range(1, 4):
        ws.cell(row=tr, column=j).fill = PatternFill("solid", fgColor=CLARO)
        ws.cell(row=tr, column=j).border = BORDER
    ws.cell(row=tr + 2, column=1,
            value="Nota: llena creditos y promedio de cada semestre pasado. El Sem 9 se calcula solo.").font = Font(italic=True, size=9)


# ----------------------------------------------------------------------------
# HOJA 4: PRIORIDADES (AUTO)
# ----------------------------------------------------------------------------
def hoja_prioridades(wb):
    ws = wb.create_sheet("Prioridades (auto)")
    titulo(ws, "A1", "PRIORIDADES ACADEMICAS (automatico)", "D1")
    for j, (txt, w) in enumerate([("Materia", 32), ("Nota", 10),
                                   ("Prioridad", 14), ("Que hacer", 34)], start=1):
        header(ws, 3, j, txt, width=w)
    first = 4
    n = len(MATERIAS)
    src_first = 4
    src_last = 3 + n
    for i in range(n):
        r = first + i
        sr = src_first + i
        ws.cell(row=r, column=1, value=f"='Resumen Semestre'!A{sr}").alignment = LEFT
        ws.cell(row=r, column=2, value=f"='Resumen Semestre'!C{sr}").alignment = CENTER
        ws.cell(row=r, column=3, value=f"='Resumen Semestre'!F{sr}").alignment = CENTER
        ws.cell(row=r, column=4,
                value=(f'=IF(B{r}="","-",'
                       f'IF(B{r}<{APRUEBA},"URGENTE: bloques diarios + hablar con el profe",'
                       f'IF(B{r}<3.5,"Reforzar esta semana",'
                       f'"Mantener, repaso ligero")))')).alignment = LEFT
        for j in range(1, 5):
            ws.cell(row=r, column=j).border = BORDER
    last = first + n - 1
    # resaltar prioridad ALTA
    ws.conditional_formatting.add(f"C{first}:C{last}",
        FormulaRule(formula=[f'$C{first}="ALTA"'], fill=PatternFill("solid", fgColor=ROJO)))
    ws.conditional_formatting.add(f"C{first}:C{last}",
        FormulaRule(formula=[f'$C{first}="MEDIA"'], fill=PatternFill("solid", fgColor=AMAR)))


def main():
    wb = Workbook()
    hoja_resumen(wb)
    hoja_calculadora(wb)
    hoja_carrera(wb)
    hoja_prioridades(wb)
    wb.save("Notas-y-Promedios.xlsx")
    print("OK -> Notas-y-Promedios.xlsx")


if __name__ == "__main__":
    main()
