#!/usr/bin/env python3
"""
'Mi-Vida-360.xlsx' - El tablero maestro.
Hojas:
  1) Inicio (Dashboard)   -> metricas clave, se actualizan solas
  2) Ingresos             -> registro mensual de ingresos
  3) Egresos              -> registro mensual de gastos
  4) Deudas               -> deudas, cuotas, meses restantes, estrategia
  5) Presupuesto 50/30/20 -> reparto automatico segun ingresos
  6) Pasantia - Plan      -> pasos accionables con estado
  7) Pasantia - Aplicaciones -> rastreador de postulaciones
  8) Libros               -> biblioteca recomendada + avance de lectura
Moneda: COP.  Todo con formulas.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

AZUL="1A237E"; GRIS="37474F"; CLARO="ECEFF1"; ACENTO="3949AB"
ROJO="FFCDD2"; AMAR="FFF9C4"; VERDE="C8E6C9"; ROSA="F8BBD0"
COP = '"$"#,##0'
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def titulo(ws, a, texto, span, color=AZUL, size=14):
    ws.merge_cells(f"{a}:{span}")
    c = ws[a]; c.value = texto
    c.font = Font(bold=True, size=size, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color); c.alignment = CENTER

def head(ws, row, col, txt, w=None, color=GRIS):
    c = ws.cell(row=row, column=col, value=txt)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=color)
    c.alignment = CENTER; c.border = BORDER
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c

def fmt_rows(ws, rows, cols, money_cols=()):
    for r in rows:
        for c in cols:
            cell = ws.cell(row=r, column=c); cell.border = BORDER
            cell.alignment = CENTER if c not in (1,) else LEFT
            if c in money_cols: cell.number_format = COP


# ---------------------------------------------------------------- 2) INGRESOS
def hoja_ingresos(wb):
    ws = wb.create_sheet("Ingresos")
    titulo(ws, "A1", "INGRESOS DEL MES", "E1", color="1B5E20")
    for j,(t,w) in enumerate([("Fecha",12),("Concepto",26),("Categoria",18),("Monto",16),("Recurrente",12)],1):
        head(ws, 2, j, t, w)
    ej = [("", "Mesada / apoyo familiar", "Familia", 0, "Si"),
          ("", "Trabajo / freelance", "Trabajo", 0, "No"),
          ("", "Otro", "Otro", 0, "No")]
    r0 = 3
    for i,(f,c,cat,m,rec) in enumerate(ej):
        r = r0+i
        ws.cell(row=r,column=1,value=f); ws.cell(row=r,column=2,value=c)
        ws.cell(row=r,column=3,value=cat); ws.cell(row=r,column=4,value=m)
        ws.cell(row=r,column=5,value=rec)
    rN = r0+19  # espacio para 20 filas
    fmt_rows(ws, range(r0, rN+1), range(1,6), money_cols=(4,))
    tr = rN+1
    ws.cell(row=tr,column=3,value="TOTAL INGRESOS").font=Font(bold=True)
    t=ws.cell(row=tr,column=4,value=f"=SUM(D{r0}:D{rN})"); t.font=Font(bold=True,size=12,color="1B5E20")
    t.number_format=COP; t.fill=PatternFill("solid",fgColor=CLARO); t.border=BORDER
    return f"Ingresos!D{tr}"

# ---------------------------------------------------------------- 3) EGRESOS
def hoja_egresos(wb):
    ws = wb.create_sheet("Egresos")
    titulo(ws, "A1", "EGRESOS DEL MES", "E1", color="B71C1C")
    for j,(t,w) in enumerate([("Fecha",12),("Concepto",26),("Categoria",18),("Monto",16),("Necesario?",12)],1):
        head(ws, 2, j, t, w)
    ej = [("","Transporte (TransMilenio/SITP)","Transporte",0,"Si"),
          ("","Alimentacion","Comida",0,"Si"),
          ("","Gimnasio + Box","Salud",0,"Si"),
          ("","Clases de ingles","Educacion",0,"Si"),
          ("","Universidad / materiales","Educacion",0,"Si"),
          ("","Salidas con novia","Ocio",0,"No"),
          ("","Suscripciones","Ocio",0,"No")]
    r0=3
    for i,(f,c,cat,m,n) in enumerate(ej):
        r=r0+i
        ws.cell(row=r,column=1,value=f);ws.cell(row=r,column=2,value=c)
        ws.cell(row=r,column=3,value=cat);ws.cell(row=r,column=4,value=m);ws.cell(row=r,column=5,value=n)
    rN=r0+24
    fmt_rows(ws, range(r0,rN+1), range(1,6), money_cols=(4,))
    tr=rN+1
    ws.cell(row=tr,column=3,value="TOTAL EGRESOS").font=Font(bold=True)
    t=ws.cell(row=tr,column=4,value=f"=SUM(D{r0}:D{rN})");t.font=Font(bold=True,size=12,color="B71C1C")
    t.number_format=COP;t.fill=PatternFill("solid",fgColor=CLARO);t.border=BORDER
    # gasto "no necesario"
    ws.cell(row=tr+1,column=3,value="Gasto evitable (No)").font=Font(italic=True,color="B71C1C")
    g=ws.cell(row=tr+1,column=4,value=f'=SUMIF(E{r0}:E{rN},"No",D{r0}:D{rN})')
    g.number_format=COP
    return f"Egresos!D{tr}"

# ---------------------------------------------------------------- 4) DEUDAS
def hoja_deudas(wb):
    ws = wb.create_sheet("Deudas")
    titulo(ws,"A1","DEUDAS Y PLAN DE PAGO","G1",color="E65100")
    cols=[("Acreedor",22),("Monto total",15),("Saldo actual",15),("Tasa % mes",11),
          ("Cuota mensual",14),("Meses restantes",14),("Prioridad",12)]
    for j,(t,w) in enumerate(cols,1): head(ws,2,j,t,w)
    r0=3; rN=r0+11
    for r in range(r0,rN+1):
        ws.cell(row=r,column=6,value=f'=IFERROR(IF(E{r}=0,"-",ROUNDUP(C{r}/E{r},0)),"-")')
        ws.cell(row=r,column=7,value=f'=IF(C{r}="","-",IF(D{r}>=3,"ALTA (tasa alta)",IF(C{r}>0,"Normal","Pagada")))')
    fmt_rows(ws, range(r0,rN+1), range(1,8), money_cols=(2,3,5))
    tr=rN+1
    ws.cell(row=tr,column=1,value="DEUDA TOTAL").font=Font(bold=True)
    t=ws.cell(row=tr,column=3,value=f"=SUM(C{r0}:C{rN})");t.font=Font(bold=True,size=12,color="E65100")
    t.number_format=COP;t.fill=PatternFill("solid",fgColor=CLARO);t.border=BORDER
    # estrategia
    ws.cell(row=tr+2,column=1,value="Estrategia: paga primero la deuda de MAYOR tasa (metodo avalancha). Si necesitas motivacion, paga primero la mas pequena (bola de nieve).").font=Font(italic=True,size=9)
    ws.conditional_formatting.add(f"G{r0}:G{rN}",
        FormulaRule(formula=[f'LEFT($G{r0},4)="ALTA"'],fill=PatternFill("solid",fgColor=ROJO)))
    return f"Deudas!C{tr}"

# ---------------------------------------------------------------- 5) PRESUPUESTO
def hoja_presupuesto(wb, ingresos_ref, egresos_ref):
    ws = wb.create_sheet("Presupuesto 50-30-20")
    titulo(ws,"A1","PRESUPUESTO 50 / 30 / 20","D1",color=ACENTO)
    ws.column_dimensions["A"].width=30;ws.column_dimensions["B"].width=18;ws.column_dimensions["C"].width=18;ws.column_dimensions["D"].width=22
    ws.cell(row=3,column=1,value="Ingreso mensual total").font=Font(bold=True)
    c=ws.cell(row=3,column=2,value=f"={ingresos_ref}");c.number_format=COP
    rows=[("Necesidades (50%)",0.5,"Arriendo, comida, transporte, U"),
          ("Gustos / ocio (30%)",0.3,"Salidas, suscripciones, hobbies"),
          ("Ahorro + deudas (20%)",0.2,"Fondo emergencia, abonos a deuda")]
    for j,t in enumerate(["Categoria","Ideal","Sugerido (de tu ingreso)","Para que"],1):
        head(ws,5,j,t)
    for i,(cat,pct,desc) in enumerate(rows):
        r=6+i
        ws.cell(row=r,column=1,value=cat)
        ws.cell(row=r,column=2,value=pct).number_format="0%"
        m=ws.cell(row=r,column=3,value=f"=$B$3*B{r}");m.number_format=COP
        ws.cell(row=r,column=4,value=desc).alignment=LEFT
        fmt_rows(ws,[r],range(1,5))
    ws.cell(row=10,column=1,value="Gasto real del mes").font=Font(bold=True)
    g=ws.cell(row=10,column=3,value=f"={egresos_ref}");g.number_format=COP
    ws.cell(row=11,column=1,value="BALANCE (ingreso - gasto)").font=Font(bold=True)
    b=ws.cell(row=11,column=3,value=f"={ingresos_ref}-{egresos_ref}");b.number_format=COP;b.font=Font(bold=True,size=12)
    ws.conditional_formatting.add("C11",CellIsRule(operator="lessThan",formula=["0"],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add("C11",CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=PatternFill("solid",fgColor=VERDE)))

# ---------------------------------------------------------------- 6) PASANTIA PLAN
def hoja_pasantia_plan(wb):
    ws=wb.create_sheet("Pasantia - Plan")
    titulo(ws,"A1","PLAN DE PASANTIA - PASO A PASO","D1",color="004D40")
    for j,(t,w) in enumerate([("Paso",6),("Que hacer",60),("Estado",14),("Fecha meta",14)],1):
        head(ws,2,j,t,w)
    pasos=[
        "Definir objetivo: area (estructuras, vias, hidraulica, construccion, interventoria) y ciudad.",
        "Actualizar hoja de vida a 1 pagina, enfocada en logros y el proyecto estructural.",
        "Optimizar LinkedIn: foto profesional, titular claro, activar 'Open to work'.",
        "Armar lista de 25-30 empresas objetivo (constructoras, consultoras, interventoria).",
        "Conseguir contactos/referidos: profes, egresados, ferias, oficina de practicas de la U.",
        "Crear carta de presentacion base (personalizable por empresa).",
        "Preparar portafolio con tus proyectos academicos (incluye el de estructural).",
        "Aplicar minimo 5 vacantes por semana (registrar en hoja 'Aplicaciones').",
        "Revisar portales: LinkedIn, elempleo, Computrabajo, Magneto365, bolsa de la U.",
        "Preparar entrevista: preguntas tecnicas + comportamentales (metodo STAR).",
        "Hacer seguimiento a cada aplicacion a los 7 dias.",
        "Al recibir oferta: revisar convenio U, ARL y afiliacion antes de firmar.",
    ]
    r0=3
    for i,p in enumerate(pasos):
        r=r0+i
        ws.cell(row=r,column=1,value=i+1)
        ws.cell(row=r,column=2,value=p).alignment=LEFT
        ws.cell(row=r,column=3,value="Pendiente")
        fmt_rows(ws,[r],range(1,5))
    rN=r0+len(pasos)-1
    dv=DataValidation(type="list",formula1='"Pendiente,En curso,Hecho"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"C{r0}:C{rN}")
    ws.conditional_formatting.add(f"C{r0}:C{rN}",CellIsRule(operator="equal",formula=['"Hecho"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"C{r0}:C{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))
    # progreso
    ws.cell(row=rN+2,column=2,value="Progreso del plan").font=Font(bold=True)
    p=ws.cell(row=rN+2,column=3,value=f'=ROUND(COUNTIF(C{r0}:C{rN},"Hecho")/{len(pasos)},2)')
    p.number_format="0%";p.font=Font(bold=True)
    return f"'Pasantia - Plan'!C{rN+2}"

# ---------------------------------------------------------------- 7) APLICACIONES
def hoja_aplicaciones(wb):
    ws=wb.create_sheet("Pasantia - Aplicaciones")
    titulo(ws,"A1","RASTREADOR DE APLICACIONES","G1",color="00695C")
    for j,(t,w) in enumerate([("Empresa",24),("Cargo",22),("Fecha",12),("Estado",16),
                               ("Contacto",20),("Proximo paso",24),("Notas",24)],1):
        head(ws,2,j,t,w)
    r0=3; rN=r0+39
    fmt_rows(ws, range(r0,rN+1), range(1,8))
    dv=DataValidation(type="list",formula1='"Por aplicar,Aplicada,En proceso,Entrevista,Oferta,Rechazada"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"D{r0}:D{rN}")
    # contadores
    ws.cell(row=2,column=9,value="Aplicadas").font=Font(bold=True)
    ws.cell(row=3,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Aplicada")+COUNTIF(D{r0}:D{rN},"En proceso")+COUNTIF(D{r0}:D{rN},"Entrevista")+COUNTIF(D{r0}:D{rN},"Oferta")')
    ws.cell(row=4,column=9,value="Entrevistas").font=Font(bold=True)
    ws.cell(row=5,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Entrevista")')
    ws.cell(row=6,column=9,value="Ofertas").font=Font(bold=True)
    ws.cell(row=7,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Oferta")')
    ws.column_dimensions["I"].width=12
    colmap={"Oferta":VERDE,"Entrevista":"B3E5FC","Rechazada":ROJO,"Aplicada":AMAR}
    for val,color in colmap.items():
        ws.conditional_formatting.add(f"D{r0}:D{rN}",CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",fgColor=color)))
    return f"'Pasantia - Aplicaciones'!I3"

# ---------------------------------------------------------------- 8) LIBROS
def hoja_libros(wb):
    ws=wb.create_sheet("Libros")
    titulo(ws,"A1","BIBLIOTECA RECOMENDADA PARA TI","F1",color="4A148C")
    for j,(t,w) in enumerate([("Titulo",30),("Autor",20),("Categoria",18),("Estado",14),("% avance",10),("Para que te sirve",34)],1):
        head(ws,2,j,t,w)
    libros=[
        ("Habitos Atomicos","James Clear","Habitos","Por leer","Construir disciplina con pequenos sistemas (empieza por este)"),
        ("Enfocate (Deep Work)","Cal Newport","Productividad","Por leer","Concentracion profunda para proyecto e ingles"),
        ("Los 7 habitos de la gente altamente efectiva","Stephen Covey","Desarrollo","Por leer","Priorizar y vivir con proposito"),
        ("Padre Rico, Padre Pobre","R. Kiyosaki","Finanzas","Por leer","Educacion financiera basica"),
        ("El hombre mas rico de Babilonia","George Clason","Finanzas","Por leer","Ahorro e inversion en parabolas simples"),
        ("Como ganar amigos e influir sobre las personas","Dale Carnegie","Relaciones","Por leer","Comunicacion (util en pareja y entrevistas)"),
        ("Mindset (La actitud del exito)","Carol Dweck","Mentalidad","Por leer","Mentalidad de crecimiento"),
        ("El poder del ahora","Eckhart Tolle","Bienestar","Por leer","Manejar ansiedad y vivir el presente"),
        ("Inteligencia emocional","Daniel Goleman","Bienestar","Por leer","Gestionar emociones y estres academico"),
        ("Despertando al gigante interior","Tony Robbins","Motivacion","Por leer","Motivacion y toma de decisiones"),
    ]
    r0=3
    for i,(t,a,cat,est,para) in enumerate(libros):
        r=r0+i
        ws.cell(row=r,column=1,value=t);ws.cell(row=r,column=2,value=a)
        ws.cell(row=r,column=3,value=cat);ws.cell(row=r,column=4,value=est)
        ws.cell(row=r,column=5,value=0).number_format="0%"
        ws.cell(row=r,column=6,value=para).alignment=LEFT
        fmt_rows(ws,[r],range(1,7))
    rN=r0+len(libros)-1
    dv=DataValidation(type="list",formula1='"Por leer,Leyendo,Leido"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"D{r0}:D{rN}")
    ws.conditional_formatting.add(f"D{r0}:D{rN}",CellIsRule(operator="equal",formula=['"Leido"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"D{r0}:D{rN}",CellIsRule(operator="equal",formula=['"Leyendo"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.cell(row=rN+2,column=1,value="Leidos / Total").font=Font(bold=True)
    ws.cell(row=rN+2,column=4,value=f'=COUNTIF(D{r0}:D{rN},"Leido")&" / "&{len(libros)}').font=Font(bold=True)
    return f'Libros!D{rN+2}', len(libros)

# ---------------------------------------------------------------- 1) DASHBOARD
def hoja_inicio(wb, refs):
    ws=wb.create_sheet("Inicio", 0)
    titulo(ws,"A1","MI VIDA 360  -  TABLERO MAESTRO","F1",size=16)
    ws.merge_cells("A2:F2")
    ws["A2"].value="Llena las hojas y este tablero se actualiza solo. Tu futuro se construye con datos + accion."
    ws["A2"].font=Font(italic=True,color="455A64"); ws["A2"].alignment=CENTER
    tarjetas=[
        ("Ingresos del mes", f"={refs['ing']}", COP, "1B5E20"),
        ("Egresos del mes", f"={refs['egr']}", COP, "B71C1C"),
        ("Balance del mes", f"={refs['ing']}-{refs['egr']}", COP, "0D47A1"),
        ("Deuda total", f"={refs['deu']}", COP, "E65100"),
        ("Avance plan pasantia", f"={refs['pas']}", "0%", "004D40"),
        ("Aplicaciones enviadas", f"={refs['apl']}", "0", "00695C"),
        ("Libros leidos", f"={refs['lib']}", "General", "4A148C"),
    ]
    r=4
    for i,(t,f,fmt,color) in enumerate(tarjetas):
        rr=4+i
        lab=ws.cell(row=rr,column=1,value=t);lab.font=Font(bold=True,size=11)
        lab.fill=PatternFill("solid",fgColor=CLARO);lab.border=BORDER;lab.alignment=LEFT
        val=ws.cell(row=rr,column=2,value=f);val.font=Font(bold=True,size=12,color=color)
        val.border=BORDER;val.alignment=CENTER
        if fmt not in ("General",): val.number_format=fmt
        ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=3)
    ws.column_dimensions["A"].width=26
    ws.column_dimensions["B"].width=16;ws.column_dimensions["C"].width=8
    ws.conditional_formatting.add("B6",CellIsRule(operator="lessThan",formula=["0"],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add("B6",CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=PatternFill("solid",fgColor=VERDE)))
    # mini guia
    ws.cell(row=13,column=1,value="COMO USARLO").font=Font(bold=True,size=12,color=AZUL)
    guia=["1. Registra ingresos y egresos cada que ocurran.",
          "2. Carga tus deudas y sigue el metodo avalancha.",
          "3. Trabaja el plan de pasantia y registra cada aplicacion.",
          "4. Marca tus libros a medida que avanzas.",
          "5. Revisa este tablero cada domingo en tu planeacion semanal."]
    for i,g in enumerate(guia):
        ws.cell(row=14+i,column=1,value=g).alignment=LEFT
        ws.merge_cells(start_row=14+i,start_column=1,end_row=14+i,end_column=6)


def main():
    wb=Workbook(); wb.remove(wb.active)
    ing=hoja_ingresos(wb)
    egr=hoja_egresos(wb)
    deu=hoja_deudas(wb)
    hoja_presupuesto(wb, ing, egr)
    pas=hoja_pasantia_plan(wb)
    apl=hoja_aplicaciones(wb)
    lib_ref,_=hoja_libros(wb)
    hoja_inicio(wb, {"ing":ing,"egr":egr,"deu":deu,"pas":pas,"apl":apl,"lib":lib_ref})
    wb.save("Mi-Vida-360.xlsx")
    print("OK -> Mi-Vida-360.xlsx")

if __name__=="__main__":
    main()
