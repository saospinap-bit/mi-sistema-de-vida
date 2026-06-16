#!/usr/bin/env python3
"""
'Becas-Maestria-Cursos.xlsx'
Hojas:
  1) Becas            -> programas reales, cobertura, requisitos, link, tracker
  2) Destinos Maestria-> paises por costo (datos 2026), pros/contras
  3) Cursos CV        -> cursos online para fortalecer la hoja de vida + avance
  4) Ruta Posgrado    -> linea de tiempo de pasos para aplicar
Fuentes citadas en celdas. Info reformulada por cumplimiento de licencias.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

AZUL="0D47A1"; GRIS="37474F"; CLARO="ECEFF1"; VERDE="C8E6C9"; AMAR="FFF9C4"; ROJO="FFCDD2"
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def titulo(ws,a,t,span,color=AZUL):
    ws.merge_cells(f"{a}:{span}"); c=ws[a]; c.value=t
    c.font=Font(bold=True,size=14,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER
def head(ws,row,col,t,w=None,color=GRIS):
    c=ws.cell(row=row,column=col,value=t); c.font=Font(bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER; c.border=BORDER
    if w: ws.column_dimensions[get_column_letter(col)].width=w

def rellenar(ws, r0, datos, ncols, left_cols=()):
    for i,fila in enumerate(datos):
        r=r0+i
        for j,val in enumerate(fila,1):
            c=ws.cell(row=r,column=j,value=val); c.border=BORDER
            c.alignment=LEFT if j in left_cols else CENTER
    return r0+len(datos)-1

# ----------------------------------------------------------- 1) BECAS
def hoja_becas(wb):
    ws=wb.active; ws.title="Becas"
    titulo(ws,"A1","BECAS PARA TU MAESTRIA (revisar fechas cada ano)","G1")
    cols=[("Beca/Programa",24),("Pais/Region",16),("Que cubre",30),
          ("Requisito clave",26),("Link",30),("Estado",14)]
    for j,(t,w) in enumerate(cols,1): head(ws,2,j,t,w)
    datos=[
        ("COLFUTURO","Colombia->mundo","Credito-beca hasta US$25k/ano (max 2 anos)","Admision + excelencia academica","colfuturo.org","Por investigar"),
        ("Fulbright Colombia","EE.UU.","Matricula + sostenimiento maestria","Ingles (TOEFL), liderazgo","fulbright.edu.co","Por investigar"),
        ("Chevening","Reino Unido","Totalmente financiada (matricula+vida)","Experiencia laboral + admision","chevening.org","Por investigar"),
        ("DAAD","Alemania","Becas maestria (EPOS para desarrollo)","Buen promedio + propuesta","daad.de","Por investigar"),
        ("Erasmus Mundus","Europa (varios)","Maestrias conjuntas 100% financiadas","Admision al programa conjunto","erasmus-plus.ec.europa.eu","Por investigar"),
        ("Becas de la universidad","Destino elegido","Descuentos/becas parciales","Admision temprana","Web de cada universidad","Por investigar"),
        ("ICETEX","Colombia","Creditos y becas convocatorias","Segun convocatoria","icetex.gov.co","Por investigar"),
    ]
    rN=rellenar(ws,3,datos,6,left_cols=(1,3,4,5))
    dv=DataValidation(type="list",formula1='"Por investigar,Aplicando,Aplicada,Admitido,Descartada"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"F3:F{rN}")
    ws.conditional_formatting.add(f"F3:F{rN}",CellIsRule(operator="equal",formula=['"Admitido"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"F3:F{rN}",CellIsRule(operator="equal",formula=['"Aplicando"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.cell(row=rN+2,column=1,value="Nota: confirma SIEMPRE las fechas de convocatoria en la web oficial; cambian cada ano.").font=Font(italic=True,size=9)

# ----------------------------------------------------------- 2) DESTINOS
def hoja_destinos(wb):
    ws=wb.create_sheet("Destinos Maestria")
    titulo(ws,"A1","MEJORES DESTINOS POR COSTO (datos 2026, confirmar)","F1",color="1B5E20")
    cols=[("Pais",14),("Costo matricula aprox.",24),("Idioma",14),("A favor",26),("En contra",24)]
    for j,(t,w) in enumerate(cols,1): head(ws,2,j,t,w)
    datos=[
        ("Alemania","Gratis (solo ~250-350 EUR/sem)","Aleman/Ingles","Top calidad-costo, industria fuerte","Algunos programas piden aleman"),
        ("Austria","~726 EUR/semestre (no-UE)","Aleman","Muy economica, buena calidad","Cupos y idioma"),
        ("Francia","~2.770-3.770 EUR/ano (no-UE)","Frances/Ingles","Ingenieria prestigiosa","Idioma frances util"),
        ("Italia","~900-4.000 EUR/ano (segun ingreso)","Italiano/Ingles","Becas por renta, costo bajo","Tramites lentos"),
        ("Espana","Medio","Espanol","Sin barrera de idioma","Menos becas full"),
        ("Noruega","YA cobra a no-UE desde 2023","Ingles/Noruego","Alta calidad de vida","Dejo de ser gratis"),
        ("Reino Unido","Alto","Ingles","Maestrias de 1 ano, prestigio","Caro: depende de Chevening/Colfuturo"),
        ("EE.UU.","Alto","Ingles","Investigacion top","Caro: depende de Fulbright"),
        ("Canada","Medio-alto","Ingles/Frances","Buena via a residencia","Costo de vida"),
    ]
    rN=rellenar(ws,3,datos,5,left_cols=(2,4,5))
    ws.cell(row=rN+2,column=1,value="Recomendacion: si el costo manda, Alemania/Austria; si quieres beca full, apunta a Chevening (UK), Fulbright (USA) o Erasmus Mundus.").font=Font(italic=True,size=9)
    ws.cell(row=rN+3,column=1,value="Fuentes: mastersportal.com, thecollegemonk.com, leapscholar.com (reformulado).").font=Font(italic=True,size=8,color="888888")

# ----------------------------------------------------------- 3) CURSOS
def hoja_cursos(wb):
    ws=wb.create_sheet("Cursos CV")
    titulo(ws,"A1","CURSOS ONLINE PARA TU HOJA DE VIDA","G1",color="4A148C")
    cols=[("Curso",34),("Plataforma",14),("Por que te sirve",30),("Prioridad",12),("Estado",14),("% avance",10)]
    for j,(t,w) in enumerate(cols,1): head(ws,2,j,t,w)
    datos=[
        ("Construction Management Specialization (Columbia)","Coursera","Gestion de obra, muy valorado","ALTA","Por empezar",0),
        ("BIM Fundamentals for Engineers","Coursera","BIM: el futuro del diseno","ALTA","Por empezar",0),
        ("Autodesk Revit / AutoCAD","Coursera/Udemy","Software clave en ofertas","ALTA","Por empezar",0),
        ("Engineering Project Management Specialization","Coursera","Liderar proyectos","MEDIA","Por empezar",0),
        ("Google Project Management Certificate","Coursera","Certificado respetado por reclutadores","MEDIA","Por empezar",0),
        ("Preparacion IELTS/TOEFL","Coursera/edX","Ingles certificado = becas","ALTA","Por empezar",0),
        ("Excel + analisis de datos","Coursera","Productividad y datos","MEDIA","Por empezar",0),
        ("Python para ingenieros","Coursera/edX","Automatizar calculos","BAJA","Por empezar",0),
        ("SAP2000 / ETABS (modelado estructural)","Udemy/YouTube","Directo a tu area","ALTA","Por empezar",0),
    ]
    rN=rellenar(ws,3,datos,6,left_cols=(1,3))
    for r in range(3,rN+1): ws.cell(row=r,column=6).number_format="0%"
    dv=DataValidation(type="list",formula1='"Por empezar,En curso,Terminado"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"E3:E{rN}")
    ws.conditional_formatting.add(f"E3:E{rN}",CellIsRule(operator="equal",formula=['"Terminado"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"E3:E{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.conditional_formatting.add(f"D3:D{rN}",CellIsRule(operator="equal",formula=['"ALTA"'],fill=PatternFill("solid",fgColor=ROJO)))
    ws.cell(row=rN+2,column=1,value="Tip: en Coursera puedes pedir 'ayuda economica' (financial aid) para tomar cursos certificados gratis.").font=Font(italic=True,size=9)
    ws.cell(row=rN+3,column=1,value="Cursos terminados:").font=Font(bold=True)
    ws.cell(row=rN+3,column=5,value=f'=COUNTIF(E3:E{rN},"Terminado")&" / "&{len(datos)}').font=Font(bold=True)

# ----------------------------------------------------------- 4) RUTA
def hoja_ruta(wb):
    ws=wb.create_sheet("Ruta Posgrado")
    titulo(ws,"A1","RUTA HACIA LA MAESTRIA (linea de tiempo)","D1",color="00695C")
    for j,(t,w) in enumerate([("Cuando",16),("Que hacer",58),("Estado",14)],1): head(ws,2,j,t,w)
    datos=[
        ("Ahora (9o sem)","Subir el promedio (PAPA) y salvar la materia critica","Pendiente"),
        ("Ahora","Empezar ingles y apuntar a IELTS/TOEFL","Pendiente"),
        ("3-6 meses","Definir pais y 5-8 universidades objetivo","Pendiente"),
        ("3-6 meses","Tomar 1-2 cursos de la hoja 'Cursos CV'","Pendiente"),
        ("6-9 meses","Conseguir cartas de recomendacion de profes","Pendiente"),
        ("6-9 meses","Escribir carta de motivacion (statement of purpose)","Pendiente"),
        ("9-12 meses","Aplicar a becas (Colfuturo, Chevening, DAAD, etc.)","Pendiente"),
        ("9-12 meses","Aplicar a admisiones de universidades","Pendiente"),
        ("12-15 meses","Entrevistas y resultados","Pendiente"),
        ("Despues de grado","Pasantia/experiencia que suma para la beca","Pendiente"),
    ]
    rN=rellenar(ws,3,datos,3,left_cols=(2,))
    dv=DataValidation(type="list",formula1='"Pendiente,En curso,Hecho"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"C3:C{rN}")
    ws.conditional_formatting.add(f"C3:C{rN}",CellIsRule(operator="equal",formula=['"Hecho"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"C3:C{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))

def main():
    wb=Workbook()
    hoja_becas(wb); hoja_destinos(wb); hoja_cursos(wb); hoja_ruta(wb)
    wb.save("Becas-Maestria-Cursos.xlsx")
    print("OK -> Becas-Maestria-Cursos.xlsx")

if __name__=="__main__":
    main()
