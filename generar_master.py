#!/usr/bin/env python3
"""
SISTEMA DE VIDA - EXCEL MAESTRO (todo en un solo archivo, separado por pestanas)
Archivo: Mi-Sistema-de-Vida.xlsx
Pestanas:
  1 Inicio (dashboard maestro)     11 Deudas
  2 Horario                        12 Presupuesto 50-30-20
  3 Habitos                        13 Pasantia - Plan
  4 Resumen Semestre               14 Pasantia - Aplicaciones
  5 Calculadora Notas              15 Becas
  6 Entregas                       16 Destinos Maestria
  7 PAPA Carrera                   17 Cursos CV
  8 Prioridades                    18 Ruta Posgrado
  9 Ingresos                       19 Libros
 10 Egresos
Todo con formulas. Moneda COP. Escala notas 0-5 (aprueba 3.0).
"""
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ---- paleta ----
NAVY="1A237E"; GRIS="37474F"; CLARO="ECEFF1"; ACENTO="3949AB"
ROJO="FFCDD2"; AMAR="FFF9C4"; VERDE="C8E6C9"; ROSA="F8BBD0"; AZULC="B3E5FC"
COP='"$"#,##0'; APRUEBA=3.0
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def titulo(ws,a,t,span,color=NAVY,size=14):
    ws.merge_cells(f"{a}:{span}"); c=ws[a]; c.value=t
    c.font=Font(bold=True,size=size,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER
def head(ws,row,col,t,w=None,color=GRIS):
    c=ws.cell(row=row,column=col,value=t); c.font=Font(bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=CENTER; c.border=BORDER
    if w: ws.column_dimensions[get_column_letter(col)].width=w
def fill_table(ws,r0,datos,left_cols=(),date_cols=(),pct_cols=()):
    for i,fila in enumerate(datos):
        r=r0+i
        for j,val in enumerate(fila,1):
            c=ws.cell(row=r,column=j,value=val); c.border=BORDER
            c.alignment=LEFT if j in left_cols else CENTER
            if j in date_cols: c.number_format="yyyy-mm-dd"
            if j in pct_cols: c.number_format="0%"
    return r0+len(datos)-1

# ============================================================ HORARIO
DIAS=["Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"]
CAT_COLOR={"ingles":"B3E5FC","clase":"FFCDD2","estructural":"F44336","proyecto":"FFE0B2",
           "novia":"F8BBD0","gym":"C8E6C9","box":"A5D6A7","psicologia":"E1BEE7",
           "lectura":"FFF9C4","viaje":"ECEFF1","descanso":"F5F5F5"}
EVENTOS=[
    ("Lunes","06:00","07:30","Ingles (virtual)","ingles"),("Lunes","08:00","09:30","Gimnasio (fuerza)","gym"),
    ("Lunes","10:00","12:30","PROYECTO estructural","proyecto"),("Lunes","13:00","13:30","Lectura 30 min","lectura"),
    ("Lunes","14:00","16:30","PROYECTO / estudio","proyecto"),("Lunes","16:30","17:30","Psicologia","psicologia"),
    ("Lunes","18:00","20:00","Clase opcional","clase"),
    ("Martes","06:00","07:30","Ingles (virtual)","ingles"),("Martes","08:00","09:00","Viaje a la U","viaje"),
    ("Martes","09:00","11:00","Clase OBLIGATORIA","clase"),("Martes","11:00","12:30","Estudio / pasantia","proyecto"),
    ("Martes","12:30","13:30","Almuerzo con novia","novia"),("Martes","13:30","16:00","PROYECTO estructural","proyecto"),
    ("Martes","17:30","18:00","Recoger a novia","novia"),("Martes","19:30","21:00","Box","box"),
    ("Miercoles","06:00","07:30","Ingles (virtual)","ingles"),("Miercoles","08:00","09:00","Viaje a la U","viaje"),
    ("Miercoles","09:00","11:00","Clase OBLIGATORIA","clase"),("Miercoles","11:00","12:30","Repaso / tareas","proyecto"),
    ("Miercoles","12:30","13:30","Almuerzo con novia","novia"),("Miercoles","14:00","16:00","ESTRUCTURAL - NO FALTAR","estructural"),
    ("Miercoles","16:00","17:30","Estudio / dudas profe","proyecto"),("Miercoles","17:30","18:00","Recoger a novia","novia"),
    ("Jueves","06:00","07:30","Ingles (virtual)","ingles"),("Jueves","08:00","09:00","Viaje a la U","viaje"),
    ("Jueves","09:00","11:00","Clase OBLIGATORIA","clase"),("Jueves","11:00","12:30","Estudio / pasantia","proyecto"),
    ("Jueves","12:30","13:30","Almuerzo con novia","novia"),("Jueves","13:30","16:00","PROYECTO estructural","proyecto"),
    ("Jueves","17:30","18:00","Recoger a novia","novia"),("Jueves","19:30","21:00","Box","box"),
    ("Viernes","06:00","07:30","Ingles (virtual)","ingles"),("Viernes","08:00","09:00","Viaje a la U","viaje"),
    ("Viernes","09:00","11:00","Clase OBLIGATORIA","clase"),("Viernes","12:30","13:30","Almuerzo con novia","novia"),
    ("Viernes","14:00","16:00","ESTRUCTURAL - NO FALTAR","estructural"),("Viernes","17:30","18:00","Recoger a novia","novia"),
    ("Viernes","19:00","21:00","Tiempo de calidad con novia","novia"),
    ("Sabado","08:00","09:30","Gimnasio (fuerza)","gym"),("Sabado","10:00","12:30","PROYECTO (bloque largo)","proyecto"),
    ("Sabado","14:00","18:00","Tiempo con novia / social","novia"),("Sabado","20:00","20:30","Lectura 30 min","lectura"),
    ("Domingo","10:00","11:00","Descanso / recargar","descanso"),("Domingo","17:00","17:45","Revision y planeacion semanal","psicologia"),
]
def _franjas():
    out=[]; t=datetime.strptime("06:00","%H:%M"); fin=datetime.strptime("21:00","%H:%M")
    while t<fin: out.append(t.strftime("%H:%M")); t+=timedelta(minutes=30)
    return out
FRANJAS=_franjas()
def _hm(s): return datetime.strptime(s,"%H:%M")
def _rejilla():
    ini={}; occ={}
    for dia,a,b,tit,cat in EVENTOS:
        t0,t1=_hm(a),_hm(b); n=int((t1-t0).total_seconds()//1800)
        ini[(dia,a)]=(tit,cat,max(n,1)); tt=t0
        while tt<t1: occ[(dia,tt.strftime("%H:%M"))]=cat; tt+=timedelta(minutes=30)
    return ini,occ

def hoja_horario(wb):
    ws=wb.create_sheet("Horario")
    titulo(ws,"A1","MI HORARIO SEMANAL","H1",color=GRIS)
    head(ws,2,1,"Hora",8)
    for j,d in enumerate(DIAS,2): head(ws,2,j,d,18,color="546E7A")
    ini,occ=_rejilla()
    for i,fr in enumerate(FRANJAS):
        r=i+3
        hc=ws.cell(row=r,column=1,value=fr); hc.font=Font(bold=True,size=9); hc.alignment=CENTER
        hc.fill=PatternFill("solid",fgColor=CLARO); hc.border=BORDER
        ws.row_dimensions[r].height=20
        for j,dia in enumerate(DIAS,2):
            c=ws.cell(row=r,column=j); c.border=BORDER; c.alignment=CENTER
            if (dia,fr) in ini:
                tit,cat,n=ini[(dia,fr)]; c.value=tit
                c.fill=PatternFill("solid",fgColor=CAT_COLOR.get(cat,"FFFFFF"))
                c.font=Font(size=8,bold=cat in("estructural","clase"))
                if n>1: ws.merge_cells(start_row=r,start_column=j,end_row=r+n-1,end_column=j)
            elif (dia,fr) in occ:
                c.fill=PatternFill("solid",fgColor=CAT_COLOR.get(occ[(dia,fr)],"FFFFFF"))
    ws.freeze_panes="B3"

def hoja_habitos(wb):
    ws=wb.create_sheet("Habitos")
    titulo(ws,"A1","TRACKER SEMANAL DE HABITOS","H1",color="2E7D32")
    cols=["Habito","L","M","X","J","V","S","D"]
    for j,t in enumerate(cols,1): head(ws,2,j,t,(24 if j==1 else 5))
    habitos=["Dormir 7-8h","Ingles","Proyecto estructural","Gym / Box","Lectura 30 min","Tiempo c/ novia","Agua + comer bien"]
    for i,h in enumerate(habitos):
        r=3+i; ws.cell(row=r,column=1,value=h).border=BORDER; ws.cell(row=r,column=1).alignment=LEFT
        for j in range(2,9):
            c=ws.cell(row=r,column=j); c.border=BORDER; c.alignment=CENTER

# ============================================================ NOTAS
# (Materia, creditos, [(evaluacion, % peso), ...])  -- creditos editables
MATERIAS_DATA=[
    ("Diseno Estructural",3,[("Parcial 1",20),("Parcial 2",20),("Parcial 3",20),("Primera entrega proyecto",10),("Entrega final proyecto",30)]),
    ("Uitoto",3,[("Mito",30),("Talleres grupales",30),("Propuesta final",40)]),
    ("PEPI",3,[("Parcial 1",20),("Talleres",40),("Quices",10),("Proyecto",30)]),
    ("Saneamiento Ambiental",3,[("Proyecto",40),("Talleres",35),("Quices",15),("Participacion",10)]),
    ("Alcantarillados",3,[("Informe de salida",20),("Planes de gobierno",10),("Parcial 1",20),("Parcial 2",20),("Proyecto",30)]),
    ("Fundamentos de Construccion",3,[("Parcial 1",25),("Parcial 2",25),("Proyecto 1",20),("Talleres",30)]),
]
MATERIAS=[(n,c) for n,c,_ in MATERIAS_DATA]
N_EVAL=7; BLOCK=N_EVAL+9
def _bs(i): return 3+i*BLOCK
def _def(i): return f"D{_bs(i)+N_EVAL+4}"
def _falta(i): return f"D{_bs(i)+N_EVAL+6}"

def hoja_resumen(wb):
    ws=wb.create_sheet("Resumen Semestre")
    titulo(ws,"A1","RESUMEN DEL SEMESTRE (9o) - Aprueba 3.0","G1",color=NAVY)
    for j,(t,w) in enumerate([("Materia",32),("Creditos",9),("Nota Definitiva",13),("Aporte",12),("Estado",13),("Prioridad",12),("Falta %",9)],1):
        head(ws,3,j,t,w)
    first=4
    for i,(nom,cr) in enumerate(MATERIAS):
        r=first+i
        ws.cell(row=r,column=1,value=nom).alignment=LEFT
        ws.cell(row=r,column=2,value=cr).alignment=CENTER
        ws.cell(row=r,column=3,value=f"='Calculadora Notas'!{_def(i)}").alignment=CENTER
        ws.cell(row=r,column=4,value=f"=B{r}*C{r}").alignment=CENTER
        ws.cell(row=r,column=5,value=f'=IF(C{r}="","-",IF(C{r}>={APRUEBA},"Aprobada","EN RIESGO"))').alignment=CENTER
        ws.cell(row=r,column=6,value=f'=IF(C{r}="","-",IF(C{r}<{APRUEBA},"ALTA",IF(C{r}<3.5,"MEDIA","OK")))').alignment=CENTER
        ws.cell(row=r,column=7,value=f"='Calculadora Notas'!{_falta(i)}").alignment=CENTER
        for j in range(1,8): ws.cell(row=r,column=j).border=BORDER
    last=first+len(MATERIAS)-1; tr=last+1
    ws.cell(row=tr,column=1,value="PAPA SEMESTRE").font=Font(bold=True)
    ws.cell(row=tr,column=2,value=f"=SUM(B{first}:B{last})").alignment=CENTER
    papa=ws.cell(row=tr,column=3,value=f"=IFERROR(ROUND(SUMPRODUCT(B{first}:B{last},C{first}:C{last})/SUM(B{first}:B{last}),2),0)")
    papa.font=Font(bold=True,size=12,color="1B5E20"); papa.alignment=CENTER
    for j in range(1,4): ws.cell(row=tr,column=j).fill=PatternFill("solid",fgColor=CLARO); ws.cell(row=tr,column=j).border=BORDER
    ws.cell(row=tr+2,column=1,value="Materias EN RIESGO:").font=Font(bold=True,color="B71C1C")
    ws.cell(row=tr+2,column=2,value=f'=COUNTIF(E{first}:E{last},"EN RIESGO")').font=Font(bold=True,color="B71C1C")
    rng=f"C{first}:C{last}"
    ws.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=[str(APRUEBA)],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="between",formula=[str(APRUEBA),"3.49"],fill=PatternFill("solid",fgColor=AMAR)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="greaterThanOrEqual",formula=["3.5"],fill=PatternFill("solid",fgColor=VERDE)))
    return {"papa":"'Resumen Semestre'!C"+str(tr),"riesgo":"'Resumen Semestre'!B"+str(tr+2)}

def hoja_calculadora(wb):
    ws=wb.create_sheet("Calculadora Notas")
    titulo(ws,"A1","CALCULADORA POR MATERIA (pon % y nota 0-5)","D1",color=NAVY)
    for col,w in [("A",30),("B",12),("C",12),("D",14)]: ws.column_dimensions[col].width=w
    for i,(nom,_) in enumerate(MATERIAS):
        s=_bs(i)
        ws.merge_cells(f"A{s}:D{s}"); c=ws[f"A{s}"]; c.value=f"MATERIA: {nom}"
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=GRIS); c.alignment=LEFT
        for j,t in enumerate(["Evaluacion","% (peso)","Nota (0-5)","Aporte"],1):
            hc=ws.cell(row=s+1,column=j,value=t); hc.font=Font(bold=True); hc.fill=PatternFill("solid",fgColor=CLARO)
            hc.alignment=CENTER; hc.border=BORDER
        ev0=s+2; evals=MATERIAS_DATA[i][2]
        for k in range(N_EVAL):
            r=ev0+k
            if k<len(evals):
                ws.cell(row=r,column=1,value=evals[k][0]).alignment=LEFT
                ws.cell(row=r,column=2,value=evals[k][1]).alignment=CENTER
            else:
                ws.cell(row=r,column=1,value=f"Eval {k+1}").alignment=LEFT
            ws.cell(row=r,column=4,value=f"=IFERROR(B{r}*C{r}/100,0)").alignment=CENTER
            for j in range(1,5): ws.cell(row=r,column=j).border=BORDER
        evN=ev0+N_EVAL-1
        rs=evN+1; ws.cell(row=rs,column=1,value="Suma % (debe dar 100)").font=Font(bold=True)
        ws.cell(row=rs,column=2,value=f"=SUM(B{ev0}:B{evN})").alignment=CENTER
        ws.cell(row=rs,column=3,value=f'=IF(B{rs}=100,"OK","REVISAR")').alignment=CENTER
        ra=evN+2; ws.cell(row=ra,column=1,value="Nota acumulada").font=Font(bold=True)
        ws.cell(row=ra,column=4,value=f"=ROUND(SUM(D{ev0}:D{evN}),2)").alignment=CENTER
        rd=evN+3; ws.cell(row=rd,column=1,value="NOTA DEFINITIVA").font=Font(bold=True,color="1B5E20")
        dd=ws.cell(row=rd,column=4,value=f"=ROUND(SUM(D{ev0}:D{evN}),2)"); dd.font=Font(bold=True,color="1B5E20"); dd.alignment=CENTER
        rc=evN+4; ws.cell(row=rc,column=1,value="% ya cursado").alignment=LEFT
        ws.cell(row=rc,column=4,value=f'=SUMIF(C{ev0}:C{evN},">0",B{ev0}:B{evN})').alignment=CENTER
        rf=evN+5; ws.cell(row=rf,column=1,value="% que falta").font=Font(bold=True)
        ws.cell(row=rf,column=4,value=f"=100-D{rc}").alignment=CENTER
        rn=evN+6; ws.cell(row=rn,column=1,value=f"Nota que necesito (en lo que falta) para pasar {APRUEBA}").font=Font(bold=True,color="B71C1C")
        nn=ws.cell(row=rn,column=4,value=f'=IFERROR(IF(D{rf}=0,"-",ROUND(({APRUEBA}-D{ra})/(D{rf}/100),2)),"-")'); nn.font=Font(bold=True,color="B71C1C"); nn.alignment=CENTER
        for rr in (rs,ra,rd,rc,rf,rn):
            for j in range(1,5): ws.cell(row=rr,column=j).border=BORDER

def hoja_entregas(wb):
    ws=wb.create_sheet("Entregas")
    titulo(ws,"A1","ENTREGAS - que debo entregar (semana / mes)","H1",color="C62828")
    for j,(t,w) in enumerate([("Materia",24),("Entregable",26),("Tipo",14),("% nota",9),
                               ("Fecha",13),("Dias restantes",13),("Estado",14),("Alerta",16)],1):
        head(ws,2,j,t,w)
    r0=3; rN=r0+24
    ejemplos=[
        ("Diseno Estructural","Parcial 1","Parcial",20,"","Pendiente"),
        ("Diseno Estructural","Parcial 2","Parcial",20,"","Pendiente"),
        ("Diseno Estructural","Parcial 3","Parcial",20,"","Pendiente"),
        ("Diseno Estructural","Primera entrega proyecto","Proyecto",10,"","Pendiente"),
        ("Diseno Estructural","Entrega final proyecto","Proyecto",30,"","Pendiente"),
    ]
    for i,(mat,ent,tip,pc,fec,est) in enumerate(ejemplos):
        r=r0+i
        ws.cell(row=r,column=1,value=mat); ws.cell(row=r,column=2,value=ent)
        ws.cell(row=r,column=3,value=tip); ws.cell(row=r,column=4,value=pc)
        ws.cell(row=r,column=7,value=est)
    for r in range(r0,rN+1):
        ws.cell(row=r,column=4).number_format="0"
        ws.cell(row=r,column=5).number_format="yyyy-mm-dd"
        ws.cell(row=r,column=6,value=f'=IF(E{r}="","",E{r}-TODAY())').alignment=CENTER
        ws.cell(row=r,column=8,value=(f'=IF(G{r}="Entregado","Listo",IF(E{r}="","",'
                f'IF(F{r}<0,"VENCIDO",IF(F{r}<=3,"Vence pronto",IF(F{r}<=7,"Esta semana","Mas adelante")))))')).alignment=CENTER
        for j in range(1,9):
            c=ws.cell(row=r,column=j); c.border=BORDER
            if j in (1,2): c.alignment=LEFT
            elif j not in (1,2): c.alignment=CENTER
    dv=DataValidation(type="list",formula1='"Pendiente,En curso,Entregado"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"G{r0}:G{rN}")
    dvt=DataValidation(type="list",formula1='"Taller,Quiz,Parcial,Proyecto,Informe,Exposicion"',allow_blank=True)
    ws.add_data_validation(dvt); dvt.add(f"C{r0}:C{rN}")
    ws.conditional_formatting.add(f"H{r0}:H{rN}",CellIsRule(operator="equal",formula=['"VENCIDO"'],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add(f"H{r0}:H{rN}",CellIsRule(operator="equal",formula=['"Vence pronto"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.conditional_formatting.add(f"H{r0}:H{rN}",CellIsRule(operator="equal",formula=['"Listo"'],fill=PatternFill("solid",fgColor=VERDE)))
    # resumen
    sem=rN+2
    ws.cell(row=sem,column=1,value="Entregas esta semana (<=7 dias):").font=Font(bold=True)
    ws.cell(row=sem,column=3,value=(f'=SUMPRODUCT((E{r0}:E{rN}<>"")*(E{r0}:E{rN}>=TODAY())*'
            f'(E{r0}:E{rN}<=TODAY()+7)*(G{r0}:G{rN}<>"Entregado"))')).font=Font(bold=True,color="C62828")
    mes=rN+3
    ws.cell(row=mes,column=1,value="Entregas este mes (<=30 dias):").font=Font(bold=True)
    ws.cell(row=mes,column=3,value=(f'=SUMPRODUCT((E{r0}:E{rN}<>"")*(E{r0}:E{rN}>=TODAY())*'
            f'(E{r0}:E{rN}<=TODAY()+30)*(G{r0}:G{rN}<>"Entregado"))')).font=Font(bold=True,color="E65100")
    ven=rN+4
    ws.cell(row=ven,column=1,value="VENCIDAS sin entregar:").font=Font(bold=True,color="B71C1C")
    ws.cell(row=ven,column=3,value=(f'=SUMPRODUCT((E{r0}:E{rN}<>"")*(E{r0}:E{rN}<TODAY())*(G{r0}:G{rN}<>"Entregado"))')).font=Font(bold=True,color="B71C1C")
    return {"semana":"Entregas!C"+str(sem),"mes":"Entregas!C"+str(mes)}

def hoja_carrera(wb):
    ws=wb.create_sheet("PAPA Carrera")
    titulo(ws,"A1","PAPA - actual y PROYECTADO tras este semestre","D1",color=NAVY)
    ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=16
    filas=[
        ("PAPA actual (antes de este semestre)",4.2,"0.00"),
        ("% de avance actual",0.817,"0.0%"),
        ("Creditos totales del plan (editable)",180,"0"),
        ("Creditos ya cursados (calculado)","=ROUND(B4*B5,0)","0"),
        ("Promedio de ESTE semestre (de tus notas)","='Resumen Semestre'!C10","0.00"),
        ("Creditos de este semestre","='Resumen Semestre'!B10","0"),
        ("PAPA PROYECTADO al terminar el semestre","=IFERROR(ROUND((B3*B6+B7*B8)/(B6+B8),2),0)","0.00"),
        ("Nuevo % de avance","=IFERROR(ROUND((B6+B8)/B5,3),0)","0.0%"),
    ]
    r=3
    for lab,val,fmt in filas:
        l=ws.cell(row=r,column=1,value=lab); l.border=BORDER; l.alignment=LEFT
        v=ws.cell(row=r,column=2,value=val); v.border=BORDER; v.alignment=CENTER; v.number_format=fmt
        if lab.startswith("PAPA PROYECTADO"):
            v.font=Font(bold=True,size=12,color="1B5E20"); l.font=Font(bold=True)
            v.fill=PatternFill("solid",fgColor=VERDE); l.fill=PatternFill("solid",fgColor=CLARO)
        r+=1
    # Tabla de sensibilidad: como te quedaria el PAPA segun el promedio del semestre
    r+=1
    ws.cell(row=r,column=1,value="SI tu promedio del semestre es...").font=Font(bold=True)
    ws.cell(row=r,column=2,value="...tu PAPA quedaria").font=Font(bold=True); r+=1
    for s in [3.0,3.5,4.0,4.2,4.5,5.0]:
        ws.cell(row=r,column=1,value=s).alignment=CENTER; ws.cell(row=r,column=1).number_format="0.0"; ws.cell(row=r,column=1).border=BORDER
        v=ws.cell(row=r,column=2,value=f"=IFERROR(ROUND((B3*B6+{s}*B8)/(B6+B8),2),0)"); v.alignment=CENTER; v.number_format="0.00"; v.border=BORDER
        r+=1
    ws.cell(row=r+1,column=1,value="Todas tus materias tienen 3 creditos -> el promedio del semestre = promedio simple de las 6 notas.").font=Font(italic=True,size=9)
    ws.cell(row=r+2,column=1,value="Ajusta 'Creditos totales del plan' si conoces el numero exacto de tu pensum (UNAL Ing. Civil ~180).").font=Font(italic=True,size=9)

def hoja_prioridades(wb):
    ws=wb.create_sheet("Prioridades")
    titulo(ws,"A1","PRIORIDADES ACADEMICAS (automatico)","D1",color=NAVY)
    for j,(t,w) in enumerate([("Materia",32),("Nota",10),("Prioridad",14),("Que hacer",40)],1): head(ws,3,j,t,w)
    first=4; n=len(MATERIAS)
    for i in range(n):
        r=first+i; sr=4+i
        ws.cell(row=r,column=1,value=f"='Resumen Semestre'!A{sr}").alignment=LEFT
        ws.cell(row=r,column=2,value=f"='Resumen Semestre'!C{sr}").alignment=CENTER
        ws.cell(row=r,column=3,value=f"='Resumen Semestre'!F{sr}").alignment=CENTER
        ws.cell(row=r,column=4,value=(f'=IF(B{r}="","-",IF(B{r}<{APRUEBA},"URGENTE: bloques diarios + hablar con el profe",'
                f'IF(B{r}<3.5,"Reforzar esta semana","Mantener, repaso ligero")))')).alignment=LEFT
        for j in range(1,5): ws.cell(row=r,column=j).border=BORDER
    last=first+n-1
    ws.conditional_formatting.add(f"C{first}:C{last}",FormulaRule(formula=[f'$C{first}="ALTA"'],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add(f"C{first}:C{last}",FormulaRule(formula=[f'$C{first}="MEDIA"'],fill=PatternFill("solid",fgColor=AMAR)))

# ============================================================ FINANZAS
def hoja_ingresos(wb):
    ws=wb.create_sheet("Ingresos"); titulo(ws,"A1","INGRESOS DEL MES","E1",color="1B5E20")
    for j,(t,w) in enumerate([("Fecha",12),("Concepto",26),("Categoria",18),("Monto",16),("Recurrente",12)],1): head(ws,2,j,t,w)
    ej=[("","Mesada / apoyo familiar","Familia",0,"Si"),("","Trabajo / freelance","Trabajo",0,"No"),("","Otro","Otro",0,"No")]
    r0=3
    for i,f in enumerate(ej):
        for j,v in enumerate(f,1): ws.cell(row=r0+i,column=j,value=v)
    rN=r0+19
    for r in range(r0,rN+1):
        for j in range(1,6):
            c=ws.cell(row=r,column=j); c.border=BORDER; c.alignment=LEFT if j==2 else CENTER
            if j==4: c.number_format=COP
    tr=rN+1; ws.cell(row=tr,column=3,value="TOTAL INGRESOS").font=Font(bold=True)
    t=ws.cell(row=tr,column=4,value=f"=SUM(D{r0}:D{rN})"); t.font=Font(bold=True,size=12,color="1B5E20"); t.number_format=COP
    t.fill=PatternFill("solid",fgColor=CLARO); t.border=BORDER
    return f"Ingresos!D{tr}"

def hoja_egresos(wb):
    ws=wb.create_sheet("Egresos"); titulo(ws,"A1","EGRESOS DEL MES","E1",color="B71C1C")
    for j,(t,w) in enumerate([("Fecha",12),("Concepto",26),("Categoria",18),("Monto",16),("Necesario?",12)],1): head(ws,2,j,t,w)
    ej=[("","Transporte (TM/SITP)","Transporte",0,"Si"),("","Alimentacion","Comida",0,"Si"),("","Gimnasio + Box","Salud",0,"Si"),
        ("","Clases de ingles","Educacion",0,"Si"),("","Universidad / materiales","Educacion",0,"Si"),
        ("","Salidas con novia","Ocio",0,"No"),("","Suscripciones","Ocio",0,"No")]
    r0=3
    for i,f in enumerate(ej):
        for j,v in enumerate(f,1): ws.cell(row=r0+i,column=j,value=v)
    rN=r0+24
    for r in range(r0,rN+1):
        for j in range(1,6):
            c=ws.cell(row=r,column=j); c.border=BORDER; c.alignment=LEFT if j==2 else CENTER
            if j==4: c.number_format=COP
    tr=rN+1; ws.cell(row=tr,column=3,value="TOTAL EGRESOS").font=Font(bold=True)
    t=ws.cell(row=tr,column=4,value=f"=SUM(D{r0}:D{rN})"); t.font=Font(bold=True,size=12,color="B71C1C"); t.number_format=COP
    t.fill=PatternFill("solid",fgColor=CLARO); t.border=BORDER
    ws.cell(row=tr+1,column=3,value="Gasto evitable (No)").font=Font(italic=True,color="B71C1C")
    g=ws.cell(row=tr+1,column=4,value=f'=SUMIF(E{r0}:E{rN},"No",D{r0}:D{rN})'); g.number_format=COP
    return f"Egresos!D{tr}"

def hoja_deudas(wb):
    ws=wb.create_sheet("Deudas"); titulo(ws,"A1","DEUDAS Y PLAN DE PAGO","G1",color="E65100")
    for j,(t,w) in enumerate([("Acreedor",22),("Monto total",15),("Saldo actual",15),("Tasa % mes",11),("Cuota mensual",14),("Meses restantes",14),("Prioridad",12)],1): head(ws,2,j,t,w)
    r0=3; rN=r0+11
    for r in range(r0,rN+1):
        ws.cell(row=r,column=6,value=f'=IFERROR(IF(E{r}=0,"-",ROUNDUP(C{r}/E{r},0)),"-")')
        ws.cell(row=r,column=7,value=f'=IF(C{r}="","-",IF(D{r}>=3,"ALTA (tasa alta)",IF(C{r}>0,"Normal","Pagada")))')
        for j in range(1,8):
            c=ws.cell(row=r,column=j); c.border=BORDER; c.alignment=LEFT if j==1 else CENTER
            if j in (2,3,5): c.number_format=COP
    tr=rN+1; ws.cell(row=tr,column=1,value="DEUDA TOTAL").font=Font(bold=True)
    t=ws.cell(row=tr,column=3,value=f"=SUM(C{r0}:C{rN})"); t.font=Font(bold=True,size=12,color="E65100"); t.number_format=COP
    t.fill=PatternFill("solid",fgColor=CLARO); t.border=BORDER
    ws.cell(row=tr+2,column=1,value="Estrategia avalancha: paga primero la deuda de MAYOR tasa. Bola de nieve: la mas pequena primero (motivacion).").font=Font(italic=True,size=9)
    ws.conditional_formatting.add(f"G{r0}:G{rN}",FormulaRule(formula=[f'LEFT($G{r0},4)="ALTA"'],fill=PatternFill("solid",fgColor=ROJO)))
    return f"Deudas!C{tr}"

def hoja_presupuesto(wb,ing,egr):
    ws=wb.create_sheet("Presupuesto 50-30-20"); titulo(ws,"A1","PRESUPUESTO 50 / 30 / 20","D1",color=ACENTO)
    for col,w in [("A",30),("B",18),("C",20),("D",24)]: ws.column_dimensions[col].width=w
    ws.cell(row=3,column=1,value="Ingreso mensual total").font=Font(bold=True)
    c=ws.cell(row=3,column=2,value=f"={ing}"); c.number_format=COP
    for j,t in enumerate(["Categoria","Ideal","Sugerido","Para que"],1): head(ws,5,j,t)
    rows=[("Necesidades (50%)",0.5,"Arriendo, comida, transporte, U"),("Gustos / ocio (30%)",0.3,"Salidas, suscripciones, hobbies"),("Ahorro + deudas (20%)",0.2,"Fondo emergencia, abonos a deuda")]
    for i,(cat,pct,desc) in enumerate(rows):
        r=6+i; ws.cell(row=r,column=1,value=cat); ws.cell(row=r,column=2,value=pct).number_format="0%"
        m=ws.cell(row=r,column=3,value=f"=$B$3*B{r}"); m.number_format=COP
        ws.cell(row=r,column=4,value=desc).alignment=LEFT
        for j in range(1,5): ws.cell(row=r,column=j).border=BORDER
    ws.cell(row=10,column=1,value="Gasto real del mes").font=Font(bold=True)
    g=ws.cell(row=10,column=3,value=f"={egr}"); g.number_format=COP
    ws.cell(row=11,column=1,value="BALANCE (ingreso - gasto)").font=Font(bold=True)
    b=ws.cell(row=11,column=3,value=f"={ing}-{egr}"); b.number_format=COP; b.font=Font(bold=True,size=12)
    ws.conditional_formatting.add("C11",CellIsRule(operator="lessThan",formula=["0"],fill=PatternFill("solid",fgColor=ROJO)))
    ws.conditional_formatting.add("C11",CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=PatternFill("solid",fgColor=VERDE)))

# ============================================================ PASANTIA
def hoja_pasantia_plan(wb):
    ws=wb.create_sheet("Pasantia - Plan"); titulo(ws,"A1","PLAN DE PASANTIA - PASO A PASO","D1",color="004D40")
    for j,(t,w) in enumerate([("Paso",6),("Que hacer",60),("Estado",14),("Fecha meta",14)],1): head(ws,2,j,t,w)
    pasos=["Definir objetivo: area (estructuras, vias, hidraulica, construccion, interventoria) y ciudad.",
        "Actualizar hoja de vida a 1 pagina, enfocada en logros y el proyecto estructural.",
        "Optimizar LinkedIn: foto, titular claro, activar 'Open to work'.",
        "Armar lista de 25-30 empresas objetivo (constructoras, consultoras, interventoria).",
        "Conseguir contactos/referidos: profes, egresados, oficina de practicas de la U.",
        "Crear carta de presentacion base personalizable.",
        "Armar portafolio con proyectos academicos (incluye el de estructural).",
        "Aplicar minimo 5 vacantes por semana (registrar en 'Aplicaciones').",
        "Revisar portales: LinkedIn, elempleo, Computrabajo, Magneto365, bolsa de la U.",
        "Preparar entrevista: tecnicas + comportamentales (metodo STAR).",
        "Seguimiento a cada aplicacion a los 7 dias.",
        "Al recibir oferta: revisar convenio U, ARL y afiliacion antes de firmar."]
    r0=3
    for i,p in enumerate(pasos):
        r=r0+i; ws.cell(row=r,column=1,value=i+1); ws.cell(row=r,column=2,value=p).alignment=LEFT
        ws.cell(row=r,column=3,value="Pendiente")
        for j in range(1,5): ws.cell(row=r,column=j).border=BORDER
    rN=r0+len(pasos)-1
    dv=DataValidation(type="list",formula1='"Pendiente,En curso,Hecho"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"C{r0}:C{rN}")
    ws.conditional_formatting.add(f"C{r0}:C{rN}",CellIsRule(operator="equal",formula=['"Hecho"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"C{r0}:C{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.cell(row=rN+2,column=2,value="Progreso del plan").font=Font(bold=True)
    p=ws.cell(row=rN+2,column=3,value=f'=ROUND(COUNTIF(C{r0}:C{rN},"Hecho")/{len(pasos)},2)'); p.number_format="0%"; p.font=Font(bold=True)
    return f"'Pasantia - Plan'!C{rN+2}"

def hoja_aplicaciones(wb):
    ws=wb.create_sheet("Pasantia - Aplicaciones"); titulo(ws,"A1","RASTREADOR DE APLICACIONES","G1",color="00695C")
    for j,(t,w) in enumerate([("Empresa",24),("Cargo",22),("Fecha",12),("Estado",16),("Contacto",20),("Proximo paso",24),("Notas",24)],1): head(ws,2,j,t,w)
    r0=3; rN=r0+39
    for r in range(r0,rN+1):
        for j in range(1,8): ws.cell(row=r,column=j).border=BORDER; ws.cell(row=r,column=j).alignment=LEFT if j in (1,2,5,6,7) else CENTER
    dv=DataValidation(type="list",formula1='"Por aplicar,Aplicada,En proceso,Entrevista,Oferta,Rechazada"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"D{r0}:D{rN}")
    ws.cell(row=2,column=9,value="Aplicadas").font=Font(bold=True)
    ws.cell(row=3,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Aplicada")+COUNTIF(D{r0}:D{rN},"En proceso")+COUNTIF(D{r0}:D{rN},"Entrevista")+COUNTIF(D{r0}:D{rN},"Oferta")')
    ws.cell(row=4,column=9,value="Entrevistas").font=Font(bold=True); ws.cell(row=5,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Entrevista")')
    ws.cell(row=6,column=9,value="Ofertas").font=Font(bold=True); ws.cell(row=7,column=9,value=f'=COUNTIF(D{r0}:D{rN},"Oferta")')
    ws.column_dimensions["I"].width=12
    for val,color in {"Oferta":VERDE,"Entrevista":AZULC,"Rechazada":ROJO,"Aplicada":AMAR}.items():
        ws.conditional_formatting.add(f"D{r0}:D{rN}",CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",fgColor=color)))
    return f"'Pasantia - Aplicaciones'!I3"

# ============================================================ BECAS / CURSOS
def hoja_becas(wb):
    ws=wb.create_sheet("Becas"); titulo(ws,"A1","BECAS PARA TU MAESTRIA (revisar fechas cada ano)","F1",color="0D47A1")
    for j,(t,w) in enumerate([("Beca/Programa",24),("Pais/Region",16),("Que cubre",30),("Requisito clave",26),("Link",26),("Estado",14)],1): head(ws,2,j,t,w)
    datos=[("COLFUTURO","Colombia->mundo","Credito-beca hasta US$25k/ano (max 2 anos)","Admision + excelencia","colfuturo.org","Por investigar"),
        ("Fulbright Colombia","EE.UU.","Matricula + sostenimiento","Ingles (TOEFL), liderazgo","fulbright.edu.co","Por investigar"),
        ("Chevening","Reino Unido","Totalmente financiada","Experiencia + admision","chevening.org","Por investigar"),
        ("DAAD","Alemania","Becas maestria (EPOS)","Buen promedio + propuesta","daad.de","Por investigar"),
        ("Erasmus Mundus","Europa (varios)","Maestrias conjuntas 100%","Admision al programa","erasmus-plus.ec.europa.eu","Por investigar"),
        ("Becas de la universidad","Destino","Descuentos/becas parciales","Admision temprana","Web de cada U","Por investigar"),
        ("ICETEX","Colombia","Creditos y becas","Segun convocatoria","icetex.gov.co","Por investigar")]
    rN=fill_table(ws,3,datos,left_cols=(1,3,4,5))
    dv=DataValidation(type="list",formula1='"Por investigar,Aplicando,Aplicada,Admitido,Descartada"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"F3:F{rN}")
    ws.conditional_formatting.add(f"F3:F{rN}",CellIsRule(operator="equal",formula=['"Admitido"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"F3:F{rN}",CellIsRule(operator="equal",formula=['"Aplicando"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.cell(row=rN+2,column=1,value="Confirma SIEMPRE fechas en la web oficial; cambian cada ano.").font=Font(italic=True,size=9)

def hoja_destinos(wb):
    ws=wb.create_sheet("Destinos Maestria"); titulo(ws,"A1","MEJORES DESTINOS POR COSTO (2026, confirmar)","E1",color="1B5E20")
    for j,(t,w) in enumerate([("Pais",14),("Costo matricula aprox.",26),("Idioma",16),("A favor",26),("En contra",24)],1): head(ws,2,j,t,w)
    datos=[("Alemania","Gratis (~250-350 EUR/sem)","Aleman/Ingles","Top calidad-costo, industria fuerte","Algunos piden aleman"),
        ("Austria","~726 EUR/semestre (no-UE)","Aleman","Muy economica","Cupos y idioma"),
        ("Francia","~2.770-3.770 EUR/ano (no-UE)","Frances/Ingles","Ingenieria prestigiosa","Idioma frances util"),
        ("Italia","~900-4.000 EUR/ano (segun renta)","Italiano/Ingles","Becas por renta","Tramites lentos"),
        ("Espana","Medio","Espanol","Sin barrera de idioma","Menos becas full"),
        ("Noruega","YA cobra a no-UE desde 2023","Ingles/Noruego","Alta calidad de vida","Dejo de ser gratis"),
        ("Reino Unido","Alto","Ingles","Maestrias de 1 ano","Caro: depende de Chevening/Colfuturo"),
        ("EE.UU.","Alto","Ingles","Investigacion top","Caro: depende de Fulbright"),
        ("Canada","Medio-alto","Ingles/Frances","Buena via a residencia","Costo de vida")]
    rN=fill_table(ws,3,datos,left_cols=(2,4,5))
    ws.cell(row=rN+2,column=1,value="Si el costo manda: Alemania/Austria. Si quieres beca full: Chevening (UK), Fulbright (USA) o Erasmus Mundus.").font=Font(italic=True,size=9)
    ws.cell(row=rN+3,column=1,value="Fuentes: mastersportal.com, thecollegemonk.com, leapscholar.com (reformulado).").font=Font(italic=True,size=8,color="888888")

def hoja_cursos(wb):
    ws=wb.create_sheet("Cursos CV"); titulo(ws,"A1","CURSOS ONLINE PARA TU HOJA DE VIDA","F1",color="4A148C")
    for j,(t,w) in enumerate([("Curso",34),("Plataforma",14),("Por que sirve",30),("Prioridad",12),("Estado",14),("% avance",10)],1): head(ws,2,j,t,w)
    datos=[("Construction Management Specialization (Columbia)","Coursera","Gestion de obra, muy valorado","ALTA","Por empezar",0),
        ("BIM Fundamentals for Engineers","Coursera","BIM: el futuro del diseno","ALTA","Por empezar",0),
        ("Autodesk Revit / AutoCAD","Coursera/Udemy","Software clave en ofertas","ALTA","Por empezar",0),
        ("Engineering Project Management Specialization","Coursera","Liderar proyectos","MEDIA","Por empezar",0),
        ("Google Project Management Certificate","Coursera","Respetado por reclutadores","MEDIA","Por empezar",0),
        ("Preparacion IELTS/TOEFL","Coursera/edX","Ingles certificado = becas","ALTA","Por empezar",0),
        ("Excel + analisis de datos","Coursera","Productividad y datos","MEDIA","Por empezar",0),
        ("Python para ingenieros","Coursera/edX","Automatizar calculos","BAJA","Por empezar",0),
        ("SAP2000 / ETABS (modelado)","Udemy/YouTube","Directo a tu area","ALTA","Por empezar",0)]
    rN=fill_table(ws,3,datos,left_cols=(1,3),pct_cols=(6,))
    dv=DataValidation(type="list",formula1='"Por empezar,En curso,Terminado"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"E3:E{rN}")
    ws.conditional_formatting.add(f"E3:E{rN}",CellIsRule(operator="equal",formula=['"Terminado"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"E3:E{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.conditional_formatting.add(f"D3:D{rN}",CellIsRule(operator="equal",formula=['"ALTA"'],fill=PatternFill("solid",fgColor=ROJO)))
    ws.cell(row=rN+2,column=1,value="Tip: pide 'financial aid' en Coursera para certificados gratis.").font=Font(italic=True,size=9)
    ws.cell(row=rN+3,column=1,value="Cursos terminados:").font=Font(bold=True)
    ws.cell(row=rN+3,column=5,value=f'=COUNTIF(E3:E{rN},"Terminado")&" / "&{len(datos)}').font=Font(bold=True)
    return f"'Cursos CV'!E{rN+3}"

def hoja_ruta(wb):
    ws=wb.create_sheet("Ruta Posgrado"); titulo(ws,"A1","RUTA HACIA LA MAESTRIA (linea de tiempo)","C1",color="00695C")
    for j,(t,w) in enumerate([("Cuando",16),("Que hacer",58),("Estado",14)],1): head(ws,2,j,t,w)
    datos=[("Ahora (9o sem)","Subir el PAPA y salvar la materia critica","Pendiente"),
        ("Ahora","Empezar ingles y apuntar a IELTS/TOEFL","Pendiente"),
        ("3-6 meses","Definir pais y 5-8 universidades objetivo","Pendiente"),
        ("3-6 meses","Tomar 1-2 cursos de 'Cursos CV'","Pendiente"),
        ("6-9 meses","Conseguir cartas de recomendacion","Pendiente"),
        ("6-9 meses","Escribir carta de motivacion (statement of purpose)","Pendiente"),
        ("9-12 meses","Aplicar a becas (Colfuturo, Chevening, DAAD...)","Pendiente"),
        ("9-12 meses","Aplicar a admisiones de universidades","Pendiente"),
        ("12-15 meses","Entrevistas y resultados","Pendiente"),
        ("Tras el grado","Pasantia/experiencia que suma para la beca","Pendiente")]
    rN=fill_table(ws,3,datos,left_cols=(2,))
    dv=DataValidation(type="list",formula1='"Pendiente,En curso,Hecho"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"C3:C{rN}")
    ws.conditional_formatting.add(f"C3:C{rN}",CellIsRule(operator="equal",formula=['"Hecho"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"C3:C{rN}",CellIsRule(operator="equal",formula=['"En curso"'],fill=PatternFill("solid",fgColor=AMAR)))

def hoja_libros(wb):
    ws=wb.create_sheet("Libros"); titulo(ws,"A1","BIBLIOTECA RECOMENDADA PARA TI","F1",color="4A148C")
    for j,(t,w) in enumerate([("Titulo",32),("Autor",20),("Categoria",16),("Estado",14),("% avance",10),("Para que sirve",34)],1): head(ws,2,j,t,w)
    libros=[("Habitos Atomicos","James Clear","Habitos","Por leer",0,"Disciplina (empieza por este)"),
        ("Enfocate (Deep Work)","Cal Newport","Productividad","Por leer",0,"Concentracion para proyecto e ingles"),
        ("Los 7 habitos","Stephen Covey","Desarrollo","Por leer",0,"Priorizar y vivir con proposito"),
        ("Padre Rico, Padre Pobre","R. Kiyosaki","Finanzas","Por leer",0,"Educacion financiera"),
        ("El hombre mas rico de Babilonia","George Clason","Finanzas","Por leer",0,"Ahorro e inversion"),
        ("Como ganar amigos...","Dale Carnegie","Relaciones","Por leer",0,"Comunicacion (pareja y entrevistas)"),
        ("Mindset","Carol Dweck","Mentalidad","Por leer",0,"Mentalidad de crecimiento"),
        ("El poder del ahora","Eckhart Tolle","Bienestar","Por leer",0,"Manejar ansiedad"),
        ("Inteligencia emocional","Daniel Goleman","Bienestar","Por leer",0,"Gestionar el estres"),
        ("Despertando al gigante interior","Tony Robbins","Motivacion","Por leer",0,"Motivacion y decisiones")]
    rN=fill_table(ws,3,libros,left_cols=(1,6),pct_cols=(5,))
    dv=DataValidation(type="list",formula1='"Por leer,Leyendo,Leido"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"D3:D{rN}")
    ws.conditional_formatting.add(f"D3:D{rN}",CellIsRule(operator="equal",formula=['"Leido"'],fill=PatternFill("solid",fgColor=VERDE)))
    ws.conditional_formatting.add(f"D3:D{rN}",CellIsRule(operator="equal",formula=['"Leyendo"'],fill=PatternFill("solid",fgColor=AMAR)))
    ws.cell(row=rN+2,column=1,value="Leidos / Total").font=Font(bold=True)
    ws.cell(row=rN+2,column=4,value=f'=COUNTIF(D3:D{rN},"Leido")&" / "&{len(libros)}').font=Font(bold=True)
    return f"Libros!D{rN+2}"

# ============================================================ SALUD: GYM
def hoja_rutina_gym(wb):
    ws=wb.create_sheet("Rutina Gym")
    titulo(ws,"A1","RUTINA DE GIMNASIO - TONIFICAR (Lun y Sab)","E1",color="2E7D32")
    ws.merge_cells("A2:E2"); ws["A2"].value="Tonificar = peso moderado, 10-15 reps, baja controlada, descanso 45-60s. El Box (Mar/Jue) cubre tu cardio."
    ws["A2"].font=Font(italic=True,color="455A64"); ws["A2"].alignment=LEFT
    bloques=[
        ("DIA A - LUNES (full body, enfasis tren superior)",[
            ("Calentamiento (cardio + movilidad)","-","5-10 min","-"),
            ("Sentadilla (barra/mancuernas)","3","12","60s"),
            ("Press de banca / pecho en maquina","3","12","60s"),
            ("Remo sentado (espalda)","3","12","60s"),
            ("Press militar (hombros)","3","12","60s"),
            ("Curl de biceps","3","15","45s"),
            ("Extension de triceps","3","15","45s"),
            ("Plancha abdominal","3","40 s","30s"),
            ("Cardio final (caminadora/eliptica)","-","10 min","-")]),
        ("DIA B - SABADO (full body, enfasis tren inferior + core)",[
            ("Calentamiento (cardio + movilidad)","-","5-10 min","-"),
            ("Peso muerto rumano","3","12","60s"),
            ("Zancadas (lunges)","3","12 x pierna","60s"),
            ("Jalon al pecho (dorsales)","3","12","60s"),
            ("Aperturas / pec deck","3","15","45s"),
            ("Elevaciones laterales (hombro)","3","15","45s"),
            ("Hip thrust / gluteo","3","15","45s"),
            ("Elevacion de piernas + crunch","3","20","30s"),
            ("Cardio final","-","15 min","-")]),
    ]
    r=4
    for nombre,ejercicios in bloques:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
        c=ws.cell(row=r,column=1,value=nombre); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=GRIS); c.alignment=LEFT; r+=1
        for j,t in enumerate(["Ejercicio","Series","Reps","Descanso","Peso / notas"],1):
            hc=ws.cell(row=r,column=j,value=t); hc.font=Font(bold=True); hc.fill=PatternFill("solid",fgColor=CLARO)
            hc.alignment=CENTER; hc.border=BORDER
        r+=1
        for ej,se,re,de in ejercicios:
            ws.cell(row=r,column=1,value=ej).alignment=LEFT
            ws.cell(row=r,column=2,value=se).alignment=CENTER
            ws.cell(row=r,column=3,value=re).alignment=CENTER
            ws.cell(row=r,column=4,value=de).alignment=CENTER
            for j in range(1,6): ws.cell(row=r,column=j).border=BORDER
            r+=1
        r+=1
    for col,w in [("A",36),("B",9),("C",13),("D",11),("E",24)]: ws.column_dimensions[col].width=w
    ws.cell(row=r,column=1,value="Progresion: sube un poco el peso cada 2 semanas si completas todas las reps con buena tecnica.").font=Font(italic=True,size=9)

# ============================================================ SALUD: COMIDAS
def hoja_comidas(wb):
    ws=wb.create_sheet("Comidas Diarias")
    titulo(ws,"A1","PLAN DE COMIDAS DIARIO (enfoque tonificar)","F1",color="00695C")
    ws.merge_cells("A2:F2"); ws["A2"].value="Proteina en cada comida + carbohidrato bueno + verduras. Ajusta porciones a tu hambre. 2-2.5 L de agua al dia."
    ws["A2"].font=Font(italic=True,color="455A64"); ws["A2"].alignment=LEFT
    for j,(t,w) in enumerate([("Dia",12),("Desayuno (5:30-7)",30),("Media manana",20),("Almuerzo (12:30)",30),("Pre/Post entreno",22),("Cena",26)],1):
        head(ws,3,j,t,w)
    dias=[
        ("Lunes","Huevos + arepa + fruta + cafe","Yogur + frutos secos","Pollo + arroz + ensalada","Banano + agua","Crema de verduras + huevo"),
        ("Martes","Avena con leche + fruta","Manzana","Carne magra + papa + verduras","Batido proteina (post-box)","Atun + ensalada"),
        ("Miercoles","Huevos + pan integral + jugo","Yogur griego","Pescado + arroz + brocoli","Frutos secos","Sopa + pollo desmechado"),
        ("Jueves","Tostadas + aguacate + huevo","Fruta","Lentejas + arroz + ensalada","Batido (post-box)","Tortilla de huevo + verduras"),
        ("Viernes","Avena + banano + mani","Yogur","Pollo a la plancha + quinua","Fruta","Ensalada grande + atun"),
        ("Sabado","Huevos + arepa (post-gym)","Batido proteina","Carne + arroz + patacon","Fruta","Wrap integral de pollo"),
        ("Domingo","Desayuno libre (con medida)","Fruta","Almuerzo familiar","-","Algo ligero: sopa o ensalada"),
    ]
    fill_table(ws,4,dias,left_cols=(2,3,4,5,6))

# ============================================================ SALUD: MERCADO
def hoja_mercado(wb):
    ws=wb.create_sheet("Mercado Semanal")
    titulo(ws,"A1","LISTA DE MERCADO SEMANAL","E1",color="33691E")
    for j,(t,w) in enumerate([("Categoria",16),("Producto",26),("Cantidad",14),("Precio aprox (COP)",18),("Comprado",12)],1):
        head(ws,2,j,t,w)
    items=[
        ("Proteina","Huevos","30 und",0,""),("Proteina","Pechuga de pollo","1.5 kg",0,""),
        ("Proteina","Carne magra","1 kg",0,""),("Proteina","Pescado / atun","4 latas / 0.5 kg",0,""),
        ("Proteina","Lentejas / frijol","500 g",0,""),("Proteina","Proteina en polvo (opc.)","1 bolsa",0,""),
        ("Carbohidrato","Arroz","1 kg",0,""),("Carbohidrato","Avena","500 g",0,""),
        ("Carbohidrato","Pan integral","1 paquete",0,""),("Carbohidrato","Arepas / maiz","1 paquete",0,""),
        ("Carbohidrato","Papa / yuca / platano","2 kg",0,""),("Carbohidrato","Quinua (opc.)","250 g",0,""),
        ("Fruta/Verdura","Banano","1 racimo",0,""),("Fruta/Verdura","Manzana","6 und",0,""),
        ("Fruta/Verdura","Frutas de temporada","variado",0,""),("Fruta/Verdura","Verduras p/ensalada","variado",0,""),
        ("Fruta/Verdura","Brocoli / verduras","1 kg",0,""),("Fruta/Verdura","Aguacate","2 und",0,""),
        ("Lacteos","Leche","2 L",0,""),("Lacteos","Yogur griego","1 L",0,""),
        ("Otros","Frutos secos / mani","250 g",0,""),("Otros","Aceite de oliva","1 botella",0,""),
        ("Otros","Cafe","1 paquete",0,""),("Otros","Agua / hidratacion","-",0,""),
    ]
    rN=fill_table(ws,3,items,left_cols=(1,2))
    for r in range(3,rN+1): ws.cell(row=r,column=4).number_format=COP
    dv=DataValidation(type="list",formula1='"Si,No"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"E3:E{rN}")
    ws.conditional_formatting.add(f"E3:E{rN}",CellIsRule(operator="equal",formula=['"Si"'],fill=PatternFill("solid",fgColor=VERDE)))
    tr=rN+1; ws.cell(row=tr,column=3,value="TOTAL ESTIMADO").font=Font(bold=True)
    t=ws.cell(row=tr,column=4,value=f"=SUM(D3:D{rN})"); t.font=Font(bold=True,size=12,color="33691E"); t.number_format=COP
    t.fill=PatternFill("solid",fgColor=CLARO); t.border=BORDER
    ws.cell(row=tr+2,column=1,value="Tip: llena el precio una vez y reutiliza la lista cada semana. El total alimenta tu pestana de Egresos.").font=Font(italic=True,size=9)
    return f"'Mercado Semanal'!D{tr}"

# ============================================================ DASHBOARD
def hoja_inicio(wb,refs):
    ws=wb.create_sheet("Inicio",0)
    titulo(ws,"A1","MI SISTEMA DE VIDA - TABLERO MAESTRO","G1",size=16)
    ws.merge_cells("A2:G2"); ws["A2"].value="Todo en un archivo. Llena cada pestana y este tablero se actualiza solo."
    ws["A2"].font=Font(italic=True,color="455A64"); ws["A2"].alignment=CENTER
    secciones=[
        ("ACADEMICO","2E7D32",[
            ("PAPA del semestre",f"={refs['papa']}","0.00"),
            ("Materias en riesgo",f"={refs['riesgo']}","0"),
            ("Entregas esta semana",f"={refs['ent_sem']}","0"),
            ("Entregas este mes",f"={refs['ent_mes']}","0")]),
        ("FINANZAS","0D47A1",[
            ("Ingresos del mes",f"={refs['ing']}",COP),
            ("Egresos del mes",f"={refs['egr']}",COP),
            ("Balance del mes",f"={refs['ing']}-{refs['egr']}",COP),
            ("Deuda total",f"={refs['deu']}",COP)]),
        ("FUTURO","4A148C",[
            ("Avance plan pasantia",f"={refs['pas']}","0%"),
            ("Aplicaciones enviadas",f"={refs['apl']}","0"),
            ("Cursos terminados",f"={refs['cur']}","General"),
            ("Libros leidos",f"={refs['lib']}","General")]),
    ]
    r=4
    for nombre,color,tarjetas in secciones:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        sc=ws.cell(row=r,column=1,value=nombre); sc.font=Font(bold=True,color="FFFFFF"); sc.fill=PatternFill("solid",fgColor=color); sc.alignment=LEFT
        r+=1
        for lab,frm,fmt in tarjetas:
            l=ws.cell(row=r,column=1,value=lab); l.font=Font(bold=True,size=11); l.fill=PatternFill("solid",fgColor=CLARO); l.border=BORDER; l.alignment=LEFT
            v=ws.cell(row=r,column=2,value=frm); v.font=Font(bold=True,size=12,color=color); v.border=BORDER; v.alignment=CENTER
            if fmt not in ("General",): v.number_format=fmt
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
            r+=1
        r+=1
    ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=8
    # balance color (fila de balance esta en seccion finanzas)
    ws.cell(row=r,column=1,value="Revisa este tablero cada domingo en tu planeacion semanal.").font=Font(italic=True,color="455A64")

def main():
    wb=Workbook(); wb.remove(wb.active)
    hoja_horario(wb); hoja_habitos(wb)
    ref_res=hoja_resumen(wb); hoja_calculadora(wb); ref_ent=hoja_entregas(wb)
    hoja_carrera(wb); hoja_prioridades(wb)
    ing=hoja_ingresos(wb); egr=hoja_egresos(wb); deu=hoja_deudas(wb); hoja_presupuesto(wb,ing,egr)
    pas=hoja_pasantia_plan(wb); apl=hoja_aplicaciones(wb)
    hoja_becas(wb); hoja_destinos(wb); cur=hoja_cursos(wb); hoja_ruta(wb)
    lib=hoja_libros(wb)
    hoja_rutina_gym(wb); hoja_comidas(wb); hoja_mercado(wb)
    refs={"papa":ref_res["papa"].split("!")[0].strip("'")+"!"+ref_res["papa"].split("!")[1] if False else ref_res["papa"],
          "riesgo":ref_res["riesgo"],"ent_sem":ref_ent["semana"],"ent_mes":ref_ent["mes"],
          "ing":ing,"egr":egr,"deu":deu,"pas":pas,"apl":apl,"cur":cur,"lib":lib}
    hoja_inicio(wb,refs)
    wb.save("Mi-Sistema-de-Vida.xlsx")
    print("OK -> Mi-Sistema-de-Vida.xlsx  (22 pestanas)")

if __name__=="__main__":
    main()
